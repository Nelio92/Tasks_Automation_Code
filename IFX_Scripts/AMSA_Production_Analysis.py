"""AMSA production-data analysis (GrR + bin flips + outliers + plots).

Reference style: Tests_Data_Analysis.py

What this script does:
- Scans setup folders (e.g. Setup1_UF22_PC03, Setup2_UF32_PC02, ...).
- Groups CSV files by insertion signature (S11P/S21P/S31P/B11P/B21P).
- Per insertion and per test of selected modules:
  - merges all runs (typically 2 runs per setup, across >=2 setups),
    - computes two GrR metrics:
        * AMSA-tool style %Tolerance: 100 * (5.15 * sigma_gage / (UTL-LTL))
        * Script %StudyVariation: 100 * (sigma_gage / sigma_total)
  - counts bin flips (same die passes in >=1 run and fails in >=1 other run),
  - counts robust outliers,
  - generates CDF overlays and wafer maps,
    - writes one report sheet per insertion, grouped by module and sorted by AMSA GrR descending,
        filtered by AMSA GrR > threshold.

Notes:
- "GrR" here is an engineering metric based on within-die run spread + run-to-run mean shift
  relative to total spread. It is robust and practical for production flat-data comparisons.
- The script adds heuristic comments including known setup/probe-card issues provided by user context.
"""

from __future__ import annotations
import csv
import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any


DEFAULT_ENCODING = "latin1"
DELIMITER = ";"


# ================================
# USER PARAMETERS (edit here)
# ================================
REPO_ROOT = Path(__file__).resolve().parents[2]

# Input/output
PROD_DATA_ROOT: Path = REPO_ROOT / "PROD_Data"
OUTPUT_FOLDER: Path = REPO_ROOT / "Reports" / "flat_test_reports"

# Folder pattern for setups
SETUP_FOLDER_GLOB = "Setup*_UF*_PC*"

# Thresholds
GRR_AMSA_TOOL_THRESHOLD_PCT: float = 25.0
OUTLIER_MAD_MULTIPLIER: float = 6.0

# Optional limits
MAX_INSERTIONS: int | None = None
MAX_FILES_PER_INSERTION_PER_SETUP: int | None = None
ENCODING: str = DEFAULT_ENCODING

REPORTS_FOLDER: Path = REPO_ROOT / "PROD_Data"
RAW_SETUPS_FOLDER: Path = REPO_ROOT / "PROD_Data"
ANNOTATION_PLOTS_FOLDER: Path = REPO_ROOT / "Reports" / "amsa_report_annotation_plots"
OVERWRITE_REPORTS: bool = False
ONLY_REPORT_FILE: str | None = None # If set, only generate report for this file name (e.g. "Setup1_UF22_PC03_S11P_20240101123000.csv")


@dataclass(frozen=True)
class FlatFileMeta:
    header: list[str]
    numeric_test_cols: list[str]
    data_start_line_index: int
    meta_rows: dict[str, dict[str, str]]


@dataclass(frozen=True)
class RunFile:
    path: Path
    setup_folder: str
    setup_index: int | None
    tester: str | None
    probe_card: str | None
    insertion_sig: str
    insertion_label: str
    run_index_in_setup: int
    run_id: str


@dataclass
class LoadedRun:
    info: RunFile
    meta: FlatFileMeta
    df: Any  # pandas.DataFrame
    part_id: Any  # pandas.Series


def _excel_col_letter(col_idx_1_based: int) -> str:
    if col_idx_1_based < 1:
        raise ValueError("Column index must be >= 1")
    n = col_idx_1_based
    letters: list[str] = []
    while n:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def _safe_sheet_name(name: str) -> str:
    safe = re.sub(r"[\\/*?:\[\]]", "_", name).strip() or "Sheet"
    return safe[:31]


def _autofit_openpyxl_columns(ws, *, min_width: int = 8, max_width: int = 70, padding: int = 2) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    for col_idx, col_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
        start=1,
    ):
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            txt = str(cell.value)
            if "\n" in txt:
                txt = max(txt.splitlines(), key=len)
            max_len = max(max_len, len(txt))
        ws.column_dimensions[_excel_col_letter(col_idx)].width = max(min_width, min(max_width, max_len + padding))


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    s = str(value).strip().strip('"')
    if s == "" or s.lower() in {"nan", "na", "none"}:
        return None
    try:
        return float(s)
    except ValueError:
        try:
            return float(s.replace(",", "."))
        except ValueError:
            return None


def _module_from_test_name(test_name: str) -> str:
    if not test_name:
        return ""
    s = str(test_name).strip().upper()
    return s[:4] if len(s) >= 4 else s


def scan_flat_file_meta(
    file_path: Path,
    *,
    encoding: str = DEFAULT_ENCODING,
    delimiter: str = DELIMITER,
    needed_meta_rows: tuple[str, ...] = ("Test Name", "Low", "High", "Unit", "Cpk", "Yield"),
    max_scan_lines: int = 250,
) -> FlatFileMeta:
    needed = {x.strip() for x in needed_meta_rows}
    meta_rows: dict[str, dict[str, str]] = {}

    with file_path.open("r", encoding=encoding, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"Empty file: {file_path}")

        header = [h.strip() for h in header]
        numeric_test_cols = [h for h in header if h.strip().isdigit()]
        if not numeric_test_cols:
            raise ValueError(f"No numeric test columns in header: {file_path.name}")

        data_start_line_index = 1
        for line_idx, row in enumerate(reader, start=1):
            if line_idx >= max_scan_lines:
                break
            if not row:
                continue
            key = (row[0] or "").strip().strip('"')
            if key.isdigit():
                data_start_line_index = line_idx
                break
            if key in needed:
                row_map: dict[str, str] = {}
                for col_name, cell in zip(header, row, strict=False):
                    if col_name in numeric_test_cols:
                        row_map[col_name] = (cell or "").strip()
                meta_rows[key] = row_map

        else:
            data_start_line_index = max_scan_lines

    return FlatFileMeta(
        header=header,
        numeric_test_cols=numeric_test_cols,
        data_start_line_index=data_start_line_index,
        meta_rows=meta_rows,
    )


def _read_unit_data(file_path: Path, *, data_start_line_index: int, usecols: list[str], encoding: str = DEFAULT_ENCODING):
    import pandas as pd

    def _skip(i: int) -> bool:
        return 0 < i < data_start_line_index

    df = pd.read_csv(
        file_path,
        sep=DELIMITER,
        encoding=encoding,
        low_memory=False,
        usecols=usecols,
        skiprows=_skip,
    )
    df.columns = [c.strip() for c in df.columns]
    return df


def _test_name_from_meta(meta: FlatFileMeta, test_col: str) -> str:
    return (meta.meta_rows.get("Test Name", {}).get(test_col) or "").strip().strip('"')


def _limits_from_meta(meta: FlatFileMeta, test_col: str) -> tuple[float | None, float | None, str | None]:
    low = _to_float(meta.meta_rows.get("Low", {}).get(test_col))
    high = _to_float(meta.meta_rows.get("High", {}).get(test_col))
    unit = (meta.meta_rows.get("Unit", {}).get(test_col) or "").strip() or None
    return low, high, unit


def _parse_setup_folder_name(folder_name: str) -> tuple[int | None, str | None, str | None]:
    # Example: Setup1_UF22_PC03
    m = re.search(r"Setup(\d+)_UF(\d+)_PC(\d+)", folder_name, flags=re.IGNORECASE)
    if not m:
        return None, None, None
    return int(m.group(1)), f"UF{m.group(2)}", f"PC{m.group(3)}"


def _parse_insertion_sig(file_name: str) -> str | None:
    # Do not use \b here: '_' is a word char, so names like ..._S11P_... would fail.
    m = re.search(r"(S11P|S21P|S31P|B11P|B21P)", file_name, flags=re.IGNORECASE)
    return m.group(1).upper() if m else None


def _insertion_label(sig: str) -> str:
    s = sig.upper()
    mapping = {
        "S11P": "S1 | Hot (135°C)",
        "S21P": "S2 | Cold (-40°C)",
        "S31P": "S3 | Ambient (25°C)",
        "B11P": "B1 | Hot (135°C)",
        "B21P": "B2 | Ambient (25°C)",
    }
    return mapping.get(s, s)


def _extract_timestamp_for_sort(path: Path) -> str:
    # ..._YYYYMMDDHHMMSS_...
    m = re.search(r"_(\d{14})_", path.name)
    return m.group(1) if m else path.name


def _build_part_id(df):
    import pandas as pd

    preferred = ["LOT", "SUBLOT", "WAFER", "X", "Y", "CHIP_ID"]
    cols = [c for c in preferred if c in df.columns]

    if {"WAFER", "X", "Y"}.issubset(df.columns):
        cols = [c for c in ["LOT", "SUBLOT", "WAFER", "X", "Y", "CHIP_ID"] if c in df.columns]

    if not cols:
        return pd.Series([f"ROW_{i}" for i in range(len(df))], index=df.index)

    parts = []
    for c in cols:
        s = df[c].astype("string").fillna("").str.strip()
        if c in {"X", "Y"}:
            s = pd.to_numeric(s, errors="coerce").map(lambda x: "" if pd.isna(x) else str(int(x)))
            s = s.astype("string")
        parts.append(s)

    out = parts[0]
    for s in parts[1:]:
        out = out + "|" + s
    out = out.fillna("").astype(str)
    return out.where(out.str.len() > 0, other=[f"ROW_{i}" for i in range(len(out))])


def _mad(values) -> float:
    import numpy as np

    med = np.nanmedian(values)
    return float(np.nanmedian(np.abs(values - med)))


def _robust_sigma(values) -> float:
    import numpy as np

    m = _mad(values)
    if m > 0:
        return 1.4826 * m
    return float(np.nanstd(values, ddof=1))


_D2_D3_TABLE: dict[int, tuple[float, float]] = {
    2: (1.128, 0.853),
    3: (1.693, 0.888),
    4: (2.059, 0.880),
    5: (2.326, 0.864),
    6: (2.534, 0.848),
    7: (2.704, 0.833),
    8: (2.847, 0.820),
    9: (2.970, 0.808),
    10: (3.078, 0.797),
    11: (3.173, 0.787),
    12: (3.258, 0.778),
    13: (3.336, 0.770),
    14: (3.407, 0.763),
    15: (3.472, 0.756),
    16: (3.532, 0.750),
    17: (3.588, 0.744),
    18: (3.640, 0.739),
    19: (3.689, 0.733),
    20: (3.735, 0.729),
    21: (3.778, 0.724),
    22: (3.819, 0.720),
    23: (3.858, 0.716),
    24: (3.895, 0.712),
    25: (3.931, 0.708),
    26: (3.964, 0.705),
    27: (3.997, 0.702),
    28: (4.027, 0.699),
    29: (4.057, 0.696),
    30: (4.086, 0.693),
    31: (4.113, 0.690),
    32: (4.139, 0.687),
    33: (4.165, 0.685),
    34: (4.189, 0.682),
    35: (4.213, 0.680),
    36: (4.236, 0.678),
    37: (4.259, 0.675),
    38: (4.280, 0.673),
    39: (4.301, 0.671),
    40: (4.322, 0.669),
    41: (4.341, 0.667),
    42: (4.361, 0.665),
    43: (4.379, 0.664),
    44: (4.398, 0.662),
    45: (4.415, 0.660),
    46: (4.433, 0.658),
    47: (4.450, 0.657),
    48: (4.466, 0.655),
    49: (4.482, 0.654),
}


def _d2_star(g: int, m: int) -> float:
    """Infineon MSA guideline d2* constant using d2/d3 table formula.

    d2*(g,m) = d2(m) * ( 1/2 + 1/2 * sqrt(1 + 2*(d3(m)/d2(m))^2 / g) )
    """
    if g < 1 or m < 2:
        raise ValueError("d2* requires g>=1 and m>=2")

    if m in _D2_D3_TABLE:
        d2, d3 = _D2_D3_TABLE[m]
    else:
        # Fallback for large m: keep stable behavior.
        d2, d3 = _D2_D3_TABLE[49]

    ratio = (d3 / d2) ** 2
    return float(d2 * (0.5 + 0.5 * math.sqrt(1.0 + 2.0 * ratio / float(g))))


_AMSA_RANGE_OUTLIER_D: dict[int, float] = {
    2: 2.268,
    3: 1.574,
    4: 1.282,
    5: 1.115,
    6: 1.004,
    7: 0.924,
    8: 0.864,
    9: 0.816,
    10: 0.777,
    11: 0.744,
    12: 0.717,
    13: 0.693,
    14: 0.672,
    15: 0.653,
    16: 0.637,
    17: 0.622,
    18: 0.609,
    19: 0.596,
    20: 0.585,
    21: 0.572,
    22: 0.564,
    23: 0.556,
    24: 0.549,
    25: 0.542,
    26: 0.536,
    27: 0.530,
    28: 0.524,
    29: 0.519,
    30: 0.509,
    31: 0.508,
    32: 0.503,
    33: 0.499,
    34: 0.494,
    35: 0.490,
    36: 0.486,
    37: 0.482,
    38: 0.478,
    39: 0.474,
    40: 0.465,
    41: 0.467,
    42: 0.464,
    43: 0.461,
    44: 0.457,
    45: 0.454,
    46: 0.451,
    47: 0.448,
    48: 0.446,
    49: 0.443,
    50: 0.435,
}


def _amsa_D(n_repeats: int) -> float:
    if n_repeats <= 2:
        return _AMSA_RANGE_OUTLIER_D[2]
    if n_repeats >= 50:
        return _AMSA_RANGE_OUTLIER_D[50]
    return _AMSA_RANGE_OUTLIER_D.get(n_repeats, _AMSA_RANGE_OUTLIER_D[50])


def _detect_amsa_range_outlier_parts(long_df) -> set[str]:
    """Detect AMSA range-outlier dies per setup and return union of outlier part_ids.

    Outlier logic follows AMSA user doc intent: detect extreme per-die repeatability
    ranges in each setup before GrR computation.
    """
    import numpy as np
    import pandas as pd

    outlier_parts: set[str] = set()
    if long_df.empty:
        return outlier_parts

    for setup_name, ds in long_df.groupby("setup_folder"):
        k = int(ds["run_index_in_setup"].nunique())
        if k < 2:
            continue

        piv = ds.pivot_table(index="part_id", columns="run_index_in_setup", values="value", aggfunc="mean")
        if piv.empty:
            continue
        ranges = piv.max(axis=1, skipna=True) - piv.min(axis=1, skipna=True)
        ranges = ranges[np.isfinite(ranges.to_numpy(dtype=float))]
        ranges = ranges[ranges >= 0]
        if ranges.empty:
            continue

        rvals = np.sort(ranges.to_numpy(dtype=float))
        med = float(np.median(rvals))
        q90 = float(np.quantile(rvals, 0.90, method="linear"))
        d_const = _amsa_D(k)

        # AMSA 1.2-style robust outlier limit using median + 90% quantile.
        limit = med + d_const * max(0.0, q90 - med)
        if not np.isfinite(limit):
            continue

        bad = ranges[ranges > limit]
        if not bad.empty:
            outlier_parts.update(str(x) for x in bad.index.tolist())

    return outlier_parts


def _calc_grr_percent(long_df):
    """Practical GrR metric from long data.

    AMSA-style sigma extraction using d2* constants (MSA Guideline 2022):
    - sigma_repeat from mean ranges over repetitions within each setup
    - sigma_reprod_std from range of setup means (with repeatability correction)
    - sigma_reprod_ext from mean of per-die setup-mean ranges (dispersive)
    - sigma_reprod = max(std, ext)
    - sigma_gage = sqrt(sigma_repeat^2 + sigma_reprod^2)

    Input columns needed: part_id, setup_folder, run_index_in_setup, value
    """
    import numpy as np
    import pandas as pd

    if long_df.empty:
        return None, None, None, None, 0, 0, None

    # Structure data as part x setup x repetition means.
    tri = (
        long_df.groupby(["part_id", "setup_folder", "run_index_in_setup"], as_index=False)["value"]
        .mean()
        .dropna(subset=["value"])
    )
    if tri.empty:
        return None, None, None, None, 0, 0, None

    setups = sorted(tri["setup_folder"].astype(str).unique().tolist())
    i_setups = len(setups)
    if i_setups < 2:
        return None, None, None, None, 0, i_setups, None

    k_reps_by_setup: dict[str, int] = {
        s: int(tri.loc[tri["setup_folder"].astype(str) == s, "run_index_in_setup"].nunique()) for s in setups
    }
    if min(k_reps_by_setup.values()) < 2:
        return None, None, None, None, 0, i_setups, None

    # Keep parts present in all setups and all repetitions per setup.
    expected = sum(k_reps_by_setup.values())
    count_by_part = tri.groupby("part_id").size()
    good_parts = count_by_part[count_by_part >= expected].index
    tri = tri[tri["part_id"].isin(good_parts)].copy()
    if tri.empty:
        return None, None, None, None, 0, i_setups, None

    j_parts = int(tri["part_id"].nunique())
    if j_parts < 2:
        return None, None, None, None, j_parts, i_setups, None

    # 1) Repeatability: per setup and part, range over repetitions.
    rbar_by_setup: list[float] = []
    for s in setups:
        ds = tri[tri["setup_folder"].astype(str) == s]
        piv = ds.pivot_table(index="part_id", columns="run_index_in_setup", values="value", aggfunc="mean")
        rr = piv.max(axis=1, skipna=True) - piv.min(axis=1, skipna=True)
        rr = rr[np.isfinite(rr.to_numpy(dtype=float))]
        if rr.empty:
            continue
        rbar_by_setup.append(float(rr.mean()))
    if not rbar_by_setup:
        return None, None, None, None, j_parts, i_setups, None

    rbarbar = float(np.mean(rbar_by_setup))
    k_common = int(min(k_reps_by_setup.values()))
    d2_rep = _d2_star(g=max(1, i_setups * j_parts), m=max(2, k_common))
    sigma_repeat = float(rbarbar / d2_rep) if d2_rep > 0 else 0.0

    # 2) Standard reproducibility: range of setup means.
    setup_means = tri.groupby("setup_folder")["value"].mean()
    rxbar = float(setup_means.max() - setup_means.min()) if not setup_means.empty else 0.0
    d2_setup = _d2_star(g=1, m=max(2, i_setups))
    term_std = (rxbar / d2_setup) ** 2 - (sigma_repeat**2) / max(1, j_parts * k_common)
    sigma_reprod_std = float(math.sqrt(max(0.0, term_std)))

    # 3) Extended (dispersive) reproducibility: range over setup-means per die.
    mean_by_part_setup = tri.groupby(["part_id", "setup_folder"], as_index=False)["value"].mean()
    piv_ps = mean_by_part_setup.pivot_table(index="part_id", columns="setup_folder", values="value", aggfunc="mean")
    rr_die = piv_ps.max(axis=1, skipna=True) - piv_ps.min(axis=1, skipna=True)
    rr_die = rr_die[np.isfinite(rr_die.to_numpy(dtype=float))]
    if rr_die.empty:
        sigma_reprod_ext = 0.0
    else:
        rbar_ext = float(rr_die.mean())
        d2_ext = _d2_star(g=max(1, rr_die.shape[0]), m=max(2, i_setups))
        sigma_reprod_ext = float(rbar_ext / d2_ext) if d2_ext > 0 else 0.0

    sigma_reprod = float(max(sigma_reprod_std, sigma_reprod_ext))

    all_vals = tri["value"].to_numpy(dtype=float)
    all_vals = all_vals[np.isfinite(all_vals)]
    total_sigma = float(np.nanstd(all_vals, ddof=1)) if all_vals.size >= 2 else 0.0

    grr_sigma = float(math.sqrt(max(0.0, sigma_repeat**2 + sigma_reprod**2)))
    if not math.isfinite(total_sigma) or total_sigma <= 0:
        return None, sigma_repeat, sigma_reprod, total_sigma, j_parts, i_setups, tri

    grr_pct = 100.0 * grr_sigma / total_sigma
    return grr_pct, sigma_repeat, sigma_reprod, total_sigma, j_parts, i_setups, tri


def _grr_amsa_tool_percent(*, sigma_gage: float | None, low_limit: float | None, high_limit: float | None) -> float | None:
    if sigma_gage is None or not math.isfinite(float(sigma_gage)):
        return None
    if low_limit is None or high_limit is None:
        return None
    tol = float(high_limit) - float(low_limit)
    if not math.isfinite(tol) or tol <= 0:
        return None
    return 100.0 * (5.15 * float(sigma_gage) / tol)


def _count_bin_flips(long_df, *, low_limit: float | None, high_limit: float | None) -> tuple[int, int]:
    import numpy as np

    if low_limit is None and high_limit is None:
        return 0, 0

    low = -np.inf if low_limit is None else float(low_limit)
    high = np.inf if high_limit is None else float(high_limit)

    d = long_df.copy()
    d["pass"] = (d["value"] >= low) & (d["value"] <= high)

    flip_count = 0
    total_parts = 0
    for _, g in d.groupby("part_id"):
        if g.shape[0] < 2:
            continue
        total_parts += 1
        pmin = bool(g["pass"].min())
        pmax = bool(g["pass"].max())
        if pmin != pmax:
            flip_count += 1
    return flip_count, total_parts


def _count_outliers(long_df, *, outlier_mad_multiplier: float) -> tuple[int, int]:
    import numpy as np

    v = long_df["value"].to_numpy(dtype=float)
    v = v[np.isfinite(v)]
    if v.size == 0:
        return 0, 0

    med = float(np.median(v))
    mad = _mad(v)
    if mad <= 0:
        return 0, 0

    d = long_df.copy()
    d = d[d["value"].notna()].copy()
    d["is_outlier"] = (d["value"] - med).abs() > (outlier_mad_multiplier * mad)

    n_obs = int(d["is_outlier"].sum())
    n_parts = int(d.loc[d["is_outlier"], "part_id"].nunique())
    return n_obs, n_parts


def _build_comment(
    *,
    module: str,
    test_name: str,
    grr_pct: float,
    repeat_sigma: float | None,
    reprod_sigma: float | None,
    total_sigma: float | None,
    low_limit: float | None,
    high_limit: float | None,
    flip_count: int,
    parts_considered_for_flip: int,
    outlier_obs: int,
    outlier_parts: int,
    long_df,
) -> str:
    import numpy as np
    import pandas as pd

    notes: list[str] = []

    # Sigma contribution note.
    if repeat_sigma is not None and reprod_sigma is not None:
        if repeat_sigma > 0 and reprod_sigma > 0:
            if repeat_sigma >= 1.5 * reprod_sigma:
                notes.append("High GrR mainly from run-to-run spread on the same die (repeatability issue)")
            elif reprod_sigma >= 1.5 * repeat_sigma:
                notes.append("High GrR mainly from setup/run mean shift (reproducibility issue)")

    # Tight distribution near limits / marginal drift.
    if low_limit is not None or high_limit is not None:
        low = -np.inf if low_limit is None else float(low_limit)
        high = np.inf if high_limit is None else float(high_limit)
        vals = long_df["value"].to_numpy(dtype=float)
        vals = vals[np.isfinite(vals)]
        if vals.size:
            spec_span = (high - low) if np.isfinite(high - low) else np.nan
            if np.isfinite(spec_span) and spec_span > 0:
                sig = _robust_sigma(vals)
                if np.isfinite(sig) and sig < 0.05 * spec_span and flip_count > 0:
                    notes.append("Tight distribution vs spec span; marginal drift can create bin flips")

            # Near-limit concentration.
            if np.isfinite(low) and np.isfinite(high) and high > low:
                rel = np.minimum(np.abs(vals - low), np.abs(high - vals)) / (high - low)
                near = float(np.mean(rel < 0.05)) if rel.size else 0.0
                if near >= 0.15:
                    notes.append("Significant population near limits (guard-band sensitive)")

    # Outlier characterization.
    if outlier_obs > 0:
        notes.append(f"Outliers detected: {outlier_obs} obs on {outlier_parts} die(s)")
        if outlier_parts <= 2 and outlier_obs >= 4:
            notes.append("Outliers concentrated on few die -> potential defective DUT/contact instability")

    # Site-related signature.
    worst_site = None
    if "SITE_NUM" in long_df.columns:
        ds = long_df.dropna(subset=["SITE_NUM", "value"]).copy()
        if not ds.empty:
            ds["SITE_NUM"] = pd.to_numeric(ds["SITE_NUM"], errors="coerce")
            ds = ds.dropna(subset=["SITE_NUM"])
            if ds["SITE_NUM"].nunique() >= 2:
                med_by_site = ds.groupby("SITE_NUM")["value"].median().sort_index()
                rng = float(med_by_site.max() - med_by_site.min())
                sig = _robust_sigma(ds["value"].to_numpy(dtype=float))
                if sig > 0 and rng / sig >= 3.0:
                    worst_site = int(med_by_site.sub(med_by_site.median()).abs().idxmax())
                    notes.append(f"Site signature: site medians spread is high (worst site={worst_site})")

    # Wafer signature.
    if "WAFER" in long_df.columns:
        dw = long_df.dropna(subset=["WAFER", "value"]).copy()
        if not dw.empty and dw["WAFER"].nunique() >= 2:
            med_by_wafer = dw.groupby("WAFER")["value"].median()
            rng = float(med_by_wafer.max() - med_by_wafer.min())
            sig = _robust_sigma(dw["value"].to_numpy(dtype=float))
            if sig > 0 and rng / sig >= 3.0:
                notes.append("Wafer signature suspected (median shifts across wafers)")

    # Coordinate signatures.
    for axis in ("X", "Y"):
        if axis in long_df.columns:
            dxy = long_df.dropna(subset=[axis, "value"]).copy()
            if dxy.shape[0] >= 50:
                rho = dxy["value"].corr(pd.to_numeric(dxy[axis], errors="coerce"), method="spearman")
                if rho is not None and np.isfinite(rho) and abs(float(rho)) >= 0.30:
                    notes.append(f"Spatial signature: spearman(value,{axis})={float(rho):+.2f}")

    # Known root-cause hints from user context.
    pcs = {str(pc).upper() for pc in long_df.get("probe_card", []) if str(pc) != "nan"}
    if "PC02" in pcs and module == "DPLL":
        if worst_site == 5 or worst_site is None:
            notes.append("Known context: PC02 site 5 DIV8 resistance can impact DPLL phase noise")
    if "PC03" in pcs and module in {"DPLL", "TXLO"}:
        if worst_site == 11 or worst_site is None:
            notes.append("Known context: PC03 Xin/contact issue at site 11 can affect DPLL/TXLO")

    # Multipass not set context.
    if flip_count > 0:
        notes.append("Context: multipass not set, failing dies from previous runs may remain in population")
        notes.append(f"Bin flips observed on {flip_count}/{parts_considered_for_flip} comparable dies")

    if grr_pct >= 30:
        notes.append("Severe measurement-system variation (high GrR)")
    elif grr_pct >= 20:
        notes.append("Strong measurement-system influence")
    else:
        notes.append("Moderate measurement-system influence")

    return "; ".join(notes) if notes else "High GrR with no clear dominant signature"


def _cdf_overlay_plot_png(
    long_df,
    *,
    title: str,
    out_path: Path,
    low_limit: float | None,
    high_limit: float | None,
    outlier_mad_multiplier: float,
) -> None:
    import numpy as np

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        return

    fig, ax = plt.subplots(figsize=(7.2, 4.2), dpi=140)

    all_v = long_df["value"].to_numpy(dtype=float)
    all_v = all_v[np.isfinite(all_v)]
    med = float(np.median(all_v)) if all_v.size else float("nan")
    mad = _mad(all_v) if all_v.size else 0.0
    has_outlier_rule = bool(np.isfinite(med) and mad > 0)
    mad_outlier_label_drawn = False

    for run_id, g in sorted(long_df.groupby("run_id"), key=lambda x: str(x[0])):
        v = g["value"].to_numpy(dtype=float)
        v = v[np.isfinite(v)]
        if v.size == 0:
            continue
        v.sort()
        y = np.arange(1, v.size + 1) / v.size
        ax.plot(
            v,
            y,
            linestyle="None",
            marker=".",
            markersize=3.2,
            alpha=0.85,
            label=f"{run_id} (N={v.size})",
        )

        if has_outlier_rule:
            out = np.abs(v - med) > (outlier_mad_multiplier * mad)
            if np.any(out):
                ax.scatter(
                    v[out],
                    y[out],
                    facecolors="none",
                    edgecolors="#D62728",
                    linewidths=1.0,
                    s=20,
                    label=("Outlier" if not mad_outlier_label_drawn else None),
                )
                mad_outlier_label_drawn = True

    if low_limit is not None and math.isfinite(low_limit):
        ax.axvline(float(low_limit), color="#D62728", linestyle="--", linewidth=1.2, label=f"LTL={low_limit:.6g}")
    if high_limit is not None and math.isfinite(high_limit):
        ax.axvline(float(high_limit), color="#D62728", linestyle="--", linewidth=1.2, label=f"UTL={high_limit:.6g}")

    ax.set_title(title)
    ax.set_xlabel("Value")
    ax.set_ylabel("CDF")
    ax.grid(True, alpha=0.25)
    ax.legend(loc="best", fontsize=8, framealpha=0.9)
    fig.tight_layout()

    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, format="png")
    plt.close(fig)


def _wafer_map_runs_plot_png(long_df, *, title: str, out_path: Path, low_limit: float | None, high_limit: float | None) -> None:
    import numpy as np
    import pandas as pd

    if not {"X", "Y"}.issubset(set(long_df.columns)):
        return

    d = long_df.copy()
    d["X"] = pd.to_numeric(d["X"], errors="coerce")
    d["Y"] = pd.to_numeric(d["Y"], errors="coerce")
    d = d.dropna(subset=["X", "Y", "value"])
    if d.empty:
        return

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.colors as mcolors
        from matplotlib.patches import Circle
    except Exception:
        return

    v = d["value"].to_numpy(dtype=float)
    v = v[np.isfinite(v)]
    if v.size == 0:
        return

    vmin = float(np.nanpercentile(v, 1))
    vmax = float(np.nanpercentile(v, 99))
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmin >= vmax:
        vmin, vmax = float(np.nanmin(v)), float(np.nanmax(v))

    try:
        cmap = plt.get_cmap("turbo")
    except Exception:
        cmap = plt.get_cmap("viridis")
    norm = mcolors.Normalize(vmin=vmin, vmax=vmax, clip=True)

    runs = sorted(d["run_id"].astype(str).unique().tolist())
    n = len(runs)
    ncols = 2 if n >= 2 else 1
    nrows = int(math.ceil(n / ncols))

    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(7.8, 4.8), dpi=140)
    if not isinstance(axes, np.ndarray):
        axes = np.array([axes])
    axes = axes.ravel()

    fig.subplots_adjust(left=0.08, right=0.86, bottom=0.10, top=0.86, wspace=0.25, hspace=0.30)

    low = -np.inf if low_limit is None else float(low_limit)
    high = np.inf if high_limit is None else float(high_limit)

    mappable = None
    for ax, run_id in zip(axes, runs, strict=False):
        g = d[d["run_id"].astype(str) == run_id]
        if g.empty:
            ax.axis("off")
            continue

        xv = g["X"].to_numpy(dtype=float)
        yv = g["Y"].to_numpy(dtype=float)
        vv = g["value"].to_numpy(dtype=float)

        fails = (vv < low) | (vv > high)
        n_fail = int(np.count_nonzero(fails))

        sc = ax.scatter(xv, yv, c=vv, cmap=cmap, norm=norm, s=16, marker="s", linewidths=0, alpha=0.95)
        mappable = sc

        if np.any(fails):
            ax.scatter(
                xv[fails],
                yv[fails],
                facecolors="none",
                edgecolors="#D62728",
                linewidths=1.1,
                s=58,
                marker="s",
            )

        cx, cy = float(np.mean(xv)), float(np.mean(yv))
        r = float(np.max(np.sqrt((xv - cx) ** 2 + (yv - cy) ** 2))) + 0.8
        ax.add_patch(Circle((cx, cy), r, fill=False, color="#666666", linewidth=0.8, alpha=0.8))

        ax.set_title(f"{run_id} | N={len(vv)} | fails={n_fail}", fontsize=8)
        ax.set_aspect("equal", adjustable="box")
        ax.grid(True, alpha=0.15)
        ax.tick_params(labelsize=7)

    for ax in axes[len(runs) :]:
        ax.axis("off")

    if mappable is not None:
        cax = fig.add_axes([0.88, 0.17, 0.03, 0.62])
        cbar = fig.colorbar(mappable, cax=cax)
        cbar.ax.tick_params(labelsize=8)
        cbar.set_label("Value", fontsize=8)

    fig.suptitle(title, fontsize=10)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, format="png")
    plt.close(fig)


def _collect_run_files(prod_root: Path, setup_glob: str) -> list[RunFile]:
    setup_dirs = sorted([p for p in prod_root.glob(setup_glob) if p.is_dir()])
    if not setup_dirs:
        raise SystemExit(f"No setup folders found under {prod_root} using pattern '{setup_glob}'")

    grouped: dict[tuple[str, str], list[Path]] = {}
    setup_meta: dict[str, tuple[int | None, str | None, str | None]] = {}

    for sd in setup_dirs:
        setup_folder = sd.name
        setup_meta[setup_folder] = _parse_setup_folder_name(setup_folder)

        for csv_path in sorted(sd.glob("*.csv")):
            if not csv_path.is_file():
                continue
            ins = _parse_insertion_sig(csv_path.name)
            if not ins:
                continue
            grouped.setdefault((setup_folder, ins), []).append(csv_path)

    runs: list[RunFile] = []
    for (setup_folder, ins), files in grouped.items():
        files_sorted = sorted(files, key=_extract_timestamp_for_sort)
        if MAX_FILES_PER_INSERTION_PER_SETUP is not None:
            files_sorted = files_sorted[:MAX_FILES_PER_INSERTION_PER_SETUP]

        setup_idx, tester, pc = setup_meta.get(setup_folder, (None, None, None))
        for i, fp in enumerate(files_sorted, start=1):
            run_id = f"{tester or setup_folder}_{pc or ''}_R{i}".replace("__", "_").strip("_")
            runs.append(
                RunFile(
                    path=fp,
                    setup_folder=setup_folder,
                    setup_index=setup_idx,
                    tester=tester,
                    probe_card=pc,
                    insertion_sig=ins,
                    insertion_label=_insertion_label(ins),
                    run_index_in_setup=i,
                    run_id=run_id,
                )
            )
    return runs


def _load_run(run: RunFile, *, modules_upper: set[str], encoding: str) -> LoadedRun | None:
    import pandas as pd

    meta = scan_flat_file_meta(run.path, encoding=encoding)

    # Keep only numeric tests whose test-name module is in selection.
    interest_cols: list[str] = []
    for tc in meta.numeric_test_cols:
        tn = _test_name_from_meta(meta, tc)
        mod = _module_from_test_name(tn)
        if mod in modules_upper:
            interest_cols.append(tc)

    if not interest_cols:
        return None

    wanted_meta_cols = [
        c
        for c in ("SITE_NUM", "WAFER", "X", "Y", "LOT", "SUBLOT", "CHIP_ID", "PF", "FIRST_FAIL_TEST")
        if c in meta.header
    ]

    usecols = wanted_meta_cols + interest_cols
    df = _read_unit_data(run.path, data_start_line_index=meta.data_start_line_index, usecols=usecols, encoding=encoding)

    # Normalize known meta columns to keep downstream robust.
    for c in wanted_meta_cols:
        if c not in df.columns:
            df[c] = pd.NA

    part_id = _build_part_id(df)
    return LoadedRun(info=run, meta=meta, df=df, part_id=part_id)


def generate_amsa_report(
    *,
    prod_data_root: Path,
    output_folder: Path,
    modules: list[str],
    grr_threshold_pct: float,
    outlier_mad_multiplier: float,
    max_insertions: int | None,
    encoding: str,
) -> Path:
    import numpy as np
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font

    output_folder.mkdir(parents=True, exist_ok=True)
    plots_root = output_folder / "amsa_plots"
    plots_root.mkdir(parents=True, exist_ok=True)

    modules_upper = {m.strip().upper() for m in modules if m.strip()}
    module_rank = {m.upper(): i for i, m in enumerate([x.strip() for x in modules if x.strip()])}
    if not modules_upper:
        raise SystemExit("No modules provided")

    run_files = _collect_run_files(prod_data_root, SETUP_FOLDER_GLOB)
    if not run_files:
        raise SystemExit("No eligible run files found")

    # Group by insertion.
    by_insertion: dict[str, list[RunFile]] = {}
    for rf in run_files:
        by_insertion.setdefault(rf.insertion_sig, []).append(rf)

    insertion_keys = sorted(by_insertion.keys())
    if max_insertions is not None:
        insertion_keys = insertion_keys[:max_insertions]

    wb = Workbook()
    wb.remove(wb.active)

    for ins in insertion_keys:
        runs_for_ins = sorted(by_insertion[ins], key=lambda r: (r.setup_folder, r.run_index_in_setup, _extract_timestamp_for_sort(r.path)))
        print(f"Insertion {ins}: {len(runs_for_ins)} run file(s)")
        if runs_for_ins:
            print("  Files:")
            for rf in runs_for_ins:
                print(f"    - {rf.path.name} ({rf.setup_folder}, run {rf.run_index_in_setup})")

        loaded_runs: list[LoadedRun] = []
        for rf in runs_for_ins:
            lr = _load_run(rf, modules_upper=modules_upper, encoding=encoding)
            if lr is not None:
                loaded_runs.append(lr)

        if not loaded_runs:
            print("  - No matching module tests for this insertion")
            continue

        # Union test columns available across runs.
        all_test_cols: set[str] = set()
        for lr in loaded_runs:
            for tc in lr.meta.numeric_test_cols:
                tn = _test_name_from_meta(lr.meta, tc)
                if _module_from_test_name(tn) in modules_upper:
                    all_test_cols.add(tc)

        if not all_test_cols:
            continue

        # Prepare sheets.
        data_sheet_name = _safe_sheet_name(ins + "_AMSA")
        plot_sheet_name = _safe_sheet_name(ins + "_PLOTS")
        ws = wb.create_sheet(data_sheet_name)
        ws_plots = wb.create_sheet(plot_sheet_name)
        ws.sheet_properties.tabColor = "4F81BD"
        ws_plots.sheet_properties.tabColor = "C0504D"

        headers = [
            "Insertion",
            "Module",
            "Test Nr",
            "Test Name",
            "Unit",
            "GrR AMSA Tool (%)",
            "GrR Script (%)",
            "Repeat sigma",
            "Reprod sigma",
            "Total sigma",
            "Flip chips",
            "Flip chips (%)",
            "Outliers",
            "N parts",
            "N runs",
            "Comments",
            "CDF Plot",
            "Wafer Map",
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)

        ws_plots.append(["Test", "CDF (embedded)", "Wafer map (embedded)"])
        for c in ws_plots[1]:
            c.font = Font(bold=True)

        rows_for_sheet: list[dict[str, Any]] = []

        # Analyze each test.
        for test_col in sorted(all_test_cols, key=lambda s: int(s)):
            # Build long table for this test across runs.
            chunks = []
            test_name = ""
            module = ""
            unit = None
            low_limit = None
            high_limit = None

            for lr in loaded_runs:
                if test_col not in lr.df.columns:
                    continue

                tn = _test_name_from_meta(lr.meta, test_col)
                mod = _module_from_test_name(tn)
                if mod not in modules_upper:
                    continue

                test_name = test_name or tn
                module = module or mod
                low, high, u = _limits_from_meta(lr.meta, test_col)
                if low_limit is None and low is not None:
                    low_limit = low
                if high_limit is None and high is not None:
                    high_limit = high
                unit = unit or u

                d = pd.DataFrame(
                    {
                        "part_id": lr.part_id.astype(str),
                        "run_id": lr.info.run_id,
                        "run_index_in_setup": lr.info.run_index_in_setup,
                        "setup_folder": lr.info.setup_folder,
                        "tester": lr.info.tester,
                        "probe_card": lr.info.probe_card,
                        "SITE_NUM": lr.df["SITE_NUM"] if "SITE_NUM" in lr.df.columns else pd.NA,
                        "WAFER": lr.df["WAFER"] if "WAFER" in lr.df.columns else pd.NA,
                        "X": lr.df["X"] if "X" in lr.df.columns else pd.NA,
                        "Y": lr.df["Y"] if "Y" in lr.df.columns else pd.NA,
                        "value": pd.to_numeric(lr.df[test_col], errors="coerce"),
                    }
                )
                d = d.dropna(subset=["value"]).copy()
                if not d.empty:
                    chunks.append(d)

            if not chunks:
                continue

            long_df = pd.concat(chunks, ignore_index=True)
            if long_df.empty or long_df["run_id"].nunique() < 2:
                continue

            amsa_range_outlier_parts = _detect_amsa_range_outlier_parts(long_df)
            if amsa_range_outlier_parts:
                long_df_for_grr = long_df[~long_df["part_id"].astype(str).isin(amsa_range_outlier_parts)].copy()
            else:
                long_df_for_grr = long_df

            grr_pct, repeat_sigma, reprod_sigma, total_sigma, n_parts, n_runs, _pivot = _calc_grr_percent(long_df_for_grr)
            if grr_pct is None or not math.isfinite(grr_pct):
                continue

            sigma_gage = None
            if repeat_sigma is not None and reprod_sigma is not None:
                sigma_gage = float(math.sqrt(max(0.0, float(repeat_sigma) ** 2 + float(reprod_sigma) ** 2)))

            grr_amsa_tool_pct = _grr_amsa_tool_percent(
                sigma_gage=sigma_gage,
                low_limit=low_limit,
                high_limit=high_limit,
            )

            # Use AMSA-tool metric as the selection reference.
            if grr_amsa_tool_pct is None or not math.isfinite(grr_amsa_tool_pct):
                continue
            if grr_amsa_tool_pct <= grr_threshold_pct:
                continue

            flip_count, flip_parts = _count_bin_flips(long_df, low_limit=low_limit, high_limit=high_limit)
            flip_pct = (100.0 * flip_count / flip_parts) if flip_parts > 0 else 0.0

            outlier_obs_mad, outlier_parts_mad = _count_outliers(long_df, outlier_mad_multiplier=outlier_mad_multiplier)
            outlier_parts_range = len(amsa_range_outlier_parts)
            outlier_obs = outlier_obs_mad
            outlier_parts = outlier_parts_mad + outlier_parts_range

            comments = _build_comment(
                module=module,
                test_name=test_name,
                grr_pct=float(grr_pct),
                repeat_sigma=repeat_sigma,
                reprod_sigma=reprod_sigma,
                total_sigma=total_sigma,
                low_limit=low_limit,
                high_limit=high_limit,
                flip_count=flip_count,
                parts_considered_for_flip=flip_parts,
                outlier_obs=outlier_obs,
                outlier_parts=outlier_parts,
                long_df=long_df,
            )

            rows_for_sheet.append(
                {
                    "Insertion": _insertion_label(ins),
                    "Module": module,
                    "Test Nr": int(test_col),
                    "Test Name": test_name,
                    "Unit": unit,
                    "GrR AMSA Tool (%)": float(grr_amsa_tool_pct),
                    "GrR Script (%)": float(grr_pct),
                    "Repeat sigma": repeat_sigma,
                    "Reprod sigma": reprod_sigma,
                    "Total sigma": total_sigma,
                    "Flip chips": int(flip_count),
                    "Flip chips (%)": float(flip_pct),
                    "Outliers": f"{outlier_obs} obs / {outlier_parts} dies (range={outlier_parts_range})",
                    "N parts": n_parts,
                    "N runs": n_runs,
                    "Comments": comments,
                    "_test_col": test_col,
                    "_low": low_limit,
                    "_high": high_limit,
                    "_long_df": long_df,
                    "_range_outlier_parts": amsa_range_outlier_parts,
                }
            )

        # Group by module and sort descending AMSA metric within each module.
        rows_for_sheet.sort(
            key=lambda r: (
                module_rank.get(str(r["Module"]).upper(), 10_000),
                -float(r["GrR AMSA Tool (%)"]),
            )
        )

        if not rows_for_sheet:
            ws.append(
                [
                    _insertion_label(ins),
                    "",
                    "",
                    "No tests above AMSA GrR threshold",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            _autofit_openpyxl_columns(ws)
            _autofit_openpyxl_columns(ws_plots)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{_excel_col_letter(ws.max_column)}{ws.max_row}"
            continue

        plot_anchor_row = 3
        for item in rows_for_sheet:
            ws.append(
                [
                    item["Insertion"],
                    item["Module"],
                    item["Test Nr"],
                    item["Test Name"],
                    item["Unit"],
                    item["GrR AMSA Tool (%)"],
                    item["GrR Script (%)"],
                    item["Repeat sigma"],
                    item["Reprod sigma"],
                    item["Total sigma"],
                    item["Flip chips"],
                    item["Flip chips (%)"],
                    item["Outliers"],
                    item["N parts"],
                    item["N runs"],
                    item["Comments"],
                    "View CDF",
                    "View Wafer",
                ]
            )

            row_idx = ws.max_row
            ws.cell(row=row_idx, column=headers.index("GrR AMSA Tool (%)") + 1).number_format = "0.0"
            ws.cell(row=row_idx, column=headers.index("GrR Script (%)") + 1).number_format = "0.0"
            ws.cell(row=row_idx, column=headers.index("Flip chips (%)") + 1).number_format = "0.0"

            test_col = item["_test_col"]
            safe_test = re.sub(r"[^A-Za-z0-9._-]+", "_", str(item["Test Name"]))[:80] or str(test_col)
            cdf_path = plots_root / ins / f"{test_col}_{safe_test}_cdf.png"
            wafer_path = plots_root / ins / f"{test_col}_{safe_test}_wafer.png"

            title = (
                f"{item['Test Name']} ({test_col}) | {item['Insertion']} | "
                f"GrR_AMSA={item['GrR AMSA Tool (%)']:.1f}% | GrR_script={item['GrR Script (%)']:.1f}%"
            )
            _cdf_overlay_plot_png(
                item["_long_df"],
                title=title,
                out_path=cdf_path,
                low_limit=item["_low"],
                high_limit=item["_high"],
                outlier_mad_multiplier=outlier_mad_multiplier,
            )
            _wafer_map_runs_plot_png(
                item["_long_df"],
                title=title,
                out_path=wafer_path,
                low_limit=item["_low"],
                high_limit=item["_high"],
            )

            ws_plots[f"A{plot_anchor_row}"] = f"{test_col} {item['Test Name']}"
            ws_plots[f"A{plot_anchor_row}"].font = Font(bold=True)

            cdf_anchor_cell = f"B{plot_anchor_row + 1}"
            wafer_anchor_cell = f"K{plot_anchor_row + 1}"

            cdf_link = ws_plots[f"B{plot_anchor_row}"]
            cdf_link.value = "Open CDF PNG"
            if cdf_path.exists():
                cdf_link.hyperlink = cdf_path.resolve().as_uri()
                cdf_link.font = Font(color="0000EE", underline="single")
                img = XLImage(str(cdf_path))
                img.width = 520
                img.height = 360
                ws_plots.add_image(img, cdf_anchor_cell)

            wafer_link = ws_plots[f"C{plot_anchor_row}"]
            wafer_link.value = "Open wafer PNG"
            if wafer_path.exists():
                wafer_link.hyperlink = wafer_path.resolve().as_uri()
                wafer_link.font = Font(color="0000EE", underline="single")
                wimg = XLImage(str(wafer_path))
                wimg.width = 520
                wimg.height = 360
                ws_plots.add_image(wimg, wafer_anchor_cell)

            # Hyperlinks from report sheet to plot sheet anchors.
            cdf_target = f"#{plot_sheet_name}!{cdf_anchor_cell}"
            wafer_target = f"#{plot_sheet_name}!{wafer_anchor_cell}"

            cdf_cell = ws.cell(row=row_idx, column=headers.index("CDF Plot") + 1)
            cdf_cell.hyperlink = cdf_target
            cdf_cell.font = Font(color="0000EE", underline="single")

            wafer_cell = ws.cell(row=row_idx, column=headers.index("Wafer Map") + 1)
            wafer_cell.hyperlink = wafer_target
            wafer_cell.font = Font(color="0000EE", underline="single")

            plot_anchor_row += 30

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{_excel_col_letter(ws.max_column)}{ws.max_row}"
        _autofit_openpyxl_columns(ws)
        _autofit_openpyxl_columns(ws_plots)

    out_path = output_folder / "AMSA_GrR_Report.xlsx"
    from datetime import datetime

    try:
        wb.save(out_path)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = output_folder / f"AMSA_GrR_Report_{ts}.xlsx"
        wb.save(alt)
        print(f"Could not overwrite (open file?): {out_path}")
        print(f"Saved instead: {alt}")
        return alt

    print(f"Saved: {out_path}")
    return out_path

def _detect_insertion_key(text: str) -> str | None:
    m = re.search(r"(^|[^A-Za-z0-9])(S1|S2|S3|B1|B2)([^A-Za-z0-9]|$)", text, flags=re.IGNORECASE)
    return m.group(2).upper() if m else None


def _insertion_key_to_sig(ins_key: str) -> str | None:
    return {
        "S1": "S11P",
        "S2": "S21P",
        "S3": "S31P",
        "B1": "B11P",
        "B2": "B21P",
    }.get(ins_key.upper())


def _safe_name(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", s).strip("_") or "item"


def _find_report_header(ws) -> tuple[int, dict[str, int]]:
    required = {
        "nr",
        "testname/pin",
        "gr&r",
        "binflips",
        "outliers",
        "stucked value",
        "comment",
        "effect to quality",
        "effect to yield",
        "known issue",
        "status",
    }
    for r in range(1, min(80, ws.max_row) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        norm = {str(v).strip().lower(): i for i, v in enumerate(vals, start=1) if v is not None}
        if required.issubset(set(norm.keys())):
            return r, norm
    raise ValueError(f"Could not find AMSA report header in sheet {ws.title}")


def _iter_report_rows(ws, *, header_row: int, nr_col: int):
    empty_streak = 0
    for r in range(header_row + 1, ws.max_row + 1):
        nr = ws.cell(r, nr_col).value
        if nr is None or str(nr).strip() == "":
            empty_streak += 1
            if empty_streak >= 15:
                break
            continue
        empty_streak = 0
        yield r


def _cdf_by_site_plot(long_df, *, title: str, out_path: Path, low_limit: float | None, high_limit: float | None, highlight_sites: set[int]) -> None:
    import numpy as np
    import pandas as pd

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        return

    d = long_df.dropna(subset=["value", "SITE_NUM"]).copy()
    if d.empty:
        return
    d["SITE_NUM"] = pd.to_numeric(d["SITE_NUM"], errors="coerce")
    d = d.dropna(subset=["SITE_NUM"])
    if d.empty:
        return
    d["SITE_NUM"] = d["SITE_NUM"].astype(int)

    fig, ax = plt.subplots(figsize=(7.0, 4.2), dpi=140)
    for site, g in sorted(d.groupby("SITE_NUM"), key=lambda x: int(x[0])):
        v = g["value"].to_numpy(dtype=float)
        v = v[np.isfinite(v)]
        if v.size == 0:
            continue
        v.sort()
        y = np.arange(1, v.size + 1) / v.size
        if int(site) in highlight_sites:
            ax.plot(v, y, linestyle="None", marker="o", markersize=3.2, color="#D62728", alpha=0.9, label=f"Site {int(site)}*")
        else:
            ax.plot(v, y, linestyle="None", marker=".", markersize=2.7, color="#7F7F7F", alpha=0.55, label=f"Site {int(site)}")

    if low_limit is not None and math.isfinite(low_limit):
        ax.axvline(float(low_limit), color="#1F77B4", linestyle="--", linewidth=1.2)
    if high_limit is not None and math.isfinite(high_limit):
        ax.axvline(float(high_limit), color="#1F77B4", linestyle="--", linewidth=1.2)

    ax.set_title(title)
    ax.set_xlabel("Value")
    ax.set_ylabel("CDF")
    ax.grid(True, alpha=0.2)
    ax.legend(loc="best", fontsize=7, ncol=2)
    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, format="png")
    plt.close(fig)


def _wafer_map_plot(long_df, *, title: str, out_path: Path, highlight_sites: set[int], low_limit: float | None, high_limit: float | None) -> None:
    import numpy as np
    import pandas as pd

    if not {"X", "Y"}.issubset(long_df.columns):
        return

    d = long_df.copy()
    d["X"] = pd.to_numeric(d["X"], errors="coerce")
    d["Y"] = pd.to_numeric(d["Y"], errors="coerce")
    d = d.dropna(subset=["X", "Y", "value"])
    if d.empty:
        return

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.colors as mcolors
    except Exception:
        return

    vv = d["value"].to_numpy(dtype=float)
    vv = vv[np.isfinite(vv)]
    if vv.size == 0:
        return

    vmin = float(np.nanpercentile(vv, 1))
    vmax = float(np.nanpercentile(vv, 99))
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmin >= vmax:
        vmin, vmax = float(np.nanmin(vv)), float(np.nanmax(vv))

    fig, ax = plt.subplots(figsize=(6.4, 4.8), dpi=140)
    cmap = plt.get_cmap("turbo")
    norm = mcolors.Normalize(vmin=vmin, vmax=vmax, clip=True)

    xv = d["X"].to_numpy(dtype=float)
    yv = d["Y"].to_numpy(dtype=float)
    val = d["value"].to_numpy(dtype=float)
    sc = ax.scatter(xv, yv, c=val, cmap=cmap, norm=norm, s=20, marker="s", linewidths=0, alpha=0.95)

    if highlight_sites and "SITE_NUM" in d.columns:
        site = pd.to_numeric(d["SITE_NUM"], errors="coerce")
        sel = site.isin(list(highlight_sites)).to_numpy()
        if sel.any():
            ax.scatter(xv[sel], yv[sel], facecolors="none", edgecolors="#D62728", linewidths=1.2, s=64, marker="s", label="anomalous site")

    if low_limit is not None or high_limit is not None:
        low = -np.inf if low_limit is None else float(low_limit)
        high = np.inf if high_limit is None else float(high_limit)
        fail = (val < low) | (val > high)
        if np.any(fail):
            ax.scatter(xv[fail], yv[fail], facecolors="none", edgecolors="#111111", linewidths=1.0, s=52, marker="s", label="spec fail")

    ax.set_title(title)
    ax.set_aspect("equal", adjustable="box")
    ax.grid(True, alpha=0.15)
    ax.legend(loc="best", fontsize=8)
    cbar = fig.colorbar(sc, ax=ax)
    cbar.set_label("Value", fontsize=8)
    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, format="png")
    plt.close(fig)


def _analyze_and_comment(
    long_df,
    *,
    low_limit: float | None,
    high_limit: float | None,
    grr: float | None,
    binflips: int,
    report_outliers: int,
    stucked_value: bool,
    module: str,
) -> dict[str, Any]:
    import numpy as np
    import pandas as pd

    notes: list[str] = []
    known: list[str] = []
    near_guardband = False

    vals = pd.to_numeric(long_df["value"], errors="coerce").dropna().to_numpy(dtype=float)
    vals = vals[np.isfinite(vals)]
    if vals.size == 0:
        return {
            "comment": "No raw-data samples found for this test in selected setup files.",
            "effect_quality": "Unknown",
            "effect_yield": "Unknown",
            "known_issue": "",
            "status": "Check data mapping",
            "site_signature": False,
            "wafer_signature": False,
            "highlight_sites": set(),
        }

    med = float(np.median(vals))
    sig = _robust_sigma(vals)
    mad = _mad(vals)

    out_obs = 0
    out_parts = 0
    if mad > 0:
        d = long_df.copy()
        d["is_out"] = (pd.to_numeric(d["value"], errors="coerce") - med).abs() > (6.0 * mad)
        out_obs = int(d["is_out"].sum())
        out_parts = int(d.loc[d["is_out"], "part_id"].astype(str).nunique())
    if out_obs > 0:
        if out_parts <= 2:
            notes.append(f"Outliers concentrated on few dies ({out_obs} obs on {out_parts} die) -> potential defective DUT/contact instability")
        else:
            notes.append(f"Outliers distributed across dies ({out_obs} obs on {out_parts} dies) -> broad tails")
    elif report_outliers > 0:
        notes.append(f"AMSA report flags outliers={report_outliers}; raw aggregated tail evidence is limited")

    site_signature = False
    highlight_sites: set[int] = set()
    if "SITE_NUM" in long_df.columns:
        ds = long_df.dropna(subset=["SITE_NUM", "value"]).copy()
        if not ds.empty:
            ds["SITE_NUM"] = pd.to_numeric(ds["SITE_NUM"], errors="coerce")
            ds = ds.dropna(subset=["SITE_NUM"])
            if ds["SITE_NUM"].nunique() >= 2:
                med_by_site = ds.groupby("SITE_NUM")["value"].median().sort_index()
                rng = float(med_by_site.max() - med_by_site.min())
                if sig > 0 and rng / sig >= 3.0:
                    site_signature = True
                    for s, m in med_by_site.items():
                        if abs(float(m) - med) >= 1.5 * sig:
                            highlight_sites.add(int(s))
                    if not highlight_sites:
                        highlight_sites.add(int(med_by_site.sub(med).abs().idxmax()))
                    notes.append(f"Site signature detected: anomalous site(s)={sorted(highlight_sites)}")

    wafer_signature = False
    if "WAFER" in long_df.columns:
        dw = long_df.dropna(subset=["WAFER", "value"]).copy()
        if not dw.empty:
            dw["WAFER"] = dw["WAFER"].astype("string").str.strip()
            dw = dw[dw["WAFER"].notna() & (dw["WAFER"] != "")]
            if not dw.empty and dw["WAFER"].nunique() >= 2:
                med_by_wafer = dw.groupby("WAFER")["value"].median()
                rng = float(med_by_wafer.max() - med_by_wafer.min())
                if sig > 0 and rng / sig >= 3.0:
                    wafer_signature = True
                    notes.append("Wafer signature detected (median shifts across wafers)")

    if binflips > 0:
        if low_limit is not None or high_limit is not None:
            low = -np.inf if low_limit is None else float(low_limit)
            high = np.inf if high_limit is None else float(high_limit)
            span = (high - low) if np.isfinite(high - low) else np.nan
            if np.isfinite(span) and span > 0:
                dmin = np.minimum(np.abs(vals - low), np.abs(high - vals)) / span
                near = float(np.mean(dmin < 0.05))
                if near >= 0.15:
                    near_guardband = True
                    notes.append(f"Bin-flips likely due to marginal distribution near limits ({near:.0%} in guard-band)")
                elif out_parts > 0:
                    notes.append("Bin-flips likely linked to unstable/contact-sensitive dies")
                else:
                    notes.append("Bin-flips present; setup/condition sensitivity suspected")
            else:
                notes.append("Bin-flips present (limit interpretation constrained)")
        else:
            notes.append("Bin-flips present but no valid limits available")

    if stucked_value:
        uniq = int(len(np.unique(np.round(vals, 12))))
        if uniq <= 4:
            notes.append("Stucked Value=TRUE with low cardinality -> likely discrete/go-no-go/digital readout")
        else:
            notes.append("Stucked Value=TRUE -> possible clipping/quantization behavior")

    pcs = {str(x).upper() for x in long_df.get("probe_card", []) if x is not None}
    if module == "DPLL" and "PC02" in pcs and (5 in highlight_sites or site_signature):
        known.append("PC02 site5 DIV8-path resistance known to affect DPLL phase noise")
    if module in {"DPLL", "TXLO"} and "PC03" in pcs and (11 in highlight_sites or site_signature):
        known.append("PC03 site11 Xin/contact resistance known to affect DPLL/TXLO")

    g = float(grr) if grr is not None else 0.0
    if g >= 40 or binflips > 0:
        effect_quality = "High"
    elif g >= 25 or site_signature or wafer_signature or out_parts > 0:
        effect_quality = "Medium"
    else:
        effect_quality = "Low"

    if binflips > 0:
        effect_yield = "Potential yield impact via pass/fail instability"
    elif (low_limit is not None or high_limit is not None) and out_parts > 0:
        effect_yield = "Localized yield risk from outlier dies"
    else:
        effect_yield = "Limited direct yield impact expected"

    # Status policy (requested): ACCEPT / LIMIT / DISCUSS
    unstable_no_signature = (not site_signature and not wafer_signature) and (
        binflips > 0 or out_obs > 0 or report_outliers > 0
    )
    high_instability = (binflips >= 3) or (out_obs >= 3) or (report_outliers >= 3) or (g >= 40)

    if unstable_no_signature and high_instability:
        status = "DISCUSS"
    elif near_guardband:
        status = "LIMIT"
    else:
        status = "ACCEPT"

    # If no fundamental root-cause signature is found, use explicit no-issue comment.
    fundamental_reason_found = bool(
        known
        or site_signature
        or wafer_signature
        or out_obs > 0
        or report_outliers > 0
        or binflips > 0
        or stucked_value
    )

    if not fundamental_reason_found:
        comment_text = "Minimal variation between the setups --> no issue."
    else:
        if grr is not None:
            notes.insert(0, f"High GrR observed ({grr:.1f}%)")
        comment_text = "; ".join(notes) if notes else "Minimal variation between the setups --> no issue."

    return {
        "comment": comment_text,
        "effect_quality": effect_quality,
        "effect_yield": effect_yield,
        "known_issue": "; ".join(known),
        "status": status,
        "site_signature": site_signature,
        "wafer_signature": wafer_signature,
        "highlight_sites": highlight_sites,
    }


def _document_from_report_row(*, grr: float | None, binflips: int, outliers: int, stucked_value: bool) -> dict[str, str]:
    notes: list[str] = []

    if binflips > 0:
        notes.append(f"Bin-flips present ({binflips}); possible marginal behavior close to limits")
    if outliers > 0:
        notes.append(f"Outliers reported ({outliers}); check for unstable DUT/contact events")
    if stucked_value:
        notes.append("Stucked Value=TRUE; likely discrete/go-no-go/digital readout behavior")

    # Only mention high GrR when another fundamental reason exists.
    if notes and grr is not None:
        g = float(grr)
        notes.insert(0, f"High GrR observed ({g:.1f}%)")

    if not notes:
        comment = "Minimal variation between the setups --> no issue"
    else:
        comment = "; ".join(notes)

    # Status policy: ACCEPT / LIMIT / DISCUSS
    if (binflips >= 3 or outliers >= 3) and not stucked_value:
        status = "DISCUSS"
    elif binflips > 0:
        status = "LIMIT"
    else:
        status = "ACCEPT"

    if status == "DISCUSS":
        effect_quality = "High"
        effect_yield = "Potential yield impact via pass/fail instability"
    elif status == "LIMIT":
        effect_quality = "Medium"
        effect_yield = "Possible yield sensitivity to limits"
    else:
        effect_quality = "Low"
        effect_yield = "Limited direct yield impact expected"

    return {
        "comment": comment,
        "effect_quality": effect_quality,
        "effect_yield": effect_yield,
        "known_issue": "",
        "status": status,
    }


def annotate_existing_amsa_report(report_path: Path) -> Path:
    import openpyxl as ox

    wb = ox.load_workbook(report_path, keep_vba=True, data_only=False)

    ws = None
    header_row = None
    col = None
    for cand in wb.worksheets:
        try:
            hr, cm = _find_report_header(cand)
            ws, header_row, col = cand, hr, cm
            break
        except Exception:
            pass
    if ws is None or header_row is None or col is None:
        raise ValueError(f"Could not find main AMSA table in workbook {report_path.name}")

    print(f"Documenting report: {report_path.name}")
    updated = 0
    for r in _iter_report_rows(ws, header_row=header_row, nr_col=col["nr"]):
        grr = _to_float(ws.cell(r, col["gr&r"]).value)
        binflips = int(_to_float(ws.cell(r, col["binflips"]).value) or 0)
        outliers = int(_to_float(ws.cell(r, col["outliers"]).value) or 0)
        stucked = bool(ws.cell(r, col["stucked value"]).value)

        doc = _document_from_report_row(
            grr=grr,
            binflips=binflips,
            outliers=outliers,
            stucked_value=stucked,
        )

        ws.cell(r, col["comment"]).value = doc["comment"]
        ws.cell(r, col["effect to quality"]).value = doc["effect_quality"]
        ws.cell(r, col["effect to yield"]).value = doc["effect_yield"]
        ws.cell(r, col["known issue"]).value = doc["known_issue"]
        ws.cell(r, col["status"]).value = doc["status"]
        updated += 1

    wb.save(report_path)
    print(f"Updated rows: {updated}")
    print(f"Saved: {report_path}")
    return report_path


def run() -> int:
    reports_folder = Path(REPORTS_FOLDER)

    if ONLY_REPORT_FILE:
        report_paths = [reports_folder / ONLY_REPORT_FILE]
    else:
        report_paths = sorted([p for p in reports_folder.glob("*.xlsm") if p.is_file()])

    if not report_paths:
        raise SystemExit(f"No AMSA .xlsm reports found in {reports_folder}")

    for rp in report_paths:
        annotate_existing_amsa_report(rp)
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
