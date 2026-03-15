"""\
Chip outlier + FAIL report (flat execution, no CLI).

What it does
------------
- Reads PROD export CSVs from an input folder.
- For each requested chip (Wafer/X/Y), checks each test for:
  1) OUTLIER: |value - median| > MAD_FACTOR * MAD, where MAD = median(|x - median(x)|)
     computed from the overall distribution of that test within the same
     insertion/temperature group (derived from filename: S11P/S21P/S31P/B11P/B21P).
  2) FAIL: value outside [Low, High] limits from the file.

- Writes an Excel report listing only the tests that are OUTLIER and/or FAIL.
- Autofits all column widths.

Notes
-----
- No CLI by design; edit the CONFIG section.
- This script is intentionally self-contained (no classes).

"""

from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable
import pandas as pd


# -----------------------------------------------------------------------------
# CONFIG (edit as needed)
# -----------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]

# Input folder containing PROD raw-data CSVs.
INPUT_FOLDER = REPO_ROOT / "PROD_Data"
INPUT_GLOB = "*.csv"
MAX_FILES: int | None = None

# Output excel file
OUTPUT_XLSX = REPO_ROOT / "chip_outlier_and_fail_report.xlsx"
SHEET_NAME = "Outliers_And_Fails"

# Outlier threshold multiplier (3*MAD typical)
MAD_FACTOR: float = 10.0

# How to define the "overall distribution" for median/MAD:
# - "per_condition": compute stats per (insertion tag, insertion type, temperature) group.
# - "global": compute stats across all files combined.
DISTRIBUTION_SCOPE: str = "per_condition"  # "per_condition" | "global"

# Chips to check: either provide CHIPS directly, or point to a chips file.
# - CHIPS string format: "WAFER/X/Y;WAFER/X/Y" (also accepts comma/colon as separators)
# - Wafer normalization strips leading zeros ("02" -> "2")
CHIPS = "02/21/2;09/23/45;09/17/46;09/13/19"  # e.g. "02,31,5;02,27,8"
CHIPS_FILE: str = ""  # optional: CSV/XLSX with columns wafer/x/y


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------


@dataclass(frozen=True)
class InsertionInfo:
    tag: str  # S1/S2/S3/B1/B2/UNK
    insertion_type: str  # FE/BE/UNK
    temperature_c: int | None


def _normalize_wafer(value: object) -> str:
    s = str(value).strip().strip('"').strip("'")
    if s.isdigit():
        return str(int(s))
    return s


def _insertion_from_filename(path: Path) -> InsertionInfo:
    name_upper = path.name.upper()

    if "S11P" in name_upper:
        return InsertionInfo(tag="S1", insertion_type="FE", temperature_c=135)
    if "S21P" in name_upper:
        return InsertionInfo(tag="S2", insertion_type="FE", temperature_c=-40)
    if "S31P" in name_upper:
        return InsertionInfo(tag="S3", insertion_type="FE", temperature_c=25)
    if "B11P" in name_upper:
        return InsertionInfo(tag="B1", insertion_type="BE", temperature_c=135)
    if "B21P" in name_upper:
        return InsertionInfo(tag="B2", insertion_type="BE", temperature_c=25)

    return InsertionInfo(tag="UNK", insertion_type="UNK", temperature_c=None)


def _find_header_row_index(path: Path, needle: str = "Test Nr") -> int | None:
    try:
        with path.open("r", encoding="latin1", errors="ignore") as f:
            for i, line in enumerate(f):
                if needle in line:
                    return i
    except Exception:
        return None
    return None


def _parse_first_two_lines_csv(path: Path) -> tuple[list[str], list[str]]:
    with path.open("r", encoding="latin1", errors="ignore") as f:
        line1 = f.readline().rstrip("\n\r")
        line2 = f.readline().rstrip("\n\r")

    header = line1.split(";")
    test_names = line2.split(";")
    if len(test_names) < len(header):
        test_names = test_names + [""] * (len(header) - len(test_names))
    return header, test_names


def _find_label_row_index(raw_df: pd.DataFrame, label_col: str, label: str) -> int | None:
    s = raw_df[label_col].astype(str).str.strip().str.upper()
    try:
        return int(s[s == label.upper()].index[0])
    except Exception:
        return None


def _as_float(value: object) -> float | None:
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.lower() == "nan":
        return None
    try:
        return float(s)
    except Exception:
        return None


def _read_chips_from_file(path: Path, *, input_folder: Path) -> set[tuple[str, int, int]]:
    # Similar robustness to Tests_Data_Extractor_Flat.py
    def _resolve_existing(p: Path) -> Path | None:
        candidates: list[Path] = [p]
        if not p.is_absolute():
            candidates.append(Path.cwd() / p)
        candidates.append(input_folder / p.name)
        try:
            candidates.append(Path(__file__).resolve().parent / p.name)
        except Exception:
            pass

        for c in candidates:
            try:
                if c.is_file():
                    return c
            except Exception:
                continue

        try:
            for found in input_folder.rglob(p.name):
                if found.is_file():
                    return found
        except Exception:
            pass

        return None

    chips_file = _resolve_existing(path)
    if chips_file is None:
        raise SystemExit(f"Chips file not found: {path}")

    if chips_file.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        df = pd.read_excel(chips_file, sheet_name=0)
    else:
        # delimiter guess
        try:
            first_line = chips_file.read_text(encoding="utf-8-sig", errors="ignore").splitlines()[0]
        except Exception:
            first_line = ""
        sep = ";" if first_line.count(";") > first_line.count(",") else ","
        try:
            df = pd.read_csv(chips_file, sep=sep, engine="python")
        except Exception:
            df = pd.read_csv(chips_file, sep=None, engine="python")

    df.columns = [str(c).strip().lower() for c in df.columns]

    def _find_col(candidates: Iterable[str]) -> str | None:
        for cand in candidates:
            if cand in df.columns:
                return cand
        for col in df.columns:
            for cand in candidates:
                if cand in col:
                    return col
        return None

    wafer_col = _find_col(["wafer", "waf"])
    x_col = _find_col(["x"])
    y_col = _find_col(["y"])

    if wafer_col is None or x_col is None or y_col is None:
        if df.shape[1] >= 3:
            wafer_col, x_col, y_col = df.columns[:3]
        else:
            raise SystemExit(f"Could not parse chips file columns: {chips_file} (cols={list(df.columns)})")

    out: set[tuple[str, int, int]] = set()
    for _, r in df.iterrows():
        wafer = _normalize_wafer(r[wafer_col])
        if wafer == "" or wafer.lower() == "nan":
            continue
        try:
            x_val = int(float(r[x_col]))
            y_val = int(float(r[y_col]))
        except Exception:
            continue
        out.add((wafer, x_val, y_val))

    return out


def _parse_chips() -> set[tuple[str, int, int]]:
    chips: set[tuple[str, int, int]] = set()

    if CHIPS_FILE:
        chips |= _read_chips_from_file(Path(CHIPS_FILE), input_folder=Path(INPUT_FOLDER))

    if CHIPS.strip():
        for chunk in re.split(r"[;\n]+", CHIPS.strip()):
            chunk = chunk.strip()
            if not chunk:
                continue
            parts = [p.strip() for p in re.split(r"[/:,]", chunk) if p.strip()]
            if len(parts) != 3:
                raise SystemExit(f"Malformed chip token '{chunk}'. Expected 'WAFER,X,Y'.")
            wafer = _normalize_wafer(parts[0])
            try:
                x_val = int(float(parts[1]))
                y_val = int(float(parts[2]))
            except Exception:
                raise SystemExit(f"Malformed chip token '{chunk}'. X/Y must be numbers.")
            chips.add((wafer, x_val, y_val))

    return chips


def _read_prod_csv(
    path: Path,
) -> tuple[pd.DataFrame, dict[int, str], pd.Series, pd.Series, pd.Series, InsertionInfo]:
    """Return (chip_rows_df, test_name_map, low, high, unit, insertion_info).

    chip_rows_df columns:
      - Wafer (str)
      - X (int)
      - Y (int)
      - <test_number:int> (float)

    low/high/unit are Series indexed by test_number:int.
    """

    insertion = _insertion_from_filename(path)

    header_row = _find_header_row_index(path, needle="Test Nr")
    if header_row is None:
        raise ValueError("header_not_found")
    if header_row != 0:
        raise ValueError(f"header_not_first_line (row={header_row})")

    header_cells, test_name_cells = _parse_first_two_lines_csv(path)
    header_cells_norm = [str(v).strip() for v in header_cells]

    # Read all columns; skip rows 6-13 (1-based) => indices 5-12 (0-based)
    try:
        raw = pd.read_csv(
            path,
            header=None,
            sep=";",
            encoding="latin1",
            skiprows=list(range(5, 13)),
            engine="c",
            low_memory=False,
        )
    except ValueError:
        raw = pd.read_csv(
            path,
            header=None,
            sep=";",
            encoding="latin1",
            skiprows=list(range(5, 13)),
            engine="python",
        )

    if raw.empty or len(raw) < 6:
        raise ValueError("too_few_rows")

    # Assign columns from the true header row
    if raw.shape[1] != len(header_cells_norm):
        if raw.shape[1] > len(header_cells_norm):
            extra = [f"EXTRA_{i}" for i in range(len(header_cells_norm) + 1, raw.shape[1] + 1)]
            header_cells_norm = header_cells_norm + extra
        else:
            header_cells_norm = header_cells_norm[: raw.shape[1]]
    raw.columns = header_cells_norm

    testnr_col = next((c for c in raw.columns if str(c).strip().upper() == "TEST NR"), None)
    if testnr_col is None:
        raise ValueError("test_nr_col_not_found")

    # Base test-name mapping from line2
    test_name_map: dict[int, str] = {}
    if len(test_name_cells) < len(header_cells_norm):
        test_name_cells = test_name_cells + [""] * (len(header_cells_norm) - len(test_name_cells))
    for h, n in zip(header_cells_norm, test_name_cells, strict=False):
        hs = str(h).strip()
        if not hs.isdigit():
            continue
        test_name_map[int(hs)] = str(n).strip()

    idx_test_name = _find_label_row_index(raw, testnr_col, "TEST NAME")
    idx_low = _find_label_row_index(raw, testnr_col, "LOW")
    idx_high = _find_label_row_index(raw, testnr_col, "HIGH")
    idx_unit = _find_label_row_index(raw, testnr_col, "UNIT")

    # Update test names from TEST NAME row if present
    if idx_test_name is not None:
        for c in raw.columns:
            cs = str(c).strip()
            if not cs.isdigit():
                continue
            v = raw.at[idx_test_name, c]
            vs = str(v).strip()
            if vs.lower() != "nan" and vs:
                test_name_map[int(cs)] = vs

    def _val(idx: int | None, col: str) -> object:
        if idx is None:
            return None
        try:
            return raw.at[idx, col]
        except Exception:
            return None

    low: dict[int, float | None] = {}
    high: dict[int, float | None] = {}
    unit: dict[int, str] = {}

    for c in raw.columns:
        cs = str(c).strip()
        if not cs.isdigit():
            continue
        tn = int(cs)
        low[tn] = _as_float(_val(idx_low, c))
        high[tn] = _as_float(_val(idx_high, c))
        u = _val(idx_unit, c)
        us = "" if u is None else str(u).strip()
        unit[tn] = "" if us.lower() == "nan" else us

    if idx_unit is not None:
        data_start = idx_unit + 1
    else:
        data_start = 5

    df = raw.iloc[data_start:].copy().dropna(how="all")
    if df.empty:
        raise ValueError("no_chip_rows")

    # Identify coordinate source
    # Requirement: for BE data (B1/B2), wafer/X/Y are taken from tests 62007/62008/62009.
    if insertion.insertion_type == "BE":
        wafer_col, x_col, y_col = "62007", "62008", "62009"
        if not all(c in df.columns for c in (wafer_col, x_col, y_col)):
            raise ValueError("be_xy_tests_not_found")
    else:
        wafer_col = next((c for c in df.columns if str(c).strip().upper() == "WAFER"), None)
        x_col = next((c for c in df.columns if str(c).strip().upper() == "X"), None)
        y_col = next((c for c in df.columns if str(c).strip().upper() == "Y"), None)
        if wafer_col is None or x_col is None or y_col is None:
            raise ValueError("wafer_xy_cols_not_found")

    df[wafer_col] = df[wafer_col].astype(str).map(_normalize_wafer)
    wafer_clean = df[wafer_col].astype(str).str.strip()
    df.loc[wafer_clean.eq("") | wafer_clean.str.lower().eq("nan"), wafer_col] = pd.NA

    df[x_col] = pd.to_numeric(df[x_col], errors="coerce")
    df[y_col] = pd.to_numeric(df[y_col], errors="coerce")

    df = df.dropna(subset=[wafer_col, x_col, y_col]).copy()
    if df.empty:
        raise ValueError("no_rows_with_xy")

    df[x_col] = df[x_col].astype(float).astype(int)
    df[y_col] = df[y_col].astype(float).astype(int)

    # Test columns are digit-only headers
    test_cols = [c for c in df.columns if str(c).strip().isdigit()]
    if not test_cols:
        raise ValueError("test_cols_not_found")

    values = df[test_cols].copy()
    for c in test_cols:
        values[c] = pd.to_numeric(values[c], errors="coerce")

    out = pd.DataFrame(index=df.index)
    out["Wafer"] = df[wafer_col].astype(str)
    out["X"] = df[x_col].astype(int)
    out["Y"] = df[y_col].astype(int)

    values.columns = values.columns.astype(int)
    out = pd.concat([out, values], axis=1)

    low_s = pd.Series(low, dtype="float64")
    high_s = pd.Series(high, dtype="float64")
    unit_s = pd.Series(unit, dtype="object")

    return out, test_name_map, low_s, high_s, unit_s, insertion


def _autofit_excel_columns(xlsx_path: Path, sheet_name: str) -> None:
    # Uses openpyxl directly (installed in this repo's venv)
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        wb.save(xlsx_path)
        return

    ws = wb[sheet_name]

    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        # A small padding; cap to avoid ridiculous widths on long comments.
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    wb.save(xlsx_path)


def _autofit_excel_all_sheets(xlsx_path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                v = cell.value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    wb.save(xlsx_path)


def _safe_sheet_name(name: str, existing: set[str]) -> str:
    # Excel constraints: <=31 chars, cannot contain : \ / ? * [ ]
    base = re.sub(r"[:\\/?*\[\]]", "_", name)
    base = base.strip() or "Sheet"
    base = base[:31]

    if base not in existing:
        existing.add(base)
        return base

    # Add numeric suffix while keeping <=31
    for i in range(2, 1000):
        suffix = f"_{i}"
        candidate = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else (base + suffix)
        if candidate not in existing:
            existing.add(candidate)
            return candidate

    raise ValueError(f"Could not create unique sheet name for '{name}'")


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------


def main() -> None:
    input_folder = Path(INPUT_FOLDER)
    if not input_folder.is_dir():
        raise SystemExit(f"Input folder not found: {input_folder}")

    chips = _parse_chips()
    if not chips:
        raise SystemExit("No chips provided. Set CHIPS and/or CHIPS_FILE in the CONFIG block.")

    files = sorted(input_folder.glob(INPUT_GLOB))
    if MAX_FILES is not None:
        files = files[: int(MAX_FILES)]

    if not files:
        raise SystemExit(f"No files matched {INPUT_GLOB} in {input_folder}")

    print(f"Input folder: {input_folder}")
    print(f"Files: {len(files)}")
    print(f"Chips: {len(chips)}")
    print(f"MAD_FACTOR: {MAD_FACTOR}")

    # Collect per-condition distributions and per-chip rows.
    # condition key: (tag, insertion_type, temperature_c)
    values_by_condition: dict[tuple[str, str, int | None], list[pd.DataFrame]] = {}
    chip_rows: list[dict] = []

    for p in files:
        try:
            chip_df, test_name_map, low_s, high_s, unit_s, ins = _read_prod_csv(p)
        except Exception as e:
            print(f"Skip {p.name}: {e}")
            continue

        if str(DISTRIBUTION_SCOPE).lower() == "global":
            condition = ("GLOBAL", "GLOBAL", None)
        else:
            condition = (ins.tag, ins.insertion_type, ins.temperature_c)

        # Distribution (all chips)
        test_cols = [c for c in chip_df.columns if isinstance(c, int)]
        dist = chip_df[test_cols].copy()
        values_by_condition.setdefault(condition, []).append(dist)

        # Target chips from this file
        chip_keys = pd.Series(
            list(zip(chip_df["Wafer"].astype(str), chip_df["X"].astype(int), chip_df["Y"].astype(int))),
            index=chip_df.index,
        )
        mask = chip_keys.isin(chips)
        target = chip_df.loc[mask].copy()
        if target.empty:
            continue

        # Store each target row + file-local metadata
        for _, r in target.iterrows():
            chip_rows.append(
                {
                    "file": p.name,
                    "condition": condition,
                    "insertion": ins.tag,
                    "insertion_type": ins.insertion_type,
                    "temperature_c": ins.temperature_c,
                    "wafer": str(r["Wafer"]),
                    "x": int(r["X"]),
                    "y": int(r["Y"]),
                    "values": r[test_cols],
                    "test_name_map": test_name_map,
                    "low": low_s,
                    "high": high_s,
                    "unit": unit_s,
                }
            )

    if not values_by_condition:
        raise SystemExit("No readable PROD CSVs found.")

    if not chip_rows:
        raise SystemExit("No matching chips found in the input data.")

    # Compute robust stats per condition
    stats_by_condition: dict[tuple[str, str, int | None], dict[str, pd.Series]] = {}
    for cond, parts in values_by_condition.items():
        all_vals = pd.concat(parts, axis=0, ignore_index=True)

        med = all_vals.median(axis=0, skipna=True)
        mad = (all_vals.sub(med, axis=1).abs()).median(axis=0, skipna=True)
        n = all_vals.notna().sum(axis=0)

        stats_by_condition[cond] = {"median": med, "mad": mad, "n": n}

    # Evaluate chips
    out_rows_by_file: dict[str, list[dict]] = {}

    column_order = [
        "Wafer",
        "X",
        "Y",
        "File",
        "Insertion",
        "Insertion Type",
        "Temperature_C",
        "Test Number",
        "Test Name",
        "Value",
        "Unit",
        "Low",
        "High",
        "Status",
        "Comment",
        "Median",
        "MAD",
        "MAD Factor",
        "Threshold",
        "N (dist)",
    ]

    for entry in chip_rows:
        cond = entry["condition"]
        st = stats_by_condition.get(cond)
        if st is None:
            continue

        values: pd.Series = pd.to_numeric(entry["values"], errors="coerce")
        med = st["median"].reindex(values.index)
        mad = st["mad"].reindex(values.index)
        n = st["n"].reindex(values.index)

        abs_dev = (values - med).abs()
        thr = float(MAD_FACTOR) * mad
        outlier_mask = abs_dev > thr

        low = entry["low"].reindex(values.index)
        high = entry["high"].reindex(values.index)
        fail_low = low.notna() & values.notna() & (values < low)
        fail_high = high.notna() & values.notna() & (values > high)
        fail_mask = fail_low | fail_high

        flagged = (outlier_mask.fillna(False)) | (fail_mask.fillna(False))
        flagged_tests = [int(t) for t in flagged.index[flagged] if pd.notna(t)]

        for t in flagged_tests:
            v = values.get(t)
            t_med = med.get(t)
            t_mad = mad.get(t)
            t_thr = thr.get(t)
            t_n = n.get(t)
            t_low = low.get(t)
            t_high = high.get(t)
            t_unit = entry["unit"].get(t, "")

            is_outlier = bool(outlier_mask.get(t)) if pd.notna(outlier_mask.get(t)) else False
            is_fail_low = bool(fail_low.get(t)) if pd.notna(fail_low.get(t)) else False
            is_fail_high = bool(fail_high.get(t)) if pd.notna(fail_high.get(t)) else False
            is_fail = is_fail_low or is_fail_high

            status = "OUTLIER" if is_outlier else ""
            if is_fail:
                status = (status + "+FAIL").strip("+")

            name = entry["test_name_map"].get(t, "")
            if not name:
                name = str(t)

            comment_parts: list[str] = []
            if is_outlier:
                comment_parts.append(
                    f"OUTLIER: |x-median|={abs_dev.get(t):.6g} > {MAD_FACTOR}*MAD={t_thr:.6g} (median={t_med:.6g}, MAD={t_mad:.6g}, n={int(t_n) if pd.notna(t_n) else ''})"
                )
            if is_fail:
                if is_fail_low:
                    comment_parts.append(f"FAIL: value {v:.6g} < Low {t_low:.6g}")
                if is_fail_high:
                    comment_parts.append(f"FAIL: value {v:.6g} > High {t_high:.6g}")

            out_rows_by_file.setdefault(entry["file"], []).append(
                {
                    "Wafer": entry["wafer"],
                    "X": entry["x"],
                    "Y": entry["y"],
                    "File": entry["file"],
                    "Insertion": entry["insertion"],
                    "Insertion Type": entry["insertion_type"],
                    "Temperature_C": entry["temperature_c"],
                    "Test Number": t,
                    "Test Name": name,
                    "Value": v,
                    "Unit": t_unit,
                    "Low": t_low,
                    "High": t_high,
                    "Status": status,
                    "Comment": " | ".join(comment_parts),
                    "Median": t_med,
                    "MAD": t_mad,
                    "MAD Factor": MAD_FACTOR,
                    "Threshold": t_thr,
                    "N (dist)": t_n,
                }
            )

    # Always write a file (even if empty) so workflows are deterministic.
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    used_sheet_names: set[str] = set()
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        total_rows = 0
        for p in files:
            rows = out_rows_by_file.get(p.name, [])
            df = pd.DataFrame(rows)
            if df.empty:
                df = pd.DataFrame(columns=column_order)
            else:
                for c in column_order:
                    if c not in df.columns:
                        df[c] = pd.NA
                df = df[column_order]
                df = df.sort_values(["Wafer", "X", "Y", "Insertion", "Test Number"]).reset_index(drop=True)
                total_rows += len(df)

            sheet = _safe_sheet_name(p.stem, used_sheet_names)
            df.to_excel(writer, index=False, sheet_name=sheet)

    _autofit_excel_all_sheets(OUTPUT_XLSX)

    print(f"Wrote {OUTPUT_XLSX} with {len(files)} sheets; total flagged rows: {total_rows}")


if __name__ == "__main__":
    main()
