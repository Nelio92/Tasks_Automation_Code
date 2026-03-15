from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE = Path(r"c:\UserData\Learning\Software_Programming\GitHub_Nelio92\Test_Architect_Silicon_Architect_Learning_Plan.xlsx")

headers = [
    "Done",
    "Status",
    "Phase",
    "Timeline",
    "Category",
    "Learning Item",
    "Goal",
    "ROI",
    "Priority",
    "Resource Type",
    "Resource / Link",
    "Practice",
    "Expected Output",
    "Notes",
]

items = [
    # Phase 1
    ("☐", "Not Started", "Phase 1 - HW/SW Foundations", "Month 0-3", "Course", "UltraFLEX instrument architecture deep dive", "Build first-principles understanding of DC/RF/timing resources and limits", "Very High", "P1", "Vendor Training", "Teradyne official UltraFLEX training portal", "Map one failing test to full instrument signal chain", "One-page measurement-chain map for 5 critical tests", "Prioritize RF + mixed-signal resources"),
    ("☐", "Not Started", "Phase 1 - HW/SW Foundations", "Month 0-3", "Course", "IG-XL advanced test program structure", "Improve maintainability and debug speed in production programs", "High", "P1", "Vendor Training", "Teradyne IG-XL advanced courses", "Refactor one legacy module with reusable APIs", "Module style guide + refactored code with before/after cycle-time", "Focus on multisite-safe coding patterns"),
    ("☐", "Not Started", "Phase 1 - HW/SW Foundations", "Month 0-3", "Book", "High-Speed Digital Design (Johnson & Graham)", "Strengthen SI intuition for loadboard/probe/socket interactions", "High", "P2", "Book", "https://www.pearson.com/en-us/subject-catalog/p/high-speed-digital-design/P200000003017/9780133957242", "Review 1 chapter/week and connect to current board issues", "SI checklist applied during debug", "Focus on return paths, coupling, and timing margins"),
    ("☐", "Not Started", "Phase 1 - HW/SW Foundations", "Month 0-3", "Book", "The Art of Electronics (Horowitz & Hill)", "Build practical analog intuition for instrumentation effects", "High", "P2", "Book", "https://artofelectronics.net", "Create summary notes for relevant analog sections", "Reference notebook for analog failure mechanisms", "Use this during HW/SW interface troubleshooting"),
    ("☐", "Not Started", "Phase 1 - HW/SW Foundations", "Month 0-3", "Exposure", "Shadow senior HW debug sessions", "Learn how experts isolate hardware-vs-software root causes", "Very High", "P1", "On-the-job", "Internal mentoring / design reviews", "Attend at least 4 debug sessions and capture hypotheses/tests", "Root-cause playbook v1", "Document assumptions and evidence"),
    ("☐", "Not Started", "Phase 1 - HW/SW Foundations", "Month 0-3", "Soft Skill", "Structured problem framing (issue trees + hypothesis log)", "Reduce time-to-root-cause on complex problems", "Very High", "P1", "Practice Framework", "https://untools.co/issue-tree/", "Use issue tree template for every critical excursion", "Consistent debug log template used by team", "Keep hypotheses falsifiable"),
    ("☐", "Not Started", "Phase 1 - HW/SW Foundations", "Month 0-3", "Soft Skill", "Technical communication for cross-functional teams", "Improve clarity with design, product, and quality teams", "High", "P1", "Course", "https://www.coursera.org/learn/technical-writing", "Write weekly 1-page technical memo", "Reusable memo template + stronger review outcomes", "Include decision, risks, and next actions"),
    # Phase 2
    ("☐", "Not Started", "Phase 2 - Correlation & Data Rigor", "Month 4-6", "Course", "Design of Experiments (DOE) for semiconductor test", "Quantify factor effects and interactions on unstable tests", "Very High", "P1", "Course", "https://www.coursera.org/learn/experimental-design-basics", "Run one 2^k or fractional DOE on a noisy RF metric", "DOE report with ANOVA and action recommendations", "Tie results to Cpk/yield impact"),
    ("☐", "Not Started", "Phase 2 - Correlation & Data Rigor", "Month 4-6", "Course", "Measurement System Analysis (MSA / GR&R)", "Separate measurement noise from real silicon variation", "Very High", "P1", "Course", "https://asq.org/training/measurement-systems-analysis-msa", "Execute GR&R on 2 key analog/RF measurements", "MSA baseline + guardband recommendations", "Make this mandatory for new critical tests"),
    ("☐", "Not Started", "Phase 2 - Correlation & Data Rigor", "Month 4-6", "Course", "Applied statistics with Python (SciPy/Statsmodels)", "Improve confidence intervals and statistical decision quality", "High", "P1", "Course", "https://www.edx.org/learn/statistics", "Build reusable notebook template for correlation studies", "Versioned Python template repo", "Use bootstrap CIs where assumptions are weak"),
    ("☐", "Not Started", "Phase 2 - Correlation & Data Rigor", "Month 4-6", "Exposure", "Bench ↔ ATE ↔ simulation correlation workshop", "Standardize mismatch triage and ownership", "Very High", "P1", "On-the-job", "Internal cross-functional workshop", "Lead 1 workshop with design + validation + product", "Correlation decision tree v1 adopted by team", "Capture common mismatch patterns"),
    ("☐", "Not Started", "Phase 2 - Correlation & Data Rigor", "Month 4-6", "Exposure", "Automate correlation report generation", "Reduce manual effort and improve reproducibility", "High", "P1", "Project", "Internal Python pipeline", "Automate ingestion, filtering, plotting, and summary", "One-click weekly report", "Track run time and data quality checks"),
    ("☐", "Not Started", "Phase 2 - Correlation & Data Rigor", "Month 4-6", "Soft Skill", "Root-cause communication in escalations", "Drive faster decisions under uncertainty", "High", "P1", "Practice", "A3 / 8D problem-solving templates", "Present hypotheses, evidence, and decision gates", "Escalation communication template", "Use confidence levels with evidence"),
    ("☐", "Not Started", "Phase 2 - Correlation & Data Rigor", "Month 4-6", "Soft Skill", "Time management for high-context engineering work", "Protect deep work while maintaining responsiveness", "Medium", "P2", "Course", "https://www.coursera.org/learn/work-smarter-not-harder", "Plan weekly deep-work blocks + WIP limits", "Predictable delivery cadence", "Review weekly metrics: lead time, carryover"),
    # Phase 3
    ("☐", "Not Started", "Phase 3 - Test Architecture", "Month 7-12", "Course", "DFT fundamentals for mixed-signal SoCs", "Improve controllability/observability decisions early", "Very High", "P1", "Course", "Synopsys/Cadence/Mentor DFT learning tracks", "Map current product tests to DFT observability gaps", "DFT-informed test-content proposal", "Coordinate with design team"),
    ("☐", "Not Started", "Phase 3 - Test Architecture", "Month 7-12", "Course", "Cost-of-test and adaptive test strategy", "Optimize quality risk vs test time", "Very High", "P1", "Course", "SEMICON / IEEE test conference tutorials", "Pilot one adaptive-screening or test-time optimization", "Business case with quality-risk analysis", "Quantify ppm impact"),
    ("☐", "Not Started", "Phase 3 - Test Architecture", "Month 7-12", "Book", "Statistical Methods for Quality Improvement", "Strengthen capability analysis and control plans", "High", "P2", "Book", "https://onlinelibrary.wiley.com/doi/book/10.1002/9781119184768", "Apply SPC charts to key RF metrics", "Control plan with trigger thresholds", "Integrate with production dashboards"),
    ("☐", "Not Started", "Phase 3 - Test Architecture", "Month 7-12", "Certification", "ASQ Six Sigma Green Belt (CSSGB)", "Formalize DOE/SPC/quality toolkit", "High", "P1", "Certification", "https://asq.org/cert/six-sigma-green-belt", "Complete certification and apply methods to one live project", "Certified credential + project case study", "Target exam by month 10-12"),
    ("☐", "Not Started", "Phase 3 - Test Architecture", "Month 7-12", "Exposure", "Lead cross-functional test strategy review", "Develop architect-level decision ownership", "Very High", "P1", "On-the-job", "Internal architecture review boards", "Run quarterly review with design/validation/product", "Test strategy review deck + decisions log", "Track closure of action items"),
    ("☐", "Not Started", "Phase 3 - Test Architecture", "Month 7-12", "Soft Skill", "Influence without authority", "Improve adoption of standards and decisions", "High", "P1", "Course", "https://www.coursera.org/learn/leading-teams-developing-yourself", "Use stakeholder map + pre-alignment before reviews", "Higher decision throughput", "Measure acceptance rate of proposals"),
    # Phase 4
    ("☐", "Not Started", "Phase 4 - Architect Transition", "Month 13-18", "Course", "SoC architecture fundamentals (PPA/testability tradeoffs)", "Build silicon-architect perspective while staying test-focused", "High", "P2", "Course", "https://www.coursera.org/specializations/fpga-design", "Analyze one subsystem with PPA + testability tradeoff matrix", "Architecture note with alternatives", "Bridge to silicon architect path"),
    ("☐", "Not Started", "Phase 4 - Architect Transition", "Month 13-18", "Book", "Computer Architecture: A Quantitative Approach", "Strengthen quantitative architecture decision making", "Medium", "P3", "Book", "https://www.elsevier.com/books/computer-architecture-a-quantitative-approach/hennessy/978-0-12-811905-1", "Read selected chapters and summarize decision frameworks", "Tradeoff template for architecture reviews", "Focus on metrics-driven choices"),
    ("☐", "Not Started", "Phase 4 - Architect Transition", "Month 13-18", "Certification", "ASQ Certified Quality Engineer (CQE)", "Expand reliability and quality systems depth", "Medium", "P2", "Certification", "https://asq.org/cert/quality-engineer", "Prepare using real plant/test examples", "CQE credential + improved reliability language", "Optional if role demands broader quality ownership"),
    ("☐", "Not Started", "Phase 4 - Architect Transition", "Month 13-18", "Exposure", "Own end-to-end test architecture for one product slice", "Demonstrate readiness for Test Architect role", "Very High", "P1", "On-the-job", "Internal project charter", "Define test content, limits, guardband, and sign-off criteria", "Test Architecture Playbook v1", "Target measurable impact on test-time + quality"),
    ("☐", "Not Started", "Phase 4 - Architect Transition", "Month 13-18", "Exposure", "Mentor 1-2 engineers on debug/data workflows", "Scale impact through team capability", "High", "P2", "On-the-job", "Internal mentoring", "Biweekly mentoring sessions with practical assignments", "Shared standards and reduced onboarding time", "Collect mentee feedback"),
    ("☐", "Not Started", "Phase 4 - Architect Transition", "Month 13-18", "Soft Skill", "Executive-level storytelling", "Communicate risk/reward clearly to management", "High", "P1", "Course", "https://www.coursera.org/learn/presentation-skills", "Present monthly architecture update in 10 slides", "Leadership-ready narrative + faster approvals", "Use single-threaded story: objective -> evidence -> decision"),
    ("☐", "Not Started", "Phase 4 - Architect Transition", "Month 13-18", "Soft Skill", "Personal operating system (planning + retrospectives)", "Increase consistency and avoid context-switch loss", "High", "P1", "Practice", "Weekly sprint + retrospective templates", "Run weekly planning and monthly skill matrix review", "Sustained productivity and predictable execution", "Track learning velocity and completion rate"),
]

PHASES = [
    "Phase 1 - HW/SW Foundations",
    "Phase 2 - Correlation & Data Rigor",
    "Phase 3 - Test Architecture",
    "Phase 4 - Architect Transition",
]

CATEGORIES = ["Course", "Certification", "Book", "Exposure", "Soft Skill"]

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

category_colors = {
    "Course": "D9E1F2",
    "Certification": "FCE4D6",
    "Book": "E2F0D9",
    "Exposure": "FFF2CC",
    "Soft Skill": "EDE2F7",
}


def style_header_row(ws, start_col: int, end_col: int, row: int = 1):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border


def create_learning_plan_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Learning Plan"
    ws.append(headers)
    for item in items:
        ws.append(item)

    style_header_row(ws, 1, len(headers), 1)

    widths = {
        1: 8, 2: 14, 3: 30, 4: 13, 5: 14, 6: 42, 7: 46,
        8: 11, 9: 10, 10: 15, 11: 48, 12: 44, 13: 42, 14: 30,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    for r in range(2, ws.max_row + 1):
        category = ws.cell(row=r, column=5).value
        row_fill = PatternFill("solid", fgColor=category_colors.get(category, "FFFFFF"))
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = left_wrap if c not in [1, 2, 8, 9, 10] else center

        resource_cell = ws.cell(row=r, column=11)
        value = str(resource_cell.value or "")
        if value.startswith("http://") or value.startswith("https://"):
            resource_cell.hyperlink = value
            resource_cell.style = "Hyperlink"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{ws.max_row}"

    status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Done"', allow_blank=False)
    checkbox_dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=False)
    priority_dv = DataValidation(type="list", formula1='"P1,P2,P3"', allow_blank=False)
    roi_dv = DataValidation(type="list", formula1='"Very High,High,Medium,Low"', allow_blank=False)

    for dv in [status_dv, checkbox_dv, priority_dv, roi_dv]:
        ws.add_data_validation(dv)

    status_dv.add(f"B2:B{ws.max_row}")
    checkbox_dv.add(f"A2:A{ws.max_row}")
    priority_dv.add(f"I2:I{ws.max_row}")
    roi_dv.add(f"H2:H{ws.max_row}")

    done_rule = FormulaRule(formula=['$B2="Done"'], fill=PatternFill("solid", fgColor="C6EFCE"))
    ws.conditional_formatting.add(f"A2:N{ws.max_row}", done_rule)
    return ws


def create_summary_sheet(wb: Workbook):
    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    style_header_row(summary, 1, 2, 1)

    summary_rows = [
        ("Plan generated on", date.today().isoformat()),
        ("Target Role", "Test Architect (Primary) + Silicon Architect skills (Secondary)"),
        ("Total items", len(items)),
        ("P1 items", '=COUNTIF(\'Learning Plan\'!I:I,"P1")'),
        ("P2 items", '=COUNTIF(\'Learning Plan\'!I:I,"P2")'),
        ("P3 items", '=COUNTIF(\'Learning Plan\'!I:I,"P3")'),
        ("Completed items", '=COUNTIF(\'Learning Plan\'!B:B,"Done")'),
        ("Completion %", '=IF(COUNTA(\'Learning Plan\'!F:F)=0,0,COUNTIF(\'Learning Plan\'!B:B,"Done")/(COUNTA(\'Learning Plan\'!F:F)-1))'),
    ]

    for row in summary_rows:
        summary.append(row)

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 72
    summary["B9"].number_format = "0.00%"
    summary.freeze_panes = "A2"

    for r in range(2, summary.max_row + 1):
        for c in range(1, 3):
            cell = summary.cell(row=r, column=c)
            cell.border = border
            cell.alignment = left_wrap

    start_phase = 12
    summary[f"A{start_phase}"] = "Phase"
    summary[f"B{start_phase}"] = "Total"
    summary[f"C{start_phase}"] = "Done"
    summary[f"D{start_phase}"] = "Completion %"
    style_header_row(summary, 1, 4, start_phase)
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 14

    for idx, phase in enumerate(PHASES, start=start_phase + 1):
        summary[f"A{idx}"] = phase
        summary[f"B{idx}"] = f'=COUNTIF(\'Learning Plan\'!C:C,A{idx})'
        summary[f"C{idx}"] = f'=COUNTIFS(\'Learning Plan\'!C:C,A{idx},\'Learning Plan\'!B:B,"Done")'
        summary[f"D{idx}"] = f'=IF(B{idx}=0,0,C{idx}/B{idx})'
        summary[f"D{idx}"].number_format = "0.00%"
        for col in range(1, 5):
            cell = summary.cell(row=idx, column=col)
            cell.border = border
            cell.alignment = center if col > 1 else left_wrap

    start_cat = 12
    summary[f"F{start_cat}"] = "Category"
    summary[f"G{start_cat}"] = "Total"
    summary[f"H{start_cat}"] = "Done"
    summary[f"I{start_cat}"] = "Completion %"
    style_header_row(summary, 6, 9, start_cat)
    for col_letter, width in [("F", 20), ("G", 10), ("H", 10), ("I", 14)]:
        summary.column_dimensions[col_letter].width = width

    for idx, category in enumerate(CATEGORIES, start=start_cat + 1):
        summary[f"F{idx}"] = category
        summary[f"G{idx}"] = f'=COUNTIF(\'Learning Plan\'!E:E,F{idx})'
        summary[f"H{idx}"] = f'=COUNTIFS(\'Learning Plan\'!E:E,F{idx},\'Learning Plan\'!B:B,"Done")'
        summary[f"I{idx}"] = f'=IF(G{idx}=0,0,H{idx}/G{idx})'
        summary[f"I{idx}"].number_format = "0.00%"
        for col in range(6, 10):
            cell = summary.cell(row=idx, column=col)
            cell.border = border
            cell.alignment = center if col > 6 else left_wrap
        summary[f"F{idx}"].fill = PatternFill("solid", fgColor=category_colors.get(category, "FFFFFF"))

    phase_chart = BarChart()
    phase_chart.title = "Progress by Phase"
    phase_chart.y_axis.title = "Items"
    phase_chart.x_axis.title = "Phase"
    phase_chart.height = 7
    phase_chart.width = 11
    phase_data = Reference(summary, min_col=2, max_col=3, min_row=start_phase, max_row=start_phase + len(PHASES))
    phase_cats = Reference(summary, min_col=1, min_row=start_phase + 1, max_row=start_phase + len(PHASES))
    phase_chart.add_data(phase_data, titles_from_data=True)
    phase_chart.set_categories(phase_cats)
    phase_chart.style = 10
    summary.add_chart(phase_chart, "A20")

    category_chart = BarChart()
    category_chart.title = "Progress by Category"
    category_chart.y_axis.title = "Items"
    category_chart.x_axis.title = "Category"
    category_chart.height = 7
    category_chart.width = 11
    cat_data = Reference(summary, min_col=7, max_col=8, min_row=start_cat, max_row=start_cat + len(CATEGORIES))
    cat_cats = Reference(summary, min_col=6, min_row=start_cat + 1, max_row=start_cat + len(CATEGORIES))
    category_chart.add_data(cat_data, titles_from_data=True)
    category_chart.set_categories(cat_cats)
    category_chart.style = 11
    summary.add_chart(category_chart, "F20")


def create_weekly_schedule_sheet(wb: Workbook):
    ws = wb.create_sheet("Weekly Schedule")
    weekly_headers = [
        "Week",
        "Phase",
        "Focus Area",
        "Primary Learning Goal",
        "Planned Hours",
        "Actual Hours",
        "Planned Deliverable",
        "Status",
        "Notes",
    ]
    ws.append(weekly_headers)
    style_header_row(ws, 1, len(weekly_headers), 1)

    for week in range(1, 79):
        if week <= 13:
            phase = PHASES[0]
            focus = "HW/SW interface fundamentals"
        elif week <= 26:
            phase = PHASES[1]
            focus = "Correlation and statistical rigor"
        elif week <= 52:
            phase = PHASES[2]
            focus = "Test architecture decisions"
        else:
            phase = PHASES[3]
            focus = "Architect transition and leadership"

        ws.append([
            week,
            phase,
            focus,
            "",
            8,
            "",
            "",
            "Not Started",
            "",
        ])

    widths = {
        1: 8,
        2: 32,
        3: 30,
        4: 40,
        5: 14,
        6: 14,
        7: 44,
        8: 14,
        9: 36,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(weekly_headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = left_wrap if c not in [1, 5, 6, 8] else center

    status_dv = DataValidation(type="list", formula1='"Not Started,In Progress,Done"', allow_blank=False)
    ws.add_data_validation(status_dv)
    status_dv.add(f"H2:H{ws.max_row}")

    done_rule = FormulaRule(formula=['$H2="Done"'], fill=PatternFill("solid", fgColor="C6EFCE"))
    ws.conditional_formatting.add(f"A2:I{ws.max_row}", done_rule)

    ws["K2"] = "Weekly completion"
    ws["L2"] = '=IF(COUNTA(H2:H79)=0,0,COUNTIF(H2:H79,"Done")/COUNTA(H2:H79))'
    ws["L2"].number_format = "0.00%"
    ws["K3"] = "Planned hours total"
    ws["L3"] = "=SUM(E2:E79)"
    ws["K4"] = "Actual hours total"
    ws["L4"] = "=SUM(F2:F79)"
    ws["K5"] = "Hours attainment"
    ws["L5"] = "=IF(L3=0,0,L4/L3)"
    ws["L5"].number_format = "0.00%"

    for addr in ["K2", "K3", "K4", "K5"]:
        ws[addr].font = Font(bold=True)
        ws[addr].fill = PatternFill("solid", fgColor="D9E1F2")
        ws[addr].border = border
    for addr in ["L2", "L3", "L4", "L5"]:
        ws[addr].border = border
        ws[addr].alignment = center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"


def create_scorecard_sheet(wb: Workbook):
    ws = wb.create_sheet("Promotion Scorecard")
    ws.append(["Competency", "Weight %", "Self Rating (1-5)", "Weighted Score", "Evidence / Notes"])
    style_header_row(ws, 1, 5, 1)

    competencies = [
        ("ATE platform depth (UltraFLEX + IG-XL)", 15),
        ("HW-SW interface debug skill", 15),
        ("Correlation & data interpretation rigor", 12),
        ("Test architecture strategy", 14),
        ("DFT and testability awareness", 8),
        ("Automation and coding quality", 10),
        ("Cross-functional communication", 8),
        ("Root-cause leadership in escalations", 8),
        ("Execution consistency and time management", 5),
        ("Mentoring and influence", 5),
    ]

    for comp, weight in competencies:
        ws.append([comp, weight, 3, "", ""])

    for r in range(2, ws.max_row + 1):
        ws[f"D{r}"] = f"=B{r}*C{r}/5"

    ws.append(["", "", "", "", ""])
    total_row = ws.max_row + 1
    ws[f"A{total_row}"] = "Total weighted score (%)"
    ws[f"B{total_row}"] = f"=SUM(D2:D{total_row - 2})"
    ws[f"A{total_row + 1}"] = "Readiness level"
    ws[f"B{total_row + 1}"] = f'=IF(B{total_row}>=85,"Ready for Test Architect scope",IF(B{total_row}>=70,"Near-ready (close gaps)","Build foundations further"))'
    ws[f"A{total_row + 2}"] = "Auto evidence from plan"
    ws[f"B{total_row + 2}"] = '=COUNTIF(\'Learning Plan\'!B:B,"Done")&" completed items"'

    ws[f"D{total_row}"] = "=B{0}/100".format(total_row)
    ws[f"D{total_row}"].number_format = "0.00%"

    for r in range(2, total_row + 2):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = left_wrap if c in [1, 5] else center

    for cell_addr in [f"A{total_row}", f"A{total_row + 1}", f"A{total_row + 2}"]:
        ws[cell_addr].font = Font(bold=True)
        ws[cell_addr].fill = PatternFill("solid", fgColor="FFF2CC")
        ws[cell_addr].border = border

    for cell_addr in [f"B{total_row}", f"B{total_row + 1}", f"B{total_row + 2}", f"D{total_row}"]:
        ws[cell_addr].border = border
        ws[cell_addr].alignment = center

    rating_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=False)
    ws.add_data_validation(rating_dv)
    rating_dv.add(f"C2:C{1 + len(competencies)}")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 44
    ws.freeze_panes = "A2"


def create_legend_sheet(wb: Workbook):
    legend = wb.create_sheet("Legend")
    legend.append(["Category", "Meaning", "Color"])
    style_header_row(legend, 1, 3, 1)

    legend_rows = [
        ("Course", "Structured learning program or module", ""),
        ("Certification", "Formal credential with exam", ""),
        ("Book", "Self-study reference", ""),
        ("Exposure", "On-the-job learning by doing", ""),
        ("Soft Skill", "Communication/leadership/productivity skill", ""),
    ]
    for row in legend_rows:
        legend.append(row)

    legend.column_dimensions["A"].width = 20
    legend.column_dimensions["B"].width = 55
    legend.column_dimensions["C"].width = 18
    legend.freeze_panes = "A2"

    for r in range(2, legend.max_row + 1):
        category = legend.cell(row=r, column=1).value
        fill = PatternFill("solid", fgColor=category_colors.get(category, "FFFFFF"))
        for c in range(1, 4):
            cell = legend.cell(row=r, column=c)
            cell.fill = fill
            cell.border = border
            cell.alignment = left_wrap if c != 3 else center


def main():
    wb = Workbook()
    create_learning_plan_sheet(wb)
    create_summary_sheet(wb)
    create_weekly_schedule_sheet(wb)
    create_scorecard_sheet(wb)
    create_legend_sheet(wb)
    wb.save(OUTPUT_FILE)
    print(f"Created: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
