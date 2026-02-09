"""ODI import CSV generator (flat execution, no CLI).

Reads ODI (offset/drift information) from the reference Excel workbook and generates a
semicolon-separated CSV compatible with the `ImportOdi` command described in
`ImportOdi.md`.

Key rules implemented:
- Read ODI from sheets 50_TXGE ... 58_TXPS.
- Module name = substring after underscore in sheet name (e.g. 52_DPLL -> DPLL).
- Test number from column "Test Number" (formula results via openpyxl data_only=True).
- ODI values from columns whose header starts with "ODI" (also formula results).
- Only non-zero ODI values are exported.
- ODI scope: header contains LTL -> LowerLimit; contains UTL -> UpperLimit.
- Insertion name: last token of the ODI column header (e.g. "ODI LTL S1" -> S1).
- Output mapping: ODI value goes to column "S"; column "R" is kept empty.

Configure parameters in CONFIG below and run this file directly.
"""

from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
import csv
import re
from typing import Iterable, Optional
from openpyxl import load_workbook


@dataclass(frozen=True)
class OdiImportConfig:
	reference_xlsx: Path
	output_csv: Path

	# Selection
	modules: Optional[set[str]] = None  # None => all modules (50..58)
	odi_columns: Optional[set[str]] = None  # None => all ODI columns; else match by header

	# Output fields
	comment: str = ""
	jira_tasks: str = ""  # comma-separated (e.g. "RSIPPTE-493,ABC-123")
	odi_source: str = ""  # e.g. "ATE", "INS_OFF"
	test_variants: str = ""  # comma-separated variants (e.g. "8188")

	# Behavior
	include_zero_odi: bool = False  # default exports only non-zero values


CONFIG = OdiImportConfig(
	# Reference workbook (the repo contains CTRX8188A_TE_TX.xlsx)
	reference_xlsx=Path(__file__).resolve().parents[2] / "CTRX8188A_TE_TX.xlsx",

	# Output file to generate
	output_csv=Path(__file__).resolve().parents[2] / "ODIs_TX.csv",

	# Optional filters (examples):
	# modules={"DPLL", "TXGE"},
	# odi_columns={"ODI UTL B1", "ODI UTL B2"},
	#modules=None, 
	#odi_columns=None,
    #modules={"TXGE","DPLL","TXPA","TXPB","TXPC","TXPD","TXLO","TXPS"},
	#odi_columns={"ODI LTL S1","ODI UTL S1","ODI LTL S2","ODI UTL S2","ODI LTL B1","ODI UTL B1","ODI LTL B2","ODI UTL B2","ODI LTL Q1","ODI UTL Q1","ODI LTL Q2","ODI UTL Q2","ODI LTL Q3","ODI UTL Q3"},
	modules={"TXGE","TXPD"},
	odi_columns={"ODI LTL S1","ODI UTL S1","ODI LTL S2","ODI UTL S2","ODI LTL B1","ODI UTL B1","ODI LTL B2","ODI UTL B2","ODI LTL Q1","ODI UTL Q1","ODI LTL Q2","ODI UTL Q2","ODI LTL Q3","ODI UTL Q3"},
	# Output metadata
	comment="TX updates from the Test-D and Test-Q ELFR verifications",
	jira_tasks="RSIPPTE-600,RSIPPTE-587",
	odi_source="INS_OFF",
	test_variants="8188",
)


CSV_HEADER = [
	"TestNumber",
	"OdiSource",
	"Insertions",
	"TestVariants",
	"OdiScope",
	"R",
	"S",
	"CalcMethod",
	"Link2Evidence",
	"Comment",
	"JiraTasks",
]


@dataclass(frozen=True)
class GenerationResult:
	output_csv: Path
	row_count: int
	processed_sheets: tuple[str, ...]
	selected_modules: Optional[tuple[str, ...]]
	selected_odi_columns: Optional[tuple[str, ...]]


def _normalize_header(value: str) -> str:
	return " ".join(value.strip().split()).casefold()


def _find_test_number_idx(headers: list[str]) -> Optional[int]:
	# Be tolerant: workbooks sometimes vary the header text slightly.
	candidates = {"test number", "testnumber", "test_number"}
	for idx, header in enumerate(headers):
		if not header:
			continue
		if _normalize_header(header) in candidates:
			return idx
	return None


def _is_nonzero(value: object) -> bool:
	if value is None:
		return False
	if isinstance(value, bool):
		return bool(value)
	if isinstance(value, (int, float)):
		return abs(float(value)) > 1e-12
	if isinstance(value, str):
		text = value.strip()
		if text == "":
			return False
		try:
			return abs(float(text)) > 1e-12
		except ValueError:
			return True
	return True


def _to_float(value: object) -> float:
	if isinstance(value, (int, float)):
		return float(value)
	if isinstance(value, str):
		return float(value.strip())
	raise TypeError(f"Unsupported ODI value type: {type(value)}")


def _normalize_s_value(value: float) -> float:
	# Excel cached results often contain tiny FP noise; normalize for grouping and output.
	if abs(value) <= 1e-12:
		return 0.0
	return round(value, 12)


def _format_s_value(value: float) -> str:
	if abs(value - round(value)) <= 1e-12:
		return str(int(round(value)))
	# Significant digits formatting to avoid artifacts like 0.9000000000000004
	return format(value, ".15g")



_INSERTION_TOKEN_RE = re.compile(r"^[A-Za-z]{1,3}\d{1,2}$")


def _parse_insertion_tokens(odi_header: str) -> list[str]:
	# Extract tokens like S1, S2, B1, Q3 even if punctuation is attached.
	# Also supports combined tokens like "B1/B2" or "B1,B2".
	tokens = [t.strip(" ,;()[]{}\t\n\r") for t in odi_header.strip().split()]
	insertions: list[str] = []
	for token in tokens:
		if not token:
			continue
		# Split combined forms.
		parts = re.split(r"[,/;+&]", token)
		for part in parts:
			p = part.strip(" ,;()[]{}\t\n\r").upper()
			if not p:
				continue
			if _INSERTION_TOKEN_RE.match(p):
				insertions.append(p)
	return insertions


def _extract_scope_and_insertions(odi_header: str) -> tuple[str, tuple[str, ...]]:
	normalized = _normalize_header(odi_header)
	if "ltl" in normalized:
		scope = "LowerLimit"
	elif "utl" in normalized:
		scope = "UpperLimit"
	else:
		raise ValueError(
			f"ODI column header must contain LTL or UTL to derive scope: {odi_header!r}"
		)

	insertions = _parse_insertion_tokens(odi_header)
	if not insertions:
		# Backwards-compatible fallback: last token.
		tokens = odi_header.strip().split()
		if not tokens:
			raise ValueError("Empty ODI header")
		insertions = [tokens[-1].strip(" ,;()[]{}\t\n\r").upper()]

	# De-duplicate while preserving a stable order.
	seen: set[str] = set()
	ordered: list[str] = []
	for ins in insertions:
		if ins not in seen:
			seen.add(ins)
			ordered.append(ins)
	return scope, tuple(ordered)


def _iter_target_sheets(sheetnames: Iterable[str]) -> list[str]:
	targets: list[str] = []
	for name in sheetnames:
		# Expect e.g. "50_TXGE" ... "58_TXPS"
		if "_" not in name:
			continue
		prefix, _ = name.split("_", 1)
		if not prefix.isdigit():
			continue
		idx = int(prefix)
		if 50 <= idx <= 58:
			targets.append(name)
	return targets


def generate_odi_csv(config: OdiImportConfig) -> GenerationResult:
	if not config.reference_xlsx.exists():
		raise FileNotFoundError(f"Reference workbook not found: {config.reference_xlsx}")

	config.output_csv.parent.mkdir(parents=True, exist_ok=True)

	# data_only=True is required to read cached formula results.
	wb = load_workbook(config.reference_xlsx, data_only=True, read_only=True)
	sheetnames = _iter_target_sheets(wb.sheetnames)

	selected_modules_norm: Optional[set[str]] = (
		None
		if config.modules is None
		else {m.strip().casefold() for m in config.modules if m.strip()}
	)

	normalized_selected_odi_cols: Optional[set[str]] = (
		None
		if config.odi_columns is None
		else {_normalize_header(c) for c in config.odi_columns}
	)

	# Also support semantic selection: match by (scope, insertion) even if the sheet header
	# contains extra tokens (e.g. "ODI LTL Q1 (new)") and wouldn't be an exact match.
	requested_scope_insertion: Optional[set[tuple[str, str]]] = None
	if config.odi_columns is not None:
		requested_scope_insertion = set()
		for c in config.odi_columns:
			try:
				scope, insertions = _extract_scope_and_insertions(c)
			except Exception:
				continue
			for ins in insertions:
				requested_scope_insertion.add((scope, ins))

	# Group rows when multiple insertions have same TestNumber + scope + S value
	# (Example ODI_Test.csv groups B1,B2 for the same value)
	grouped: dict[tuple[int, str, float], set[str]] = {}
	processed_sheets: list[str] = []
	seen_odi_headers_norm: set[str] = set()

	for sheet_name in sheetnames:
		module = sheet_name.split("_", 1)[1]
		if selected_modules_norm is not None and module.strip().casefold() not in selected_modules_norm:
			continue
		processed_sheets.append(sheet_name)

		ws = wb[sheet_name]

		header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
		if not header_row:
			continue

		headers_str = [h if isinstance(h, str) else "" for h in header_row]

		test_number_idx = _find_test_number_idx(headers_str)
		if test_number_idx is None:
			continue

		odi_cols: list[tuple[int, str]] = []  # (0-based index, header)
		for idx, header in enumerate(headers_str):
			if not header:
				continue
			if not _normalize_header(header).startswith("odi"):
				continue
			if normalized_selected_odi_cols is not None:
				norm = _normalize_header(header)
				seen_odi_headers_norm.add(norm)
				if norm not in normalized_selected_odi_cols:
					# Fallback semantic match
					if requested_scope_insertion is None:
						continue
					try:
						scope, insertions = _extract_scope_and_insertions(header)
					except Exception:
						continue
					if not any((scope, ins) in requested_scope_insertion for ins in insertions):
						continue
			else:
				seen_odi_headers_norm.add(_normalize_header(header))
			odi_cols.append((idx, header))

		if not odi_cols:
			continue

		blank_streak = 0
		for row in ws.iter_rows(min_row=2, values_only=True):
			if test_number_idx >= len(row):
				continue

			test_number = row[test_number_idx]
			if test_number in (None, ""):
				blank_streak += 1
				if blank_streak > 200:
					break
				continue
			blank_streak = 0

			# Excel might store it as float even if displayed as int
			try:
				test_number_int = int(float(test_number))
			except Exception:
				continue

			for col_idx, header in odi_cols:
				if col_idx >= len(row):
					continue
				odi_value = row[col_idx]
				if not config.include_zero_odi and not _is_nonzero(odi_value):
					continue

				scope, insertions = _extract_scope_and_insertions(header)
				s_value = _normalize_s_value(_to_float(odi_value))
				if abs(s_value) <= 1e-12:
					continue

				key = (test_number_int, scope, s_value)
				grouped.setdefault(key, set()).update(insertions)

	rows = []
	for (test_number, scope, s_value), insertions in grouped.items():
		rows.append(
			{
				"TestNumber": str(test_number),
				"OdiSource": config.odi_source,
				"Insertions": ",".join(sorted(insertions)),
				"TestVariants": config.test_variants,
				"OdiScope": scope,
				"R": "",
				"S": _format_s_value(s_value),
				"CalcMethod": "",
				"Link2Evidence": "",
				"Comment": config.comment,
				"JiraTasks": config.jira_tasks,
			}
		)

	rows.sort(key=lambda r: (int(r["TestNumber"]), r["OdiScope"], r["Insertions"]))

	with config.output_csv.open("w", newline="", encoding="utf-8") as f:
		writer = csv.DictWriter(f, fieldnames=CSV_HEADER, delimiter=";")
		writer.writeheader()
		writer.writerows(rows)

	return GenerationResult(
		output_csv=config.output_csv,
		row_count=len(rows),
		processed_sheets=tuple(processed_sheets),
		selected_modules=None if config.modules is None else tuple(sorted(config.modules)),
		selected_odi_columns=None if config.odi_columns is None else tuple(sorted(config.odi_columns)),
	)


def main() -> None:
	result = generate_odi_csv(CONFIG)
	print(f"Generated ODI CSV: {result.output_csv}")
	print(f"Rows written: {result.row_count}")
	print(f"Sheets processed: {len(result.processed_sheets)}")

	# Extra diagnostics for common "missing ODI" cases.
	if CONFIG.odi_columns is not None:
		# Note: this checks whether requested ODI column headers exist in the workbook.
		# If formulas have no cached results, headers may exist but values will still be empty.
		requested = {_normalize_header(c) for c in CONFIG.odi_columns}
		# Re-load a tiny bit of metadata to get actual headers found (cheap even for big XLSX)
		try:
			wb = load_workbook(CONFIG.reference_xlsx, data_only=True, read_only=True)
			found: set[str] = set()
			for sheet_name in _iter_target_sheets(wb.sheetnames):
				ws = wb[sheet_name]
				header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
				if not header_row:
					continue
				for h in header_row:
					if isinstance(h, str) and _normalize_header(h).startswith("odi"):
						found.add(_normalize_header(h))
			missing = sorted(requested - found)
			if missing:
				print(
					"WARNING: Some requested ODI column headers were not found in the workbook.\n"
					"Check exact spelling/spacing in the Excel header row. Missing headers (normalized):\n"
					+ "\n".join(f"- {m}" for m in missing)
				)
		except Exception:
			# Diagnostics must never block generation.
			pass

	if result.row_count == 0:
		selected_modules = "ALL" if result.selected_modules is None else ", ".join(result.selected_modules)
		selected_odi_cols = "ALL" if result.selected_odi_columns is None else ", ".join(result.selected_odi_columns)
		print(
			"WARNING: No ODI rows were produced. Common causes:\n"
			"- The selected module(s) or ODI column name(s) do not match the sheet headers exactly.\n"
			"- The workbook contains formulas but does not have cached results saved.\n"
			"  openpyxl cannot calculate formulas; it can only read cached values (data_only=True).\n"
			"  Fix: open the workbook in Excel, force recalculation (Ctrl+Alt+F9), then save it.\n"
			f"Selected modules: {selected_modules}\n"
			f"Selected ODI columns: {selected_odi_cols}\n"
			f"Sheets processed: {result.processed_sheets}"
		)


if __name__ == "__main__":
	main()

