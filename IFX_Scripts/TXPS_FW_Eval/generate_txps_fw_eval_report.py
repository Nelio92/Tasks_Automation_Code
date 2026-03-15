from __future__ import annotations
import argparse
import csv
import importlib.util
import math
import re
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from scipy.stats import ks_2samp

SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE_ROOT = Path(__file__).resolve().parents[3]
TEST_DATA_ANALYSIS_DIR = SCRIPT_DIR.parent / "Test_Data_Analysis"


def _load_stdf_helpers():
    module_path = TEST_DATA_ANALYSIS_DIR / "stdf_to_flat_csv.py"
    spec = importlib.util.spec_from_file_location("stdf_to_flat_csv", module_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Unable to load STDF converter helpers from {module_path}")

    module = importlib.util.module_from_spec(spec)
    sys.modules.setdefault("stdf_to_flat_csv", module)
    spec.loader.exec_module(module)
    return module.convert_stdf_file, module.csv_name_for_source


convert_stdf_file, csv_name_for_source = _load_stdf_helpers()

FIXED_COLUMNS = [
    "UNIT_ID",
    "SITE_NUM",
    "WAFER",
    "X",
    "Y",
    "LOT",
    "SUBLOT",
    "CHIP_ID",
    "PF",
    "FIRST_FAIL_TEST",
]

CASE_LABELS = {
    "PhErr0": "Phase Error 0°",
    "PhErr90": "Phase Error 90°",
    "PhErr180": "Phase Error 180°",
    "PhErr270": "Phase Error 270°",
    "PhRan1Err": "Phase Range 1 Error (0° → 180°)",
    "PhRan2Err": "Phase Range 2 Error (90° → 270°)",
}

INSERTION_TEMPERATURE = {
    "S1": "Hot (135°C)",
    "S2": "Cold (-40°C)",
    "S3": "Ambient (25°C)",
    "B1": "Hot (135°C)",
    "B2": "Ambient (25°C)",
}

TARGET_TEST_RANGE = tuple(range(58020, 58068))
TARGET_FW_TEST_RANGE = tuple(range(158020, 158068))
TEST_NAME_PATTERN = re.compile(
    r"^TXPS(?P<fw>_FW)?_(?P<case>PhErr(?:0|90|180|270)|PhRan[12]Err)Tx(?P<tx>[1-8])"
)


@dataclass
class SourceFile:
    input_path: Path
    csv_path: Path
    source_kind: str


@dataclass
class CaseSeries:
    manual_values: np.ndarray
    fw_values: np.ndarray
    paired_manual: np.ndarray
    paired_fw: np.ndarray
    paired_delta: np.ndarray
    manual_columns: list[str]
    fw_columns: list[str]
    manual_by_tx: dict[int, np.ndarray]
    fw_by_tx: dict[int, np.ndarray]
    paired_delta_by_tx: dict[int, np.ndarray]
    low_limit: float | None
    high_limit: float | None


@dataclass
class NumericStats:
    count: int
    mean: float | None
    median: float | None
    std: float | None
    minimum: float | None
    maximum: float | None
    p05: float | None
    p95: float | None
    abs_mean: float | None
    abs_median: float | None


def _format_value(value: float | None, digits: int = 4) -> str:
    if value is None or (isinstance(value, float) and not math.isfinite(value)):
        return "n/a"
    return f"{value:.{digits}f}"


def _numeric_stats(values: np.ndarray) -> NumericStats:
    if values.size == 0:
        return NumericStats(0, None, None, None, None, None, None, None, None, None)
    return NumericStats(
        count=int(values.size),
        mean=float(np.mean(values)),
        median=float(np.median(values)),
        std=float(np.std(values, ddof=1)) if values.size > 1 else 0.0,
        minimum=float(np.min(values)),
        maximum=float(np.max(values)),
        p05=float(np.percentile(values, 5)),
        p95=float(np.percentile(values, 95)),
        abs_mean=float(np.mean(np.abs(values))),
        abs_median=float(np.median(np.abs(values))),
    )


def _cdf_xy(values: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    if values.size == 0:
        return np.array([]), np.array([])
    sorted_values = np.sort(values)
    probabilities = np.arange(1, sorted_values.size + 1, dtype=float) / sorted_values.size
    return sorted_values, probabilities


def _safe_ks(manual_values: np.ndarray, fw_values: np.ndarray) -> tuple[float | None, float | None]:
    if manual_values.size < 2 or fw_values.size < 2:
        return None, None
    result = ks_2samp(manual_values, fw_values)
    return float(result.statistic), float(result.pvalue)


def _extract_primary_test_name(cell_value: str) -> str:
    return str(cell_value).split(" <>", 1)[0].strip()


def _is_data_row(row: Sequence[str]) -> bool:
    if not row:
        return False
    return row[0].strip().isdigit()


def _parse_insertion_and_temperature(file_name: str) -> tuple[str, str]:
    match = re.search(r"(?:^|[_-])(S[123]|B[12])(?:[_-]|$)", file_name, re.IGNORECASE)
    if not match:
        return "Unknown", "Unknown"
    insertion = match.group(1).upper()
    return insertion, INSERTION_TEMPERATURE.get(insertion, "Unknown")


def _parse_numeric_cell(row: Sequence[str], index: int) -> float | None:
    if index >= len(row):
        return None
    value = str(row[index]).strip()
    if not value:
        return None
    try:
        parsed = float(value)
    except ValueError:
        return None
    return parsed if math.isfinite(parsed) else None


def _display_path(path: Path) -> str:
    try:
        return str(path.relative_to(WORKSPACE_ROOT))
    except ValueError:
        return str(path)


def _read_flat_csv(csv_path: Path) -> tuple[pd.DataFrame, dict[str, str], dict[str, float | None], dict[str, float | None]]:
    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=";"))
    if len(rows) < 3:
        raise ValueError(f"Unexpected CSV format in {csv_path}")

    header = rows[0]
    test_name_row = rows[1]
    row_by_label = {row[0].strip(): row for row in rows[2:] if row and row[0].strip()}
    low_row = row_by_label.get("Low")
    high_row = row_by_label.get("High")
    data_start_index = next((index for index, row in enumerate(rows[2:], start=2) if _is_data_row(row)), None)
    if data_start_index is None:
        raise ValueError(f"No measurement rows found in {csv_path}")

    data_rows = rows[data_start_index:]
    dataframe = pd.DataFrame(data_rows, columns=header)
    dataframe = dataframe.replace("", pd.NA)

    test_name_by_column: dict[str, str] = {}
    low_limit_by_column: dict[str, float | None] = {}
    high_limit_by_column: dict[str, float | None] = {}
    for column_name, raw_test_name in zip(header, test_name_row):
        if column_name not in FIXED_COLUMNS:
            test_name_by_column[column_name] = _extract_primary_test_name(raw_test_name)
            column_index = header.index(column_name)
            low_limit_by_column[column_name] = _parse_numeric_cell(low_row, column_index) if low_row else None
            high_limit_by_column[column_name] = _parse_numeric_cell(high_row, column_index) if high_row else None

    return dataframe, test_name_by_column, low_limit_by_column, high_limit_by_column


def _collect_input_files(input_folder: Path, converted_dir: Path) -> list[SourceFile]:
    sources: list[SourceFile] = []
    for path in sorted(input_folder.iterdir(), key=lambda item: item.name.lower()):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix in {".std", ".stdf"}:
            csv_path = converted_dir / csv_name_for_source(path.name)
            if not csv_path.exists() or csv_path.stat().st_mtime < path.stat().st_mtime:
                print(f"Converting {path.name} to flat CSV...")
                convert_stdf_file(path, csv_path, artifacts_output_folder=converted_dir / "artifacts")
            sources.append(SourceFile(input_path=path, csv_path=csv_path, source_kind="stdf"))
        elif suffix == ".csv":
            sources.append(SourceFile(input_path=path, csv_path=path, source_kind="csv"))
    return sources


def _build_case_column_map(test_name_by_column: dict[str, str]) -> dict[str, dict[str, dict[int, str]]]:
    case_columns: dict[str, dict[str, dict[int, str]]] = {
        case_key: {"manual": {}, "fw": {}} for case_key in CASE_LABELS
    }
    target_tests = {str(test_number) for test_number in (*TARGET_TEST_RANGE, *TARGET_FW_TEST_RANGE)}
    for column_name, test_name in test_name_by_column.items():
        if column_name not in target_tests:
            continue
        match = TEST_NAME_PATTERN.match(test_name)
        if not match:
            continue
        case_key = match.group("case")
        tx_index = int(match.group("tx"))
        method = "fw" if match.group("fw") else "manual"
        case_columns[case_key][method][tx_index] = column_name
    return case_columns


def _pick_case_limit(limit_by_column: dict[str, float | None], columns: Sequence[str]) -> float | None:
    finite_values = [limit_by_column.get(column) for column in columns]
    finite_values = [value for value in finite_values if value is not None and math.isfinite(value)]
    if not finite_values:
        return None
    rounded_counts = Counter(round(value, 9) for value in finite_values)
    most_common_value, _ = rounded_counts.most_common(1)[0]
    return float(most_common_value)


def _extract_case_series(
    dataframe: pd.DataFrame,
    case_columns: dict[str, dict[int, str]],
    low_limit_by_column: dict[str, float | None],
    high_limit_by_column: dict[str, float | None],
) -> CaseSeries:
    manual_values: list[float] = []
    fw_values: list[float] = []
    paired_manual: list[float] = []
    paired_fw: list[float] = []
    paired_delta: list[float] = []
    manual_by_tx: dict[int, np.ndarray] = {}
    fw_by_tx: dict[int, np.ndarray] = {}
    paired_delta_by_tx: dict[int, np.ndarray] = {}

    manual_columns = [case_columns["manual"][tx] for tx in sorted(case_columns["manual"])]
    fw_columns = [case_columns["fw"][tx] for tx in sorted(case_columns["fw"])]

    tx_indices = sorted(set(case_columns["manual"]).intersection(case_columns["fw"]))
    for tx_index in tx_indices:
        manual_column = case_columns["manual"][tx_index]
        fw_column = case_columns["fw"][tx_index]

        manual_series = pd.to_numeric(dataframe[manual_column], errors="coerce")
        fw_series = pd.to_numeric(dataframe[fw_column], errors="coerce")

        manual_channel_values = manual_series.dropna().to_numpy(dtype=float)
        fw_channel_values = fw_series.dropna().to_numpy(dtype=float)
        manual_by_tx[tx_index] = manual_channel_values
        fw_by_tx[tx_index] = fw_channel_values

        manual_values.extend(manual_channel_values.tolist())
        fw_values.extend(fw_channel_values.tolist())

        valid_pair_mask = manual_series.notna() & fw_series.notna()
        if valid_pair_mask.any():
            manual_pair = manual_series.loc[valid_pair_mask].to_numpy(dtype=float)
            fw_pair = fw_series.loc[valid_pair_mask].to_numpy(dtype=float)
            delta_pair = fw_pair - manual_pair
            paired_manual.extend(manual_pair.tolist())
            paired_fw.extend(fw_pair.tolist())
            paired_delta.extend(delta_pair.tolist())
            paired_delta_by_tx[tx_index] = delta_pair
        else:
            paired_delta_by_tx[tx_index] = np.array([], dtype=float)

    return CaseSeries(
        manual_values=np.asarray(manual_values, dtype=float),
        fw_values=np.asarray(fw_values, dtype=float),
        paired_manual=np.asarray(paired_manual, dtype=float),
        paired_fw=np.asarray(paired_fw, dtype=float),
        paired_delta=np.asarray(paired_delta, dtype=float),
        manual_columns=manual_columns,
        fw_columns=fw_columns,
        manual_by_tx=manual_by_tx,
        fw_by_tx=fw_by_tx,
        paired_delta_by_tx=paired_delta_by_tx,
        low_limit=_pick_case_limit(low_limit_by_column, [*manual_columns, *fw_columns]),
        high_limit=_pick_case_limit(high_limit_by_column, [*manual_columns, *fw_columns]),
    )


def _annotation_text(
    file_label: str,
    insertion: str,
    temperature: str,
    manual_stats: NumericStats,
    fw_stats: NumericStats,
    delta_stats: NumericStats,
    ks_statistic: float | None,
    ks_pvalue: float | None,
    low_limit: float | None,
    high_limit: float | None,
) -> str:
    return "\n".join(
        [
            f"File: {file_label}",
            f"Insertion: {insertion}",
            f"Temperature: {temperature}",
            f"Limits: low={_format_value(low_limit, 3)}  high={_format_value(high_limit, 3)}",
            "",
            f"Manual  n={manual_stats.count}  mean={_format_value(manual_stats.mean)}  median={_format_value(manual_stats.median)}",
            f"FW      n={fw_stats.count}  mean={_format_value(fw_stats.mean)}  median={_format_value(fw_stats.median)}",
            f"Δ(FW-Manual) n={delta_stats.count}  mean={_format_value(delta_stats.mean)}  median={_format_value(delta_stats.median)}",
            f"KS statistic={_format_value(ks_statistic)}  p-value={_format_value(ks_pvalue)}",
        ]
    )


def _overview_headers() -> list[str]:
    headers = [
        "file_name",
        "source_kind",
        "csv_source",
        "insertion",
        "temperature",
        "case_key",
        "case_label",
        "manual_columns",
        "fw_columns",
        "low_limit",
        "high_limit",
    ]
    metric_names = ["count", "mean", "median", "std", "min", "max", "p05", "p95", "abs_mean", "abs_median"]
    for metric_name in metric_names:
        headers.append(f"manual_{metric_name}")
        headers.append(f"fw_{metric_name}")
    headers.extend(
        [
            "paired_count",
            "delta_mean",
            "delta_median",
            "delta_std",
            "delta_min",
            "delta_max",
            "delta_p05",
            "delta_p95",
            "delta_abs_mean",
            "delta_abs_median",
            "ks_statistic",
            "ks_pvalue",
            "plot_path",
        ]
    )
    return headers


def _per_tx_headers() -> list[str]:
    headers = [
        "file_name",
        "source_kind",
        "csv_source",
        "insertion",
        "temperature",
        "case_key",
        "case_label",
        "tx_channel",
        "manual_column",
        "fw_column",
        "low_limit",
        "high_limit",
    ]
    metric_names = ["count", "mean", "median", "std", "min", "max", "p05", "p95", "abs_mean", "abs_median"]
    for metric_name in metric_names:
        headers.append(f"manual_{metric_name}")
        headers.append(f"fw_{metric_name}")
    headers.extend(
        [
            "paired_count",
            "delta_mean",
            "delta_median",
            "delta_std",
            "delta_min",
            "delta_max",
            "delta_p05",
            "delta_p95",
            "delta_abs_mean",
            "delta_abs_median",
            "ks_statistic",
            "ks_pvalue",
        ]
    )
    return headers


def _plot_case_cdf(
    output_path: Path,
    file_label: str,
    case_key: str,
    insertion: str,
    temperature: str,
    case_series: CaseSeries,
) -> dict[str, float | int | None]:
    manual_stats = _numeric_stats(case_series.manual_values)
    fw_stats = _numeric_stats(case_series.fw_values)
    delta_stats = _numeric_stats(case_series.paired_delta)
    ks_statistic, ks_pvalue = _safe_ks(case_series.manual_values, case_series.fw_values)

    figure, axis = plt.subplots(figsize=(11, 7))
    channel_colors = plt.get_cmap("tab10", 8)
    tx_indices = sorted(set(case_series.manual_by_tx).union(case_series.fw_by_tx))
    for tx_index in tx_indices:
        color = channel_colors(tx_index - 1)

        manual_values = case_series.manual_by_tx.get(tx_index, np.array([]))
        manual_x, manual_y = _cdf_xy(manual_values)
        if manual_x.size:
            axis.plot(
                manual_x,
                manual_y,
                label=f"TX{tx_index} Manual (n={manual_values.size})",
                linewidth=1.8,
                linestyle="-",
                color=color,
            )

        fw_values = case_series.fw_by_tx.get(tx_index, np.array([]))
        fw_x, fw_y = _cdf_xy(fw_values)
        if fw_x.size:
            axis.plot(
                fw_x,
                fw_y,
                label=f"TX{tx_index} FW (n={fw_values.size})",
                linewidth=1.8,
                linestyle="--",
                color=color,
            )

    if case_series.low_limit is not None:
        axis.axvline(
            case_series.low_limit,
            color="red",
            linestyle="-",
            linewidth=1.8,
            label=f"Lower limit ({_format_value(case_series.low_limit, 3)})",
        )
    if case_series.high_limit is not None:
        axis.axvline(
            case_series.high_limit,
            color="red",
            linestyle="-",
            linewidth=1.8,
            alpha=0.85,
            label=f"Upper limit ({_format_value(case_series.high_limit, 3)})",
        )

    axis.set_title(f"{CASE_LABELS[case_key]}\n{file_label}", fontsize=14, pad=14)
    axis.set_xlabel("Phase error (deg)")
    axis.set_ylabel("CDF")
    axis.set_ylim(0, 1)
    axis.grid(True, linestyle="--", alpha=0.35)
    axis.legend(loc="lower right", ncol=2, fontsize=8)

    annotation = _annotation_text(
        file_label=file_label,
        insertion=insertion,
        temperature=temperature,
        manual_stats=manual_stats,
        fw_stats=fw_stats,
        delta_stats=delta_stats,
        ks_statistic=ks_statistic,
        ks_pvalue=ks_pvalue,
        low_limit=case_series.low_limit,
        high_limit=case_series.high_limit,
    )
    axis.text(
        0.02,
        0.98,
        annotation,
        transform=axis.transAxes,
        fontsize=9,
        va="top",
        ha="left",
        bbox={"boxstyle": "round", "facecolor": "white", "alpha": 0.9, "edgecolor": "#bbbbbb"},
    )

    figure.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(figure)

    return {
        "low_limit": case_series.low_limit,
        "high_limit": case_series.high_limit,
        "manual_count": manual_stats.count,
        "manual_mean": manual_stats.mean,
        "manual_median": manual_stats.median,
        "manual_std": manual_stats.std,
        "manual_min": manual_stats.minimum,
        "manual_max": manual_stats.maximum,
        "manual_p05": manual_stats.p05,
        "manual_p95": manual_stats.p95,
        "manual_abs_mean": manual_stats.abs_mean,
        "manual_abs_median": manual_stats.abs_median,
        "fw_count": fw_stats.count,
        "fw_mean": fw_stats.mean,
        "fw_median": fw_stats.median,
        "fw_std": fw_stats.std,
        "fw_min": fw_stats.minimum,
        "fw_max": fw_stats.maximum,
        "fw_p05": fw_stats.p05,
        "fw_p95": fw_stats.p95,
        "fw_abs_mean": fw_stats.abs_mean,
        "fw_abs_median": fw_stats.abs_median,
        "paired_count": delta_stats.count,
        "delta_mean": delta_stats.mean,
        "delta_median": delta_stats.median,
        "delta_std": delta_stats.std,
        "delta_min": delta_stats.minimum,
        "delta_max": delta_stats.maximum,
        "delta_p05": delta_stats.p05,
        "delta_p95": delta_stats.p95,
        "delta_abs_mean": delta_stats.abs_mean,
        "delta_abs_median": delta_stats.abs_median,
        "ks_statistic": ks_statistic,
        "ks_pvalue": ks_pvalue,
    }


def _populate_worksheet(worksheet, headers: Sequence[str], rows: Sequence[dict[str, object]], *, enable_plot_hyperlink: bool = False) -> None:
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    if enable_plot_hyperlink and "plot_path" in headers:
        for row_index in range(2, worksheet.max_row + 1):
            plot_cell = worksheet.cell(row=row_index, column=headers.index("plot_path") + 1)
            if plot_cell.value:
                plot_cell.hyperlink = str(plot_cell.value)
                plot_cell.style = "Hyperlink"

    for column_index, header in enumerate(headers, start=1):
        values = [str(header)] + ["" if value is None else str(value) for value in (row.get(header) for row in rows)]
        width = min(max(len(value) for value in values) + 2, 42)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _write_overview_workbook(
    output_path: Path,
    overview_rows: Iterable[dict[str, object]],
    per_tx_rows: Iterable[dict[str, object]],
) -> None:
    overview_rows = list(overview_rows)
    per_tx_rows = list(per_tx_rows)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Overview"

    if not overview_rows and not per_tx_rows:
        worksheet["A1"] = "No results generated"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return

    _populate_worksheet(worksheet, _overview_headers(), overview_rows, enable_plot_hyperlink=True)

    per_tx_sheet = workbook.create_sheet("Per_TX")
    _populate_worksheet(per_tx_sheet, _per_tx_headers(), per_tx_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def generate_report(input_folder: Path, output_folder: Path) -> tuple[list[dict[str, object]], Path]:
    converted_dir = output_folder / "converted_csv"
    plots_dir = output_folder / "cdf_plots"
    converted_dir.mkdir(parents=True, exist_ok=True)
    plots_dir.mkdir(parents=True, exist_ok=True)

    source_files = _collect_input_files(input_folder, converted_dir)
    if not source_files:
        raise FileNotFoundError(f"No .std/.stdf/.csv files found in {input_folder}")

    overview_rows: list[dict[str, object]] = []
    per_tx_rows: list[dict[str, object]] = []

    for source in source_files:
        print(f"Analyzing {source.input_path.name}...")
        dataframe, test_name_by_column, low_limit_by_column, high_limit_by_column = _read_flat_csv(source.csv_path)
        case_column_map = _build_case_column_map(test_name_by_column)
        insertion, temperature = _parse_insertion_and_temperature(source.input_path.stem)

        for case_key in CASE_LABELS:
            case_series = _extract_case_series(
                dataframe,
                case_column_map[case_key],
                low_limit_by_column,
                high_limit_by_column,
            )
            if case_series.manual_values.size == 0 and case_series.fw_values.size == 0:
                print(f"  Skipping {case_key}: no matching data")
                continue

            plot_path = plots_dir / source.input_path.stem / f"{case_key}.png"
            metrics = _plot_case_cdf(
                output_path=plot_path,
                file_label=source.input_path.name,
                case_key=case_key,
                insertion=insertion,
                temperature=temperature,
                case_series=case_series,
            )

            overview_rows.append(
                {
                    "file_name": source.input_path.name,
                    "source_kind": source.source_kind,
                    "csv_source": _display_path(source.csv_path),
                    "insertion": insertion,
                    "temperature": temperature,
                    "case_key": case_key,
                    "case_label": CASE_LABELS[case_key],
                    "manual_columns": ", ".join(case_series.manual_columns),
                    "fw_columns": ", ".join(case_series.fw_columns),
                    **metrics,
                    "plot_path": _display_path(plot_path),
                }
            )

            for tx_index in sorted(set(case_series.manual_by_tx).union(case_series.fw_by_tx)):
                manual_values = case_series.manual_by_tx.get(tx_index, np.array([], dtype=float))
                fw_values = case_series.fw_by_tx.get(tx_index, np.array([], dtype=float))
                delta_values = case_series.paired_delta_by_tx.get(tx_index, np.array([], dtype=float))

                manual_stats = _numeric_stats(manual_values)
                fw_stats = _numeric_stats(fw_values)
                delta_stats = _numeric_stats(delta_values)
                ks_statistic, ks_pvalue = _safe_ks(manual_values, fw_values)

                manual_column = case_column_map[case_key]["manual"].get(tx_index, "")
                fw_column = case_column_map[case_key]["fw"].get(tx_index, "")

                per_tx_rows.append(
                    {
                        "file_name": source.input_path.name,
                        "source_kind": source.source_kind,
                        "csv_source": _display_path(source.csv_path),
                        "insertion": insertion,
                        "temperature": temperature,
                        "case_key": case_key,
                        "case_label": CASE_LABELS[case_key],
                        "tx_channel": f"TX{tx_index}",
                        "manual_column": manual_column,
                        "fw_column": fw_column,
                        "low_limit": case_series.low_limit,
                        "high_limit": case_series.high_limit,
                        "manual_count": manual_stats.count,
                        "fw_count": fw_stats.count,
                        "manual_mean": manual_stats.mean,
                        "fw_mean": fw_stats.mean,
                        "manual_median": manual_stats.median,
                        "fw_median": fw_stats.median,
                        "manual_std": manual_stats.std,
                        "fw_std": fw_stats.std,
                        "manual_min": manual_stats.minimum,
                        "fw_min": fw_stats.minimum,
                        "manual_max": manual_stats.maximum,
                        "fw_max": fw_stats.maximum,
                        "manual_p05": manual_stats.p05,
                        "fw_p05": fw_stats.p05,
                        "manual_p95": manual_stats.p95,
                        "fw_p95": fw_stats.p95,
                        "manual_abs_mean": manual_stats.abs_mean,
                        "fw_abs_mean": fw_stats.abs_mean,
                        "manual_abs_median": manual_stats.abs_median,
                        "fw_abs_median": fw_stats.abs_median,
                        "paired_count": delta_stats.count,
                        "delta_mean": delta_stats.mean,
                        "delta_median": delta_stats.median,
                        "delta_std": delta_stats.std,
                        "delta_min": delta_stats.minimum,
                        "delta_max": delta_stats.maximum,
                        "delta_p05": delta_stats.p05,
                        "delta_p95": delta_stats.p95,
                        "delta_abs_mean": delta_stats.abs_mean,
                        "delta_abs_median": delta_stats.abs_median,
                        "ks_statistic": ks_statistic,
                        "ks_pvalue": ks_pvalue,
                    }
                )

    workbook_path = output_folder / "txps_fw_eval_overview.xlsx"
    _write_overview_workbook(workbook_path, overview_rows, per_tx_rows)
    return overview_rows, workbook_path


def _build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Compare manual TXPS measurements against FW-based TXPS measurements and generate CDF plots + Excel overview."
    )
    parser.add_argument(
        "--input-folder",
        type=Path,
        default=SCRIPT_DIR,
        help="Folder containing the TXPS FW evaluation STDF/CSV files.",
    )
    parser.add_argument(
        "--output-folder",
        type=Path,
        default=SCRIPT_DIR / "Outputs" / "TXPS_FW_Comparison",
        help="Folder where plots, converted CSV files, and the overview workbook will be written.",
    )
    return parser


def main() -> None:
    parser = _build_argument_parser()
    args = parser.parse_args()

    overview_rows, workbook_path = generate_report(
        input_folder=args.input_folder.resolve(),
        output_folder=args.output_folder.resolve(),
    )
    print(f"Generated {len(overview_rows)} comparison rows")
    print(f"Overview workbook: {workbook_path}")


if __name__ == "__main__":
    main()
