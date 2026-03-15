"""Flat CSV test data analysis + reporting.

This script is designed for large, semicolon-separated "flat" CSV exports where:
- The header row contains metadata columns (e.g. LOT/WAFER/X/Y/SITE_NUM) and many
  numeric test columns (e.g. 520123, 530045, ...).
- A small block of meta rows follows the header ("Test Name", "Low", "High",
  "Unit", "Cpk", "Yield", ...).
- Unit/device rows come after that meta block, with measurements per test column.

The report is module-agnostic: a test's "module" is extracted from the first 4
characters of its test name (e.g. DPLL, TXPA, TXLO). Modules to analyze are
provided via a user configuration section at the top of this file.

Outputs:
- One Excel workbook containing one sheet per input file with only the tests of
  interest that have yield < 100% or Cpk outside thresholds.
- A per-input-file plots sheet embedding CDF plots; hyperlinks in the data sheet
    jump to the embedded plot, and clicking an embedded plot opens the full PNG.
- Optionally, a separate correlation workbook (one sheet per input file)
  computing Pearson and Spearman correlations for each module test vs all tests.

Note on "hover previews": Excel doesn't support showing an image on hyperlink
hover via openpyxl. This implementation embeds images in a dedicated plots sheet
and links to those locations.
"""

from __future__ import annotations
import csv
import math
import os
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Iterable, Literal
from xml.etree import ElementTree as ET


DEFAULT_ENCODING = "latin1"
DELIMITER = ";"
WAFER_MAP_FAIL_COORDS_MAX_ITEMS = 50
WAFER_MAP_FAIL_COORDS_WRAP_WIDTH_WIDE = 84
WAFER_MAP_FAIL_COORDS_WRAP_WIDTH_NARROW = 56


# ================================================
# INJECTED BY YAML LAUNCHER - DO NOT EDIT MANUALLY
# ================================================
# These values are intentionally populated by `run_tests_data_analysis.py`
# from a YAML config file before `run()` is called.

# Input/Output
INPUT_FOLDER: Path | None = None
OUTPUT_FOLDER: Path | None = None

# Which modules to analyze (first 4 chars of test name)
MODULES: list[str] = []

# Thresholds
YIELD_THRESHOLD: float | None = None
CPK_LOW: float | None = None
CPK_HIGH: float | None = None

# Outlier detection (|x-median| > OUTLIER_MAD_MULTIPLIER * MAD)
OUTLIER_MAD_MULTIPLIER: float | None = None

# Optional controls
MAX_FILES: int | None = None
SINGLE_FILE: str | None = None
ENCODING: str | None = None

# Optional correlation workbook
GENERATE_CORRELATION_REPORT: bool | None = None
CORRELATION_METHODS: list[Literal["pearson", "spearman"]] = []
PEARSON_ABS_MIN_FOR_REPORT: float | None = None

# Wafer map display controls
# Scales the *area* of the wafer outline circle; 2.0 => 2× area, 3.0 => 3× area.
WAFERMAP_CIRCLE_AREA_MULT: float | None = None


def _require_runtime_configuration() -> dict[str, Any]:
    missing: list[str] = []

    if INPUT_FOLDER is None:
        missing.append("INPUT_FOLDER")
    if OUTPUT_FOLDER is None:
        missing.append("OUTPUT_FOLDER")
    if not MODULES:
        missing.append("MODULES")
    if YIELD_THRESHOLD is None:
        missing.append("YIELD_THRESHOLD")
    if CPK_LOW is None:
        missing.append("CPK_LOW")
    if CPK_HIGH is None:
        missing.append("CPK_HIGH")
    if OUTLIER_MAD_MULTIPLIER is None:
        missing.append("OUTLIER_MAD_MULTIPLIER")
    if ENCODING is None:
        missing.append("ENCODING")
    if GENERATE_CORRELATION_REPORT is None:
        missing.append("GENERATE_CORRELATION_REPORT")
    if PEARSON_ABS_MIN_FOR_REPORT is None:
        missing.append("PEARSON_ABS_MIN_FOR_REPORT")
    if WAFERMAP_CIRCLE_AREA_MULT is None:
        missing.append("WAFERMAP_CIRCLE_AREA_MULT")
    if GENERATE_CORRELATION_REPORT and not CORRELATION_METHODS:
        missing.append("CORRELATION_METHODS")

    if missing:
        joined = ", ".join(missing)
        raise RuntimeError(
            "Tests_Data_Analysis runtime configuration is missing. "
            "Run via run_tests_data_analysis.py with a YAML config file. "
            f"Missing: {joined}"
        )

    return {
        "input_folder": _as_path(INPUT_FOLDER),
        "output_folder": _as_path(OUTPUT_FOLDER),
        "modules": list(MODULES),
        "yield_threshold": float(YIELD_THRESHOLD),
        "cpk_low": float(CPK_LOW),
        "cpk_high": float(CPK_HIGH),
        "outlier_mad_multiplier": float(OUTLIER_MAD_MULTIPLIER),
        "max_files": MAX_FILES,
        "single_file": SINGLE_FILE,
        "encoding": str(ENCODING),
        "generate_correlation_report": bool(GENERATE_CORRELATION_REPORT),
        "correlation_methods": list(CORRELATION_METHODS),
        "pearson_abs_min_for_report": float(PEARSON_ABS_MIN_FOR_REPORT),
        "wafermap_circle_area_mult": float(WAFERMAP_CIRCLE_AREA_MULT),
    }


def _as_path(value: str | os.PathLike[str] | Path) -> Path:
    return value if isinstance(value, Path) else Path(value)


def _excel_col_letter(col_idx_1_based: int) -> str:
    """Convert a 1-based column index to an Excel column letter."""
    if col_idx_1_based < 1:
        raise ValueError("Column index must be >= 1")
    n = col_idx_1_based
    letters: list[str] = []
    while n:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def _safe_sheet_name(name: str) -> str:
    # Excel sheet limit is 31 chars and cannot contain some characters.
    safe = re.sub(r"[\\/*?:\[\]]", "_", name)
    safe = safe.strip() or "Sheet"
    return safe[:31]


def _unique_sheet_name(name: str, existing_names) -> str:
    """Return an Excel-safe unique sheet name capped at 31 characters."""
    base = _safe_sheet_name(name)
    existing_lower = {str(item).lower() for item in existing_names}
    if base.lower() not in existing_lower:
        return base

    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}".strip() or f"Sheet{suffix}"
        if candidate.lower() not in existing_lower:
            return candidate
        counter += 1


def _excel_internal_sheet_ref(sheet_name: str) -> str:
    """Return a sheet reference safe for internal Excel hyperlinks.

    Always quote sheet names to support characters like '-' and spaces.
    """
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def _parse_excel_internal_target(target: str) -> tuple[str | None, str | None]:
    """Parse internal Excel hyperlink target like #'Sheet Name'!B4."""
    if not target or not str(target).startswith("#"):
        return None, None

    body = str(target)[1:]
    if "!" not in body:
        return None, None

    raw_sheet, raw_cell = body.split("!", 1)
    sheet = raw_sheet
    if len(sheet) >= 2 and sheet[0] == "'" and sheet[-1] == "'":
        sheet = sheet[1:-1].replace("''", "'")
    return sheet, raw_cell


def _is_valid_excel_a1_ref(cell_ref: str) -> bool:
    """Return True for basic A1 references/ranges used in internal hyperlinks."""
    if not cell_ref:
        return False
    pattern = r"^\$?[A-Z]{1,3}\$?[1-9][0-9]*(:\$?[A-Z]{1,3}\$?[1-9][0-9]*)?$"
    return re.match(pattern, str(cell_ref).upper()) is not None


def _self_check_workbook_internal_hyperlinks(workbook) -> list[str]:
    """Validate all internal (#...) hyperlinks in workbook and return issues."""
    issues: list[str] = []
    sheet_names = set(workbook.sheetnames)

    for ws in workbook.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                hl = cell.hyperlink
                if hl is None:
                    continue
                target = getattr(hl, "target", None)
                if target is None or not str(target).startswith("#"):
                    continue

                sheet_name, cell_ref = _parse_excel_internal_target(str(target))
                if not sheet_name or sheet_name not in sheet_names:
                    issues.append(
                        f"{ws.title}!{cell.coordinate}: invalid target sheet in hyperlink '{target}'"
                    )
                    continue
                if not cell_ref or not _is_valid_excel_a1_ref(cell_ref):
                    issues.append(
                        f"{ws.title}!{cell.coordinate}: invalid target cell ref in hyperlink '{target}'"
                    )

    return issues


def _normalize_ooxml_path(path: str) -> str:
    parts: list[str] = []
    for part in PurePosixPath(path).parts:
        if part in ("", ".", "/", "\\"):
            continue
        if part == "..":
            if parts:
                parts.pop()
            continue
        parts.append(part)
    return "/".join(parts)


def _resolve_ooxml_target(source_part: str, target: str) -> str:
    return _normalize_ooxml_path(str(PurePosixPath(source_part).parent / target))


def _rels_part_for(part_path: str) -> str:
    part = PurePosixPath(part_path)
    return str(part.parent / "_rels" / f"{part.name}.rels")


def _next_relationship_id(rels_root: ET.Element) -> str:
    used: set[int] = set()
    for rel in rels_root:
        rid = rel.attrib.get("Id", "")
        m = re.fullmatch(r"rId(\d+)", rid)
        if m:
            used.add(int(m.group(1)))

    next_id = 1
    while next_id in used:
        next_id += 1
    return f"rId{next_id}"


def _enable_clickable_plot_images(workbook_path: Path, sheet_image_targets: dict[str, list[str]]) -> None:
    """Patch workbook drawings so clicking an embedded plot opens the PNG file."""
    if not sheet_image_targets:
        return

    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkg = "http://schemas.openxmlformats.org/package/2006/relationships"
    ns_xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"
    drawing_rel_type = f"{ns_rel}/drawing"
    hyperlink_rel_type = f"{ns_rel}/hyperlink"

    ET.register_namespace("", ns_main)
    ET.register_namespace("r", ns_rel)
    ET.register_namespace("xdr", ns_xdr)
    ET.register_namespace("a", ns_a)

    with zipfile.ZipFile(workbook_path, "r") as zin:
        archive = {name: zin.read(name) for name in zin.namelist()}

    workbook_part = "xl/workbook.xml"
    workbook_rels_part = "xl/_rels/workbook.xml.rels"
    if workbook_part not in archive or workbook_rels_part not in archive:
        return

    workbook_root = ET.fromstring(archive[workbook_part])
    workbook_rels_root = ET.fromstring(archive[workbook_rels_part])
    workbook_rels_by_id = {
        rel.attrib.get("Id"): rel for rel in workbook_rels_root.findall(f"{{{ns_pkg}}}Relationship")
    }

    sheet_part_by_name: dict[str, str] = {}
    for sheet in workbook_root.findall(f".//{{{ns_main}}}sheet"):
        name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get(f"{{{ns_rel}}}id")
        rel = workbook_rels_by_id.get(rel_id)
        if not name or rel is None:
            continue
        target = rel.attrib.get("Target")
        if not target:
            continue
        sheet_part_by_name[name] = _resolve_ooxml_target(workbook_part, target)

    modified: dict[str, bytes] = {}
    for sheet_name, image_targets in sheet_image_targets.items():
        if not image_targets:
            continue

        sheet_part = sheet_part_by_name.get(sheet_name)
        if not sheet_part:
            continue
        sheet_rels_part = _rels_part_for(sheet_part)
        if sheet_rels_part not in archive:
            continue

        sheet_rels_root = ET.fromstring(archive[sheet_rels_part])
        drawing_target = None
        for rel in sheet_rels_root.findall(f"{{{ns_pkg}}}Relationship"):
            if rel.attrib.get("Type") == drawing_rel_type:
                drawing_target = rel.attrib.get("Target")
                break
        if not drawing_target:
            continue

        drawing_part = _resolve_ooxml_target(sheet_part, drawing_target)
        drawing_rels_part = _rels_part_for(drawing_part)
        if drawing_part not in archive or drawing_rels_part not in archive:
            continue

        drawing_root = ET.fromstring(archive[drawing_part])
        drawing_rels_root = ET.fromstring(archive[drawing_rels_part])

        pic_anchors: list[ET.Element] = []
        for anchor_tag in ("oneCellAnchor", "twoCellAnchor", "absoluteAnchor"):
            for anchor in drawing_root.findall(f"{{{ns_xdr}}}{anchor_tag}"):
                if anchor.find(f"{{{ns_xdr}}}pic") is not None:
                    pic_anchors.append(anchor)

        changed = False
        for target_uri, anchor in zip(image_targets, pic_anchors, strict=False):
            pic = anchor.find(f"{{{ns_xdr}}}pic")
            if pic is None:
                continue
            nv_pic = pic.find(f"{{{ns_xdr}}}nvPicPr")
            if nv_pic is None:
                continue
            c_nv_pr = nv_pic.find(f"{{{ns_xdr}}}cNvPr")
            if c_nv_pr is None:
                continue

            for child in list(c_nv_pr):
                if child.tag == f"{{{ns_a}}}hlinkClick":
                    c_nv_pr.remove(child)

            rel_id = _next_relationship_id(drawing_rels_root)
            ET.SubElement(
                drawing_rels_root,
                f"{{{ns_pkg}}}Relationship",
                {
                    "Id": rel_id,
                    "Type": hyperlink_rel_type,
                    "Target": target_uri,
                    "TargetMode": "External",
                },
            )
            ET.SubElement(
                c_nv_pr,
                f"{{{ns_a}}}hlinkClick",
                {
                    f"{{{ns_rel}}}id": rel_id,
                    "tooltip": "Open full-size PNG",
                },
            )
            changed = True

        if changed:
            modified[drawing_part] = ET.tostring(drawing_root, encoding="utf-8", xml_declaration=True)
            modified[drawing_rels_part] = ET.tostring(drawing_rels_root, encoding="utf-8", xml_declaration=True)

    if not modified:
        return

    temp_path = workbook_path.with_suffix(workbook_path.suffix + ".tmp")
    with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name, data in archive.items():
            zout.writestr(name, modified.get(name, data))

    temp_path.replace(workbook_path)


def _autofit_openpyxl_columns(ws, *, min_width: int = 8, max_width: int = 70, padding: int = 2) -> None:
    if ws.max_column < 1 or ws.max_row < 1:
        return

    for col_idx, col_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
        start=1,
    ):
        max_len = 0
        rotated_header = False
        header_text = ""
        for cell in col_cells:
            if cell.value is None:
                continue
            if cell.row == 1:
                alignment = getattr(cell, "alignment", None)
                text_rotation = getattr(alignment, "textRotation", 0) if alignment is not None else 0
                if text_rotation not in (0, None):
                    rotated_header = True
                    header_text = str(cell.value)
                    continue
            text = str(cell.value)
            if "\n" in text:
                text = max(text.splitlines(), key=len)
            max_len = max(max_len, len(text))

        width = max(min_width, min(max_width, max_len + padding))
        if rotated_header:
            compact_width = max(6, min(12, max_len + 2))
            if max_len == 0 and header_text:
                compact_width = max(6, min(12, len(header_text.split()[0]) + 2))
            width = min(width, compact_width)
        ws.column_dimensions[_excel_col_letter(col_idx)].width = width


def _progress_percent(current: int, total: int) -> int:
    if total <= 0:
        return 100
    bounded = min(max(current, 0), total)
    return int(round(100.0 * bounded / total))


def _print_progress(stage: str, current: int, total: int, detail: str | None = None) -> None:
    pct = _progress_percent(current, total)
    suffix = "" if not detail else f" | {detail}"
    print(f"[{stage}] {pct:3d}% ({current}/{total}){suffix}")


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    text = str(value).strip().strip('"')
    if text == "" or text.lower() in {"nan", "na", "none"}:
        return None
    try:
        return float(text)
    except ValueError:
        # Handle comma decimals.
        try:
            return float(text.replace(",", "."))
        except ValueError:
            return None


def _module_from_test_name(test_name: str) -> str:
    if not test_name:
        return ""
    s = str(test_name).strip()
    if len(s) < 4:
        return s.upper()
    return s[:4].upper()


@dataclass(frozen=True)
class FlatFileMeta:
    header: list[str]
    numeric_test_cols: list[str]
    data_start_line_index: int  # 0-based line index where unit data begins
    meta_rows: dict[str, dict[str, str]]  # row_name -> {col_name -> raw string}


def scan_flat_file_meta(
    file_path: Path,
    *,
    encoding: str = DEFAULT_ENCODING,
    delimiter: str = DELIMITER,
    needed_meta_rows: Iterable[str] = ("Test Name", "Low", "High", "Unit", "Cpk", "Yield", "Mean", "Stddev"),
    max_scan_lines: int = 200,
) -> FlatFileMeta:
    needed = {r.strip() for r in needed_meta_rows}
    meta_rows: dict[str, dict[str, str]] = {}

    with file_path.open("r", encoding=encoding, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"Empty file: {file_path}")

        header = [h.strip() for h in header]
        numeric_test_cols = [h for h in header if h.strip().isdigit()]
        if not numeric_test_cols:
            raise ValueError("Could not find numeric test columns in header")

        data_start_line_index = 1
        for line_idx, row in enumerate(reader, start=1):
            if line_idx >= max_scan_lines:
                break
            if not row:
                continue
            key = (row[0] or "").strip().strip('"')
            if key.isdigit():
                data_start_line_index = line_idx
                break

            if key in needed:
                row_map: dict[str, str] = {}
                # Zip may truncate; that's ok.
                for col_name, cell in zip(header, row, strict=False):
                    if col_name in numeric_test_cols:
                        row_map[col_name] = (cell or "").strip()
                meta_rows[key] = row_map

        else:
            # If we didn't break on a unit row, we still set start line to after scanned block.
            data_start_line_index = max_scan_lines

    return FlatFileMeta(
        header=header,
        numeric_test_cols=numeric_test_cols,
        data_start_line_index=data_start_line_index,
        meta_rows=meta_rows,
    )


def _is_analysis_input_csv_path(file_path: Path) -> bool:
    name = file_path.name.lower()
    return not name.endswith("_dtr_records.csv")


def _collect_analysis_csv_paths(
    input_folder: Path,
    *,
    single_file: str | None,
    max_files: int | None,
) -> list[Path]:
    if single_file:
        csv_paths = [input_folder / single_file]
    else:
        csv_paths = sorted(
            [p for p in input_folder.glob("*.csv") if p.is_file() and _is_analysis_input_csv_path(p)]
        )
    if max_files is not None:
        csv_paths = csv_paths[:max_files]
    return csv_paths


def _mad(values):
    import numpy as np

    med = np.nanmedian(values)
    return float(np.nanmedian(np.abs(values - med)))


def _robust_sigma(values) -> float:
    import numpy as np

    m = _mad(values)
    if m > 0:
        return 1.4826 * m
    # Fallback to std if MAD collapses.
    return float(np.nanstd(values, ddof=1))


def _count_hist_peaks(values) -> int:
    """Very lightweight multimodality heuristic based on histogram peaks."""
    import numpy as np

    v = values[np.isfinite(values)]
    if v.size < 80:
        return 1

    rounded = np.round(v.astype(float), 9)
    unique_vals, unique_counts = np.unique(rounded, return_counts=True)
    if 2 <= unique_vals.size <= 10:
        significant_mask = unique_counts >= max(3, int(math.ceil(0.10 * v.size)))
        if int(np.count_nonzero(significant_mask)) >= 2:
            return int(np.count_nonzero(significant_mask))

    q25, q75 = np.percentile(v, [25, 75])
    iqr = q75 - q25
    if iqr <= 0:
        return 1
    bin_width = 2 * iqr / (v.size ** (1 / 3))
    if not np.isfinite(bin_width) or bin_width <= 0:
        return 1
    bins = max(10, min(80, int((v.max() - v.min()) / bin_width) + 1))
    hist, _ = np.histogram(v, bins=bins)
    if hist.size < 5:
        return 1

    # Smooth with a small moving average.
    kernel = np.array([1, 2, 3, 2, 1], dtype=float)
    kernel /= kernel.sum()
    smooth = np.convolve(hist.astype(float), kernel, mode="same")
    # Local maxima with a basic prominence threshold.
    threshold = max(5.0, 0.10 * float(smooth.max()))
    peaks = 0
    for i in range(1, len(smooth) - 1):
        if smooth[i] > smooth[i - 1] and smooth[i] > smooth[i + 1] and smooth[i] >= threshold:
            peaks += 1
    return max(1, peaks)


def _cdf_plot_png(
    values,
    *,
    title: str,
    out_path: Path,
    low_limit: float | None = None,
    high_limit: float | None = None,
    proposed_l6: float | None = None,
    proposed_u6: float | None = None,
    proposed_l12: float | None = None,
    proposed_u12: float | None = None,
    zoom_to_limits: bool = False,
) -> None:
    import pandas as pd
    import numpy as np

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        # Plotting is optional; if matplotlib isn't available just skip.
        return

    v_raw = pd.to_numeric(values, errors="coerce").to_numpy(dtype=float)
    finite_mask = np.isfinite(v_raw)
    v = v_raw[finite_mask]
    if v.size == 0:
        return

    low = -np.inf if low_limit is None else float(low_limit)
    high = np.inf if high_limit is None else float(high_limit)
    has_spec = low_limit is not None or high_limit is not None
    fail_mask = (v < low) | (v > high) if has_spec else np.zeros(v.size, dtype=bool)
    n_fail = int(np.count_nonzero(fail_mask))

    order = np.argsort(v)
    v = v[order]
    fail_mask = fail_mask[order]
    y = np.arange(1, v.size + 1) / v.size

    mean_v = float(np.mean(v))
    median_v = float(np.median(v))

    fig, ax = plt.subplots(figsize=(7.0, 4.0), dpi=140)

    # CDF points (not a continuous line), with failing chips highlighted.
    pass_mask = ~fail_mask
    if np.any(pass_mask):
        ax.plot(
            v[pass_mask],
            y[pass_mask],
            linestyle="None",
            marker=".",
            markersize=3.0,
            alpha=0.80,
            color="#1F77B4",
            label=f"Pass chips={int(np.count_nonzero(pass_mask))}",
        )
    if np.any(fail_mask):
        ax.plot(
            v[fail_mask],
            y[fail_mask],
            linestyle="None",
            marker="o",
            markersize=4.4,
            alpha=0.95,
            markerfacecolor="#D62728",
            markeredgecolor="#D62728",
            label=f"Fail chips={n_fail}",
        )
    elif not np.any(pass_mask):
        ax.plot(v, y, linestyle="None", marker=".", markersize=3.0, alpha=0.85, label="Data")

    # Spec limits (red)
    if low_limit is not None and np.isfinite(low_limit):
        ax.axvline(float(low_limit), color="#D62728", linestyle="-", linewidth=1.6, label=f"LTL={_fmt_num(float(low_limit))}")
    if high_limit is not None and np.isfinite(high_limit):
        ax.axvline(float(high_limit), color="#D62728", linestyle="-", linewidth=1.6, label=f"UTL={_fmt_num(float(high_limit))}")

    # Proposed 6σ/12σ limits for Cpk-issue tests (orange/purple)
    if proposed_l6 is not None and np.isfinite(proposed_l6):
        ax.axvline(float(proposed_l6), color="#FF7F0E", linestyle="--", linewidth=1.2, label=f"LTL 6s={_fmt_1dp(float(proposed_l6))}")
    if proposed_u6 is not None and np.isfinite(proposed_u6):
        ax.axvline(float(proposed_u6), color="#FF7F0E", linestyle="--", linewidth=1.2, label=f"UTL 6s={_fmt_1dp(float(proposed_u6))}")
    if proposed_l12 is not None and np.isfinite(proposed_l12):
        ax.axvline(float(proposed_l12), color="#9467BD", linestyle=":", linewidth=1.2, label=f"LTL 12s={_fmt_1dp(float(proposed_l12))}")
    if proposed_u12 is not None and np.isfinite(proposed_u12):
        ax.axvline(float(proposed_u12), color="#9467BD", linestyle=":", linewidth=1.2, label=f"UTL 12s={_fmt_1dp(float(proposed_u12))}")

    if zoom_to_limits:
        zoom_limits = _resolve_main_distribution_zoom_limits(v)
        if zoom_limits is not None:
            ax.set_xlim(*zoom_limits)

    ax.grid(True, alpha=0.3)
    if has_spec:
        ax.set_title(title + f"\nFail chips: {n_fail}/{v.size}")
    else:
        ax.set_title(title + f"\nFail chips: N/A (no spec limits)")
    ax.set_xlabel("Value")
    ax.set_ylabel("CDF")
    ax.legend(loc="best", fontsize=8, framealpha=0.9)
    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, format="png")
    plt.close(fig)


def _cdf_plot_png_pair(
    values,
    *,
    title: str,
    out_path: Path,
    zoomed_out_path: Path | None = None,
    low_limit: float | None = None,
    high_limit: float | None = None,
    proposed_l6: float | None = None,
    proposed_u6: float | None = None,
    proposed_l12: float | None = None,
    proposed_u12: float | None = None,
) -> None:
    import numpy as np

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        return

    v_raw = np.asarray(values, dtype=float)
    v = v_raw[np.isfinite(v_raw)]
    if v.size == 0:
        return

    low = -np.inf if low_limit is None else float(low_limit)
    high = np.inf if high_limit is None else float(high_limit)
    has_spec = low_limit is not None or high_limit is not None
    fail_mask = (v < low) | (v > high) if has_spec else np.zeros(v.size, dtype=bool)
    order = np.argsort(v)
    v = v[order]
    fail_mask = fail_mask[order]
    y = np.arange(1, v.size + 1) / v.size
    n_fail = int(np.count_nonzero(fail_mask))

    fig, ax = plt.subplots(figsize=(7.0, 4.0), dpi=140)

    pass_mask = ~fail_mask
    if np.any(pass_mask):
        ax.plot(
            v[pass_mask],
            y[pass_mask],
            linestyle="None",
            marker=".",
            markersize=3.0,
            alpha=0.80,
            color="#1F77B4",
            label=f"Pass chips={int(np.count_nonzero(pass_mask))}",
        )
    if np.any(fail_mask):
        ax.plot(
            v[fail_mask],
            y[fail_mask],
            linestyle="None",
            marker="o",
            markersize=4.4,
            alpha=0.95,
            markerfacecolor="#D62728",
            markeredgecolor="#D62728",
            label=f"Fail chips={n_fail}",
        )
    elif not np.any(pass_mask):
        ax.plot(v, y, linestyle="None", marker=".", markersize=3.0, alpha=0.85, label="Data")

    if low_limit is not None and np.isfinite(low_limit):
        ax.axvline(float(low_limit), color="#D62728", linestyle="-", linewidth=1.6, label=f"LTL={_fmt_num(float(low_limit))}")
    if high_limit is not None and np.isfinite(high_limit):
        ax.axvline(float(high_limit), color="#D62728", linestyle="-", linewidth=1.6, label=f"UTL={_fmt_num(float(high_limit))}")
    if proposed_l6 is not None and np.isfinite(proposed_l6):
        ax.axvline(float(proposed_l6), color="#FF7F0E", linestyle="--", linewidth=1.2, label=f"LTL 6s={_fmt_1dp(float(proposed_l6))}")
    if proposed_u6 is not None and np.isfinite(proposed_u6):
        ax.axvline(float(proposed_u6), color="#FF7F0E", linestyle="--", linewidth=1.2, label=f"UTL 6s={_fmt_1dp(float(proposed_u6))}")
    if proposed_l12 is not None and np.isfinite(proposed_l12):
        ax.axvline(float(proposed_l12), color="#9467BD", linestyle=":", linewidth=1.2, label=f"LTL 12s={_fmt_1dp(float(proposed_l12))}")
    if proposed_u12 is not None and np.isfinite(proposed_u12):
        ax.axvline(float(proposed_u12), color="#9467BD", linestyle=":", linewidth=1.2, label=f"UTL 12s={_fmt_1dp(float(proposed_u12))}")

    ax.grid(True, alpha=0.3)
    if has_spec:
        ax.set_title(title + f"\nFail chips: {n_fail}/{v.size}")
    else:
        ax.set_title(title + f"\nFail chips: N/A (no spec limits)")
    ax.set_xlabel("Value")
    ax.set_ylabel("CDF")
    ax.legend(loc="best", fontsize=8, framealpha=0.9)
    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, format="png")

    if zoomed_out_path is not None:
        zoom_limits = _resolve_main_distribution_zoom_limits(v)
        if zoom_limits is not None:
            ax.set_xlim(*zoom_limits)
        zoomed_out_path.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(zoomed_out_path, format="png")

    plt.close(fig)


def _cdf_plot_by_site_png(
    values,
    *,
    meta_cols,
    title: str,
    out_path: Path,
    low_limit: float | None = None,
    high_limit: float | None = None,
    zoom_to_limits: bool = False,
) -> None:
    import numpy as np
    import pandas as pd

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        return

    if meta_cols is None or "SITE_NUM" not in getattr(meta_cols, "columns", []):
        return

    df = pd.DataFrame(
        {
            "v": pd.to_numeric(values, errors="coerce"),
            "SITE_NUM": pd.to_numeric(meta_cols["SITE_NUM"], errors="coerce"),
        }
    ).dropna(subset=["v", "SITE_NUM"])
    if df.empty or df["SITE_NUM"].nunique() < 2:
        return

    fig, ax = plt.subplots(figsize=(7.2, 4.2), dpi=140)
    cmap = plt.get_cmap("tab10")
    for idx, (site_num, group) in enumerate(df.groupby("SITE_NUM")):
        site_values = np.sort(group["v"].to_numpy(dtype=float))
        if site_values.size == 0:
            continue
        y = np.arange(1, site_values.size + 1) / site_values.size
        site_label = _format_site_identifier(site_num)
        ax.scatter(
            site_values,
            y,
            s=14,
            alpha=0.82,
            color=cmap(idx % 10),
            label=f"Site {site_label} (n={site_values.size})",
        )

    if low_limit is not None and math.isfinite(float(low_limit)):
        ax.axvline(float(low_limit), color="#D62728", linestyle="-", linewidth=1.4, label=f"LTL={_fmt_num(float(low_limit))}")
    if high_limit is not None and math.isfinite(float(high_limit)):
        ax.axvline(float(high_limit), color="#D62728", linestyle="-", linewidth=1.4, label=f"UTL={_fmt_num(float(high_limit))}")

    if zoom_to_limits:
        zoom_limits = _resolve_main_distribution_zoom_limits_by_groups(
            [group["v"].to_numpy(dtype=float) for _, group in df.groupby("SITE_NUM")]
        )
        if zoom_limits is not None:
            ax.set_xlim(*zoom_limits)

    ax.grid(True, alpha=0.3)
    ax.set_title(title + "\nDistribution grouped by site")
    ax.set_xlabel("Value")
    ax.set_ylabel("CDF")
    ax.legend(loc="best", fontsize=8, framealpha=0.9)
    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, format="png")
    plt.close(fig)


def _prepare_wafer_map_frame(values, *, meta_cols):
    import pandas as pd
    import numpy as np

    if meta_cols is None:
        return None, None, None, None
    if not all(c in getattr(meta_cols, "columns", []) for c in ("X", "Y")):
        return None, None, None, None

    v = pd.to_numeric(values, errors="coerce")
    x = pd.to_numeric(meta_cols["X"], errors="coerce")
    y = pd.to_numeric(meta_cols["Y"], errors="coerce")
    wafer = _normalize_wafer_ids(meta_cols["WAFER"]) if "WAFER" in meta_cols.columns else None

    df = pd.DataFrame({"v": v, "X": x, "Y": y})
    if wafer is not None:
        df["WAFER"] = wafer
    for optional_col in ("CHIP_ID", "SITE_NUM", "PF", "FIRST_FAIL_TEST"):
        if optional_col in getattr(meta_cols, "columns", []):
            df[optional_col] = meta_cols[optional_col]

    df = df.dropna(subset=["v", "X", "Y"]).copy()
    if df.empty:
        return None, None, None, None

    if "WAFER" in df.columns and df["WAFER"].notna().any():
        counts = df.dropna(subset=["WAFER"]).groupby("WAFER")["v"].size().sort_values(ascending=False)
        wafers = [str(w) for w in counts.index.tolist()]
    else:
        wafers = ["ALL"]
        df["WAFER"] = "ALL"

    wafers = wafers[:6]
    all_v = df["v"].to_numpy(dtype=float)
    vmin = float(np.nanpercentile(all_v, 1))
    vmax = float(np.nanpercentile(all_v, 99))
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmin >= vmax:
        vmin = float(np.nanmin(df["v"]))
        vmax = float(np.nanmax(df["v"]))

    return df, wafers, vmin, vmax


def _build_wafer_grid(d):
    import numpy as np

    grouped = d.groupby(["Y", "X"], as_index=False).agg(v=("v", "median"), FAIL=("FAIL", "max"))

    xv = grouped["X"].to_numpy(dtype=float)
    yv = grouped["Y"].to_numpy(dtype=float)
    vv = grouped["v"].to_numpy(dtype=float)
    fails = grouped["FAIL"].to_numpy(dtype=bool)

    x_unique = np.sort(np.unique(xv))
    y_unique = np.sort(np.unique(yv))
    x_index = {float(x): idx for idx, x in enumerate(x_unique)}
    y_index = {float(y): idx for idx, y in enumerate(y_unique)}

    grid = np.full((len(y_unique), len(x_unique)), np.nan, dtype=float)
    fail_mask = np.zeros((len(y_unique), len(x_unique)), dtype=bool)
    for x_val, y_val, v_val, fail_val in zip(xv, yv, vv, fails, strict=False):
        xi = x_index[float(x_val)]
        yi = y_index[float(y_val)]
        grid[yi, xi] = float(v_val)
        fail_mask[yi, xi] = bool(fail_val)

    return x_unique, y_unique, grid, fail_mask, int(np.count_nonzero(fails))


def _resolve_main_distribution_zoom_limits(values) -> tuple[float, float] | None:
    import numpy as np

    finite = np.asarray(values, dtype=float)
    finite = finite[np.isfinite(finite)]
    if finite.size == 0:
        return None

    if finite.size <= 8:
        left = float(np.min(finite))
        right = float(np.max(finite))
    else:
        q01, q25, q50, q75, q99 = np.nanpercentile(finite, [1, 25, 50, 75, 99])
        iqr = float(q75 - q25)
        if np.isfinite(iqr) and iqr > 0:
            left = max(float(q01), float(q25 - 1.5 * iqr))
            right = min(float(q99), float(q75 + 1.5 * iqr))
        else:
            mad = float(np.nanmedian(np.abs(finite - q50)))
            if np.isfinite(mad) and mad > 0:
                robust_sigma = 1.4826 * mad
                left = max(float(q01), float(q50 - 6.0 * robust_sigma))
                right = min(float(q99), float(q50 + 6.0 * robust_sigma))
            else:
                left = float(np.min(finite))
                right = float(np.max(finite))

    if not np.isfinite(left) or not np.isfinite(right):
        left = float(np.min(finite))
        right = float(np.max(finite))

    span = right - left
    if not np.isfinite(span) or span <= 0:
        anchor = float(np.median(finite))
        span = max(abs(anchor) * 0.1, 1.0)
        left = anchor - span / 2.0
        right = anchor + span / 2.0

    pad = max(span * 0.06, 1e-9)
    return left - pad, right + pad


def _resolve_main_distribution_zoom_limits_by_groups(group_values: Iterable[Any]) -> tuple[float, float] | None:
    limits: list[tuple[float, float]] = []
    for values in group_values:
        current = _resolve_main_distribution_zoom_limits(values)
        if current is not None:
            limits.append(current)

    if not limits:
        return None

    left = min(item[0] for item in limits)
    right = max(item[1] for item in limits)
    if not math.isfinite(left) or not math.isfinite(right):
        return None
    if right <= left:
        pad = max(abs(left) * 0.05, 1.0)
        return left - pad, right + pad
    return left, right


def _build_wafer_map_title(
    title: str,
    *,
    low_limit: float | None,
    high_limit: float | None,
    unit: str | None,
    median_v: float | None,
) -> str:
    unit_txt = unit.strip() if isinstance(unit, str) and unit.strip() else "-"
    ltl_txt = _fmt_num(float(low_limit)) if low_limit is not None and math.isfinite(float(low_limit)) else "N/A"
    utl_txt = _fmt_num(float(high_limit)) if high_limit is not None and math.isfinite(float(high_limit)) else "N/A"
    median_txt = _fmt_num(float(median_v)) if median_v is not None and math.isfinite(float(median_v)) else "N/A"
    return title + "\n" + f"LTL={ltl_txt}; UTL={utl_txt}; Unit={unit_txt}; median={median_txt}"


def _fmt_wafer_coordinate(value: float) -> str:
    value_f = float(value)
    if math.isfinite(value_f) and value_f.is_integer():
        return str(int(value_f))
    return _fmt_num(value_f)


def _summarize_failing_chip_coordinates(
    grouped,
    *,
    low_limit: float | None,
    high_limit: float | None,
    max_items: int = WAFER_MAP_FAIL_COORDS_MAX_ITEMS,
    wrap_width: int = WAFER_MAP_FAIL_COORDS_WRAP_WIDTH_NARROW,
) -> str | None:
    import textwrap

    if grouped is None or grouped.empty or "FAIL" not in grouped.columns:
        return None

    failed = grouped[grouped["FAIL"].astype(bool)].copy()
    if failed.empty:
        return None

    failed = failed.sort_values(["X", "Y"], ascending=[True, True])

    total_fail = int(failed.shape[0])
    selected = failed.head(max_items)
    coords = [
        f"X{_fmt_wafer_coordinate(x_val)}-Y{_fmt_wafer_coordinate(y_val)}"
        for x_val, y_val in zip(selected["X"], selected["Y"], strict=False)
    ]
    prefix = f"Fail coords ({len(coords)}/{total_fail}): " if total_fail > len(coords) else f"Fail coords ({total_fail}): "
    return textwrap.fill(prefix + ", ".join(coords), width=wrap_width, subsequent_indent="  ")


def _wafer_map_png(
    values,
    *,
    meta_cols,
    title: str,
    out_path: Path,
    low_limit: float | None = None,
    high_limit: float | None = None,
    unit: str | None = None,
    median_v: float | None = None,
) -> None:
    """Create a PNG wafer map with compact failing-chip coordinate annotations."""
    import numpy as np

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.colors as mcolors
    except Exception:
        return

    df, wafers, vmin, vmax = _prepare_wafer_map_frame(values, meta_cols=meta_cols)
    if df is None or wafers is None:
        return

    # Vivid multi-color gradient that makes extremes pop.
    try:
        cmap = plt.get_cmap("turbo")
    except Exception:
        cmap = plt.get_cmap("viridis")

    norm = mcolors.Normalize(vmin=vmin, vmax=vmax, clip=True)
    wafermap_scale = 1.0

    n = len(wafers)
    ncols = 3 if n >= 3 else n
    nrows = int(math.ceil(n / max(1, ncols)))

    low = -np.inf if low_limit is None else float(low_limit)
    high = np.inf if high_limit is None else float(high_limit)

    wafer_panels: list[dict[str, Any] | None] = []
    max_annotation_lines = 0
    fail_summary_wrap_width = (
        WAFER_MAP_FAIL_COORDS_WRAP_WIDTH_WIDE if ncols <= 2 else WAFER_MAP_FAIL_COORDS_WRAP_WIDTH_NARROW
    )
    for w in wafers:
        d = df[df["WAFER"].astype(str) == str(w)].copy()
        if d.empty:
            wafer_panels.append(None)
            continue

        d["FAIL"] = (d["v"].to_numpy(dtype=float) < low) | (d["v"].to_numpy(dtype=float) > high)
        grouped = d.groupby(["Y", "X"], as_index=False).agg(v=("v", "median"), FAIL=("FAIL", "max"))
        x_unique, y_unique, grid, fail_mask, n_fail = _build_wafer_grid(d)
        fail_summary = _summarize_failing_chip_coordinates(
            grouped,
            low_limit=low_limit,
            high_limit=high_limit,
            max_items=WAFER_MAP_FAIL_COORDS_MAX_ITEMS,
            wrap_width=fail_summary_wrap_width,
        )
        if fail_summary:
            max_annotation_lines = max(max_annotation_lines, fail_summary.count("\n") + 1)

        wafer_panels.append(
            {
                "wafer": str(w),
                "x_unique": x_unique,
                "y_unique": y_unique,
                "grid": grid,
                "fail_mask": fail_mask,
                "n_fail": n_fail,
                "n_points": len(grouped),
                "fail_summary": fail_summary,
            }
        )

    base_fig_w = (10.0 if n <= 2 else 12.5) * wafermap_scale
    base_fig_h = (6.8 if n <= 2 else max(7.0, 4.8 * nrows)) * wafermap_scale
    annotation_height_in = (0.16 * max_annotation_lines + (0.22 if max_annotation_lines else 0.0)) * wafermap_scale
    fig_w = base_fig_w
    fig_h = base_fig_h + (annotation_height_in * max(1, nrows))
    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(fig_w, fig_h), dpi=140)
    if not isinstance(axes, np.ndarray):
        axes = np.array([axes])
    axes = axes.ravel()

    # Reserve a right margin for the colorbar so it never covers the wafer maps.
    annotation_space = annotation_height_in / max(fig_h, 1.0)
    bottom_margin = 0.08 + (annotation_space * (1.15 if nrows == 1 else 0.75))
    hspace = 0.32 + min(0.45, annotation_space * (4.0 if nrows > 1 else 2.1))
    fig.subplots_adjust(left=0.06, right=0.86, bottom=bottom_margin, top=0.90, wspace=0.18, hspace=hspace)

    mappable = None
    for ax, panel in zip(axes, wafer_panels, strict=False):
        if panel is None:
            ax.axis("off")
            continue

        x_unique = panel["x_unique"]
        y_unique = panel["y_unique"]
        grid = panel["grid"]
        fail_mask = panel["fail_mask"]
        n_fail = panel["n_fail"]

        x_min = float(x_unique.min())
        x_max = float(x_unique.max())
        y_min = float(y_unique.min())
        y_max = float(y_unique.max())

        if x_min == x_max:
            x_min -= 0.5
            x_max += 0.5
        if y_min == y_max:
            y_min -= 0.5
            y_max += 0.5

        image = ax.imshow(
            np.ma.masked_invalid(grid),
            origin="lower",
            interpolation="nearest",
            cmap=cmap,
            norm=norm,
            aspect="auto",
            extent=(x_min, x_max, y_min, y_max),
        )
        mappable = image

        if np.any(fail_mask):
            fail_overlay = np.zeros((*fail_mask.shape, 4), dtype=float)
            fail_overlay[..., 0] = 0.84
            fail_overlay[..., 1] = 0.15
            fail_overlay[..., 2] = 0.16
            fail_overlay[..., 3] = fail_mask.astype(float) * 0.45
            ax.imshow(
                fail_overlay,
                origin="lower",
                interpolation="nearest",
                aspect="auto",
                extent=(x_min, x_max, y_min, y_max),
            )
            fail_y_idx, fail_x_idx = np.where(fail_mask)
            fail_marker_size = max(80.0, 550.0 / math.sqrt(max(1, fail_y_idx.size)))
            ax.scatter(
                x_unique[fail_x_idx],
                y_unique[fail_y_idx],
                facecolors="none",
                edgecolors="#D62728",
                linewidths=1.8,
                s=fail_marker_size,
                marker="s",
            )

        ax.set_xlim(x_min, x_max)
        ax.set_ylim(y_min, y_max)

        ax.set_title(f"WAFER={panel['wafer']}  N={panel['n_points']}  fails={n_fail}", fontsize=10 * wafermap_scale)
        ax.set_aspect("auto")
        ax.grid(True, alpha=0.12)
        ax.tick_params(labelsize=9 * wafermap_scale)
        ax.set_xlabel("X", fontsize=9 * wafermap_scale)
        ax.set_ylabel("Y", fontsize=9 * wafermap_scale)

        fail_summary = panel["fail_summary"]
        if fail_summary:
            ax.text(
                0.01,
                -0.16,
                fail_summary,
                transform=ax.transAxes,
                ha="left",
                va="top",
                fontsize=7.2 * wafermap_scale,
                linespacing=1.2,
                clip_on=False,
                bbox={
                    "boxstyle": "round,pad=0.25",
                    "facecolor": "white",
                    "edgecolor": "#D62728",
                    "alpha": 0.88,
                },
            )

    # Hide any unused axes.
    for ax in axes[len(wafers) :]:
        ax.axis("off")

    if mappable is not None:
        # Dedicated axis for colorbar (no overlap with subplot area).
        cax = fig.add_axes([0.87, 0.18, 0.03, 0.62])
        cbar = fig.colorbar(mappable, cax=cax)
        cbar.ax.tick_params(labelsize=8 * wafermap_scale)
        cbar.set_label("Value", fontsize=9 * wafermap_scale)

    fig.suptitle(
        _build_wafer_map_title(
            title,
            low_limit=low_limit,
            high_limit=high_limit,
            unit=unit,
            median_v=median_v,
        ),
        fontsize=10 * wafermap_scale,
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, format="png")
    plt.close(fig)


def _parse_filename_wafer_signature(file_name: str) -> str | None:
    m = re.search(r"\b(S11P|S21P|S31P|B11P|B21P)\b", file_name, flags=re.IGNORECASE)
    if not m:
        return None
    return m.group(1).upper()


def _parse_insertion_temperature_label(file_name: str) -> str:
    """Best-effort extraction of insertion temperature from file name."""
    upper = file_name.upper()
    if "S11P" in upper or "B11P" in upper or "HT" in upper or "Q3" in upper:
        return "Hot (135°C)"
    if "S21P" in upper or "Q2" in upper:
        return "Cold (-40°C)"
    if "S31P" in upper or "B21P" in upper or "RT" in upper or "Q1" in upper:
        return "Ambient (25°C)"
    return "Unknown"


def _supports_wafer_maps(file_name: str) -> bool:
    upper = str(file_name).upper()
    packaged_tokens = ("B11", "HT", "B21", "RT", "Q11", "Q21", "Q31")
    return not any(token in upper for token in packaged_tokens)


def _fmt_num(x: float) -> str:
    return f"{float(x):.6g}"


def _normalize_wafer_ids(values):
    """Extract numeric wafer identifiers from mixed WAFER labels."""
    import pandas as pd

    series = pd.Series(values, copy=False).astype("string").str.strip()
    series = series.mask(series.eq("") | series.str.lower().eq("nan"))
    extracted = series.str.extract(r"(\d+)", expand=False)
    numeric = pd.to_numeric(extracted, errors="coerce")
    normalized = numeric.astype("Int64").astype("string")
    return normalized.mask(normalized.isna())

def _fmt_1dp(x: float) -> str:
    return f"{float(x):.1f}"


METRIC_YIELD = "yield_fail"
METRIC_CPK_LOW = "cpk_low"
METRIC_CPK_HIGH = "cpk_high"
METRIC_SITE_DELTA = "site_to_site_delta"
METRIC_UNIQUE_VALUES = "unique_value_count"
METRIC_SKEWNESS = "skewness"
METRIC_MULTIMODALITY = "multimodality"
SKEWNESS_ABS_THRESHOLD = 1.0

METRIC_DISPLAY_ORDER: tuple[str, ...] = (
    METRIC_YIELD,
    METRIC_CPK_LOW,
    METRIC_CPK_HIGH,
    METRIC_SITE_DELTA,
    METRIC_UNIQUE_VALUES,
    METRIC_SKEWNESS,
    METRIC_MULTIMODALITY,
)

METRIC_PRIORITY: dict[str, str] = {
    METRIC_YIELD: "HIGH",
    METRIC_CPK_LOW: "HIGH",
    METRIC_CPK_HIGH: "MEDIUM",
    METRIC_SITE_DELTA: "MEDIUM",
    METRIC_UNIQUE_VALUES: "LOW",
    METRIC_SKEWNESS: "LOW",
    METRIC_MULTIMODALITY: "MEDIUM",
}

PRIORITY_RANK: dict[str, int] = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "OK": 3}
PRIORITY_FILL_COLORS: dict[str, str] = {
    "HIGH": "F4CCCC",
    "MEDIUM": "FFF2CC",
    "LOW": "DDEBF7",
    "OK": "C6E0B4",
}

_NON_ANALOG_UNIT_TOKENS = {
    "",
    "#",
    "bin",
    "bits",
    "bit",
    "bool",
    "boolean",
    "code",
    "codes",
    "count",
    "counts",
    "cnt",
    "index",
    "state",
    "status",
    "pass/fail",
    "passfail",
    "fail/pass",
    "pf",
    "logic",
}

MIN_SAMPLES_FOR_UNIQUE_VALUE_CHECK = 20
MIN_SAMPLES_PER_SITE_FOR_SITE_DELTA = 5
MIN_SITES_FOR_SITE_DELTA = 2


@dataclass(frozen=True)
class TestMetricAssessment:
    metric_keys: tuple[str, ...]
    status_text: str | None
    priority: str
    site_delta_sigma: float | None
    worst_site: str | None
    unique_value_count: int | None
    skewness: float | None
    is_analog_unit: bool
    peak_count: int
    multimodality_reason: str | None


def _format_site_identifier(site_value: Any) -> str:
    num = _to_float(site_value)
    if num is None:
        return str(site_value)
    return str(int(num)) if float(num).is_integer() else _fmt_num(num)


def _metric_label(
    metric_key: str,
    *,
    yield_threshold: float,
    cpk_low: float,
    cpk_high: float,
) -> str:
    if metric_key == METRIC_YIELD:
        return "Fails"
    if metric_key == METRIC_CPK_LOW:
        return f"Cpk<{cpk_low:g}"
    if metric_key == METRIC_CPK_HIGH:
        return f"Cpk>{cpk_high:g}"
    if metric_key == METRIC_SITE_DELTA:
        return "Site-to-Site Delta"
    if metric_key == METRIC_UNIQUE_VALUES:
        return "Unique Values"
    if metric_key == METRIC_SKEWNESS:
        return "Skewness"
    if metric_key == METRIC_MULTIMODALITY:
        return "Multimodality"
    return metric_key


def _sort_metric_keys(metric_keys: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for metric_key in METRIC_DISPLAY_ORDER:
        if metric_key in metric_keys and metric_key not in seen:
            ordered.append(metric_key)
            seen.add(metric_key)
    for metric_key in metric_keys:
        if metric_key not in seen:
            ordered.append(metric_key)
            seen.add(metric_key)
    return ordered


def _status_text_from_metric_keys(
    metric_keys: Iterable[str],
    *,
    yield_threshold: float,
    cpk_low: float,
    cpk_high: float,
) -> str | None:
    ordered = _sort_metric_keys(metric_keys)
    if not ordered:
        return None
    return " + ".join(
        _metric_label(
            metric_key,
            yield_threshold=yield_threshold,
            cpk_low=cpk_low,
            cpk_high=cpk_high,
        )
        for metric_key in ordered
    )


def _priority_from_metric_keys(metric_keys: Iterable[str]) -> str:
    priorities = [METRIC_PRIORITY.get(metric_key, "LOW") for metric_key in metric_keys]
    if not priorities:
        return "OK"
    return min(priorities, key=lambda item: PRIORITY_RANK.get(item, 999))


def _priority_fill_color(priority: str) -> str:
    return PRIORITY_FILL_COLORS.get(priority, PRIORITY_FILL_COLORS["OK"])


def _is_analog_unit(unit: str | None) -> bool:
    if unit is None:
        return False
    normalized = str(unit).strip().lower()
    if not normalized:
        return False
    return normalized.replace(" ", "") not in _NON_ANALOG_UNIT_TOKENS


def _site_delta_sigma(values, *, meta_cols) -> tuple[float | None, str | None]:
    import numpy as np
    import pandas as pd

    if "SITE_NUM" not in getattr(meta_cols, "columns", []):
        return None, None

    df_site = pd.DataFrame(
        {
            "v": pd.to_numeric(values, errors="coerce"),
            "SITE_NUM": pd.to_numeric(meta_cols["SITE_NUM"], errors="coerce"),
        }
    ).dropna(subset=["v", "SITE_NUM"])
    if df_site.empty:
        return None, None

    site_counts = df_site.groupby("SITE_NUM")["v"].size()
    eligible_sites = site_counts[site_counts >= MIN_SAMPLES_PER_SITE_FOR_SITE_DELTA].index
    if len(eligible_sites) < MIN_SITES_FOR_SITE_DELTA:
        return None, None
    df_site = df_site[df_site["SITE_NUM"].isin(eligible_sites)].copy()

    site_stats_rows: list[dict[str, float]] = []
    for site_num, group in df_site.groupby("SITE_NUM"):
        site_values = group["v"].to_numpy(dtype=float)
        if site_values.size == 0:
            continue
        site_stats_rows.append(
            {
                "SITE_NUM": float(site_num),
                "n": float(site_values.size),
                "center": float(np.median(site_values)),
                "mean": float(np.mean(site_values)),
                "sigma": float(_robust_sigma(site_values)),
            }
        )
    if len(site_stats_rows) < MIN_SITES_FOR_SITE_DELTA:
        return None, None

    site_stats = pd.DataFrame(site_stats_rows).sort_values("SITE_NUM").reset_index(drop=True)
    center_by_site = site_stats["center"].to_numpy(dtype=float)

    # Estimate site-to-site sensitivity against within-site residual spread rather than the
    # global spread, which can be inflated by the very site shifts we want to detect.
    site_center_series = df_site.groupby("SITE_NUM")["v"].transform("median")
    residuals = (df_site["v"] - site_center_series).to_numpy(dtype=float)
    sigma_candidates: list[float] = []
    residual_sigma = _robust_sigma(residuals)
    if np.isfinite(residual_sigma) and residual_sigma > 0.0:
        sigma_candidates.append(float(residual_sigma))

    site_sigmas = site_stats["sigma"].to_numpy(dtype=float)
    positive_site_sigmas = site_sigmas[np.isfinite(site_sigmas) & (site_sigmas > 0.0)]
    if positive_site_sigmas.size > 0:
        sigma_candidates.append(float(np.median(positive_site_sigmas)))
        sigma_candidates.append(float(np.sqrt(np.mean(np.square(positive_site_sigmas)))))

    sigma = min(sigma_candidates) if sigma_candidates else float(np.nanstd(residuals, ddof=1))

    global_center = float(np.median(df_site["v"].to_numpy(dtype=float)))
    deviations = np.abs(center_by_site - global_center)
    pairwise_delta = 0.0
    if center_by_site.size >= 2:
        pairwise_delta = float(np.nanmax(np.abs(center_by_site.reshape(-1, 1) - center_by_site.reshape(1, -1))))

    if np.isfinite(sigma) and sigma > 0.0:
        site_delta_sigma = max(float(np.nanmax(deviations) / sigma), float(pairwise_delta / (2.0 * sigma)))
    elif np.nanmax(deviations) > 0.0 or pairwise_delta > 0.0:
        site_delta_sigma = 99.0
    else:
        return None, None

    deviation_series = pd.Series(deviations, index=site_stats["SITE_NUM"].tolist()).dropna()
    if deviation_series.empty:
        return None, None

    worst_site = deviation_series.idxmax()
    return float(site_delta_sigma), _format_site_identifier(worst_site)


def _unique_value_count(values, *, unit: str | None) -> tuple[int | None, bool]:
    import pandas as pd

    is_analog = _is_analog_unit(unit)
    if not is_analog:
        return None, False

    series = pd.to_numeric(values, errors="coerce").dropna()
    if series.empty:
        return 0, True
    if series.size < MIN_SAMPLES_FOR_UNIQUE_VALUE_CHECK:
        return None, True
    rounded = series.round(9)
    unique_count = int(rounded.nunique(dropna=True))
    if unique_count <= 1:
        return unique_count, True

    sorted_unique = sorted(float(value) for value in rounded.dropna().unique())
    steps = [abs(curr - prev) for prev, curr in zip(sorted_unique, sorted_unique[1:], strict=False)]
    regular_quantization = False
    positive_steps = [step for step in steps if step > 0]
    if positive_steps:
        min_step = min(positive_steps)
        max_step = max(positive_steps)
        regular_quantization = min_step > 0 and max_step <= (1.5 * min_step)

    if regular_quantization and unique_count >= 5:
        return None, True
    return unique_count, True


def _skewness_value(values, *, is_analog: bool) -> float | None:
    import numpy as np

    if not is_analog:
        return None

    finite = np.asarray(values, dtype=float)
    finite = finite[np.isfinite(finite)]
    if finite.size < MIN_SAMPLES_FOR_UNIQUE_VALUE_CHECK:
        return None

    mean = float(np.mean(finite))
    centered = finite - mean
    m2 = float(np.mean(centered**2))
    if not np.isfinite(m2) or m2 <= 0.0:
        return 0.0

    m3 = float(np.mean(centered**3))
    skew = m3 / (m2 ** 1.5)
    if not np.isfinite(skew):
        return None
    return float(skew)


def _classify_wafer_process_signature(values, *, meta_cols) -> str | None:
    import numpy as np
    import pandas as pd

    if not all(col in getattr(meta_cols, "columns", []) for col in ("X", "Y")):
        return None

    df = pd.DataFrame(
        {
            "v": pd.to_numeric(values, errors="coerce"),
            "X": pd.to_numeric(meta_cols["X"], errors="coerce"),
            "Y": pd.to_numeric(meta_cols["Y"], errors="coerce"),
        }
    ).dropna()
    if df.shape[0] < 20:
        return None

    center_x = float(df["X"].mean())
    center_y = float(df["Y"].mean())
    radius = np.sqrt((df["X"] - center_x) ** 2 + (df["Y"] - center_y) ** 2)
    rho_r = _safe_spearman_correlation(df["v"], radius)
    if rho_r is not None and np.isfinite(float(rho_r)) and abs(float(rho_r)) >= 0.35:
        return "edge signature" if float(rho_r) > 0 else "center signature"

    inner = radius <= float(radius.quantile(1 / 3))
    outer = radius >= float(radius.quantile(2 / 3))
    middle = ~(inner | outer)
    sig = _robust_sigma(df["v"].to_numpy(dtype=float))
    if sig > 0 and inner.any() and middle.any() and outer.any():
        med_inner = float(df.loc[inner, "v"].median())
        med_middle = float(df.loc[middle, "v"].median())
        med_outer = float(df.loc[outer, "v"].median())
        if abs(med_middle - med_inner) >= sig and abs(med_middle - med_outer) >= sig:
            if (med_middle > med_inner and med_middle > med_outer) or (med_middle < med_inner and med_middle < med_outer):
                return "donut / ring signature"

    rho_x = _safe_spearman_correlation(df["v"], df["X"])
    rho_y = _safe_spearman_correlation(df["v"], df["Y"])
    if (
        rho_x is not None
        and np.isfinite(float(rho_x))
        and abs(float(rho_x)) >= 0.35
    ) or (
        rho_y is not None
        and np.isfinite(float(rho_y))
        and abs(float(rho_y)) >= 0.35
    ):
        return "cluster / gradient signature"

    return None


def _detect_coordinate_signature(values, *, meta_cols) -> str | None:
    import numpy as np
    import pandas as pd

    for axis in ("X", "Y"):
        if axis not in getattr(meta_cols, "columns", []):
            continue
        coord = pd.to_numeric(meta_cols[axis], errors="coerce")
        df_xy = pd.DataFrame({"v": pd.to_numeric(values, errors="coerce"), axis: coord}).dropna()
        if df_xy.shape[0] < 50:
            continue
        rho = _safe_spearman_correlation(df_xy["v"], df_xy[axis])
        if rho is not None and np.isfinite(rho) and abs(float(rho)) >= 0.30:
            return f"spearman(v,{axis})={float(rho):+.2f}"
    return None


def _is_go_nogo_test(*, unit: str | None, unique_value_count: int | None) -> bool:
    return _is_analog_unit(unit) and unique_value_count is not None and unique_value_count <= 2


def _infer_multimodality_reason(
    values,
    *,
    meta_cols,
    wafer_sig: str | None,
    site_delta_sigma: float | None,
    worst_site: str | None,
) -> str | None:
    import numpy as np
    import pandas as pd

    series = pd.to_numeric(values, errors="coerce")
    finite = series.dropna().to_numpy(dtype=float)
    finite = finite[np.isfinite(finite)]
    if finite.size == 0:
        return None

    reasons: list[str] = []
    if site_delta_sigma is not None and site_delta_sigma > 3.0 and worst_site is not None:
        reasons.append(f"site-to-site variation (site {worst_site}, Δ={site_delta_sigma:.2f}σ)")

    if "WAFER" in getattr(meta_cols, "columns", []):
        df_w = pd.DataFrame({"v": series, "WAFER": _normalize_wafer_ids(meta_cols["WAFER"])}).dropna(subset=["v"])
        if not df_w.empty and df_w["WAFER"].dropna().nunique() >= 2:
            sigma = _robust_sigma(df_w["v"].to_numpy(dtype=float))
            if sigma > 0:
                mean_by_wafer = df_w.dropna(subset=["WAFER"]).groupby("WAFER")["v"].mean()
                if not mean_by_wafer.empty:
                    global_mean = float(df_w["v"].mean())
                    wafer_shift = float(((mean_by_wafer - global_mean).abs() / sigma).max())
                    if np.isfinite(wafer_shift) and wafer_shift >= 2.0:
                        reasons.append("different wafers merged together")

    wafer_process = _classify_wafer_process_signature(series, meta_cols=meta_cols)
    if wafer_process:
        reasons.append(f"wafer process signature ({wafer_process})")
    elif wafer_sig and not reasons:
        reasons.append(f"file wafer signature {wafer_sig}")

    if not reasons:
        reasons.append("mixed populations; site/wafer root cause not dominant")
    return "; ".join(reasons[:2])


def _assess_test_metrics(
    *,
    series,
    meta_cols,
    unit: str | None,
    yield_pct: float | None,
    cpk: float | None,
    yield_threshold: float,
    cpk_low: float,
    cpk_high: float,
    wafer_sig: str | None,
) -> TestMetricAssessment:
    import numpy as np
    import pandas as pd

    metric_keys: list[str] = []

    if yield_pct is not None and yield_pct < yield_threshold:
        metric_keys.append(METRIC_YIELD)
    if cpk is not None and cpk < cpk_low:
        metric_keys.append(METRIC_CPK_LOW)
    if cpk is not None and cpk > cpk_high:
        metric_keys.append(METRIC_CPK_HIGH)

    numeric = pd.to_numeric(series, errors="coerce")
    finite = numeric.dropna().to_numpy(dtype=float)
    finite = finite[np.isfinite(finite)]

    site_delta, worst_site = _site_delta_sigma(numeric, meta_cols=meta_cols)
    if site_delta is not None and site_delta > 3.0:
        metric_keys.append(METRIC_SITE_DELTA)

    unique_values, is_analog = _unique_value_count(numeric, unit=unit)
    if is_analog and unique_values is not None and unique_values < 10:
        metric_keys.append(METRIC_UNIQUE_VALUES)

    skewness = _skewness_value(finite, is_analog=is_analog)
    if skewness is not None and abs(skewness) >= SKEWNESS_ABS_THRESHOLD:
        metric_keys.append(METRIC_SKEWNESS)

    multimodality_eligible = is_analog and not _is_go_nogo_test(unit=unit, unique_value_count=unique_values)
    peak_count = _count_hist_peaks(finite) if finite.size and multimodality_eligible else 1
    multimodality_reason = None
    if multimodality_eligible and peak_count >= 2:
        metric_keys.append(METRIC_MULTIMODALITY)
        multimodality_reason = _infer_multimodality_reason(
            numeric,
            meta_cols=meta_cols,
            wafer_sig=wafer_sig,
            site_delta_sigma=site_delta,
            worst_site=worst_site,
        )

    ordered_metric_keys = tuple(_sort_metric_keys(metric_keys))
    return TestMetricAssessment(
        metric_keys=ordered_metric_keys,
        status_text=_status_text_from_metric_keys(
            ordered_metric_keys,
            yield_threshold=yield_threshold,
            cpk_low=cpk_low,
            cpk_high=cpk_high,
        ),
        priority=_priority_from_metric_keys(ordered_metric_keys),
        site_delta_sigma=site_delta,
        worst_site=worst_site,
        unique_value_count=unique_values,
        skewness=skewness,
        is_analog_unit=is_analog,
        peak_count=peak_count,
        multimodality_reason=multimodality_reason,
    )


def _build_plot_title(
    test_name: str,
    test_col: str,
    temp_label: str,
    cpk: float | None,
    mean_v: float,
    median_v: float,
) -> str:
    first = f"{test_name} ({test_col}) | {temp_label}"
    cpk_txt = _fmt_num(float(cpk)) if cpk is not None and math.isfinite(float(cpk)) else "N/A"
    return first + "\n" + f"Cpk={cpk_txt}; mean={_fmt_num(mean_v)}; median={_fmt_num(median_v)}"


def _apply_module_group_row_colors(ws, *, module_col_header: str = "Module", first_data_row: int = 2) -> None:
    """Apply alternating light fills per contiguous module block (for filtering/readability)."""
    try:
        from openpyxl.styles import PatternFill
    except Exception:
        return

    if ws.max_row < first_data_row or ws.max_column < 1:
        return

    module_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=col_idx).value).strip() == module_col_header:
            module_col_idx = col_idx
            break
    if module_col_idx is None:
        return

    palette = [
        "D9E1F2",  # light blue
        "E2EFDA",  # light green
        "FFF2CC",  # light yellow
        "FCE4D6",  # light orange
        "EDEDED",  # light gray
    ]

    current_module = None
    color_idx = -1
    fill = None
    for row_idx in range(first_data_row, ws.max_row + 1):
        module_val = ws.cell(row=row_idx, column=module_col_idx).value
        module_str = "" if module_val is None else str(module_val)
        if module_str != current_module:
            current_module = module_str
            color_idx = (color_idx + 1) % len(palette)
            fill = PatternFill(patternType="solid", fgColor=palette[color_idx])

        if fill is None:
            continue
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _build_comment(
    *,
    series,
    meta_cols,
    outlier_mad_multiplier: float,
    low_limit: float | None,
    high_limit: float | None,
    wafer_sig: str | None,
    metric_assessment: TestMetricAssessment,
) -> str:
    import numpy as np
    import pandas as pd

    vals = pd.to_numeric(series, errors="coerce")
    finite = vals.dropna().to_numpy(dtype=float)
    finite = finite[np.isfinite(finite)]
    if finite.size == 0:
        return "No numeric data"

    med = float(np.median(finite))
    mad = _mad(finite)
    sig = _robust_sigma(finite)

    parts: list[str] = []

    if mad > 0:
        outliers = np.abs(finite - med) > (outlier_mad_multiplier * mad)
        n_out = int(outliers.sum())
        if n_out > 0:
            parts.append(f"Outliers: {n_out}/{finite.size} (>{outlier_mad_multiplier:g}×MAD)")

    # Instability / large spread heuristic.
    if sig > 0 and abs(med) > 0:
        robust_cv = sig / abs(med)
        if robust_cv >= 0.05:
            parts.append(f"Large spread (robust CV={robust_cv:.2%})")

    if metric_assessment.site_delta_sigma is not None and metric_assessment.site_delta_sigma > 3.0:
        if metric_assessment.worst_site is not None:
            parts.append(
                "Site-to-site delta "
                f"(site {metric_assessment.worst_site}, Δ={metric_assessment.site_delta_sigma:.2f}σ)"
            )
        else:
            parts.append(f"Site-to-site delta (Δ={metric_assessment.site_delta_sigma:.2f}σ)")

    if metric_assessment.is_analog_unit and metric_assessment.unique_value_count is not None:
        if metric_assessment.unique_value_count < 10:
            parts.append(f"Analog unique values low ({metric_assessment.unique_value_count})")

    if metric_assessment.skewness is not None and abs(metric_assessment.skewness) >= SKEWNESS_ABS_THRESHOLD:
        parts.append(f"Skewness high ({metric_assessment.skewness:.2f})")

    # Multi-modality heuristic.
    if metric_assessment.peak_count >= 2:
        parts.append(f"Possible multi-modal distribution (peaks≈{metric_assessment.peak_count})")
        if metric_assessment.multimodality_reason:
            parts.append(f"Multimodality reason: {metric_assessment.multimodality_reason}")

    # Site effect heuristic.
    if "SITE_NUM" in meta_cols.columns:
        df_site = pd.DataFrame({"v": vals, "SITE_NUM": pd.to_numeric(meta_cols["SITE_NUM"], errors="coerce")})
        df_site = df_site.dropna(subset=["v", "SITE_NUM"])
        if not df_site.empty and df_site["SITE_NUM"].nunique() >= 2:
            med_by_site = df_site.groupby("SITE_NUM")["v"].median().sort_index()
            rng = float(med_by_site.max() - med_by_site.min())
            if sig > 0 and rng / sig >= 3.0:
                worst = med_by_site.idxmax() if abs(med_by_site.max() - med) >= abs(med_by_site.min() - med) else med_by_site.idxmin()
                parts.append(f"Site effect (site medians range={rng:.4g}, worst site={int(worst)})")

    # Wafer signature heuristics.
    if wafer_sig:
        parts.append(f"File wafer signature: {wafer_sig}")

    if "WAFER" in meta_cols.columns:
        df_w = pd.DataFrame({"v": vals, "WAFER": _normalize_wafer_ids(meta_cols["WAFER"])})
        df_w = df_w.dropna(subset=["v"])  # keep WAFER even if blank
        wafer_series = df_w["WAFER"]
        wafers = wafer_series.dropna().unique()
        if wafers.size >= 2:
            med_by_wafer = df_w.dropna(subset=["WAFER"]).groupby("WAFER")["v"].median()
            rng = float(med_by_wafer.max() - med_by_wafer.min())
            if sig > 0 and rng / sig >= 3.0:
                parts.append("Wafer signature suspected (median shifts across wafers)")

    coordinate_signature = _detect_coordinate_signature(vals, meta_cols=meta_cols)
    if coordinate_signature:
        parts.append(f"Coordinate signature: {coordinate_signature}")

    # Fail clustering note.
    if low_limit is not None or high_limit is not None:
        low = -np.inf if low_limit is None else low_limit
        high = np.inf if high_limit is None else high_limit
        fails = (finite < low) | (finite > high)
        n_fail = int(fails.sum())
        if n_fail > 0:
            parts.append(f"Spec fails: {n_fail}/{finite.size}")

    return "; ".join(parts) if parts else "OK"


def _read_unit_data(
    file_path: Path,
    *,
    data_start_line_index: int,
    usecols: list[str],
    encoding: str = DEFAULT_ENCODING,
) -> Any:
    import pandas as pd

    df = pd.read_csv(
        file_path,
        sep=DELIMITER,
        encoding=encoding,
        low_memory=False,
        usecols=usecols,
        skiprows=range(1, data_start_line_index),
        memory_map=True,
    )
    df.columns = [c.strip() for c in df.columns]
    return df


def _limits_from_meta(meta: FlatFileMeta, test_col: str) -> tuple[float | None, float | None, str | None]:
    low_raw = meta.meta_rows.get("Low", {}).get(test_col)
    high_raw = meta.meta_rows.get("High", {}).get(test_col)
    unit_raw = meta.meta_rows.get("Unit", {}).get(test_col)
    return _to_float(low_raw), _to_float(high_raw), (unit_raw.strip() if unit_raw else None)


def _yield_cpk_from_meta(meta: FlatFileMeta, test_col: str) -> tuple[float | None, float | None]:
    y_raw = meta.meta_rows.get("Yield", {}).get(test_col)
    c_raw = meta.meta_rows.get("Cpk", {}).get(test_col)
    return _to_float(y_raw), _to_float(c_raw)


def _shorten_test_name(test_name: str | None) -> str:
    raw = "" if test_name is None else str(test_name)
    cleaned = raw.strip().strip('"')
    if not cleaned:
        return ""
    primary, _, _ = cleaned.partition("<>")
    return primary.strip()


def _test_name_from_meta(meta: FlatFileMeta, test_col: str) -> str:
    return _shorten_test_name(meta.meta_rows.get("Test Name", {}).get(test_col))


def _status_for_test(
    *,
    yield_pct: float | None,
    cpk: float | None,
    yield_threshold: float,
    cpk_low: float,
    cpk_high: float,
    site_to_site_delta: bool = False,
    unique_value_count_low: bool = False,
    skewness: bool = False,
    multimodality: bool = False,
) -> str | None:
    metric_keys: list[str] = []
    if yield_pct is not None and yield_pct < yield_threshold:
        metric_keys.append(METRIC_YIELD)
    if cpk is not None and cpk < cpk_low:
        metric_keys.append(METRIC_CPK_LOW)
    if cpk is not None and cpk > cpk_high:
        metric_keys.append(METRIC_CPK_HIGH)
    if site_to_site_delta:
        metric_keys.append(METRIC_SITE_DELTA)
    if unique_value_count_low:
        metric_keys.append(METRIC_UNIQUE_VALUES)
    if skewness:
        metric_keys.append(METRIC_SKEWNESS)
    if multimodality:
        metric_keys.append(METRIC_MULTIMODALITY)
    return _status_text_from_metric_keys(
        metric_keys,
        yield_threshold=yield_threshold,
        cpk_low=cpk_low,
        cpk_high=cpk_high,
    )


def _proposed_sigma_limits(values) -> tuple[float | None, float | None, float | None, float | None]:
    import numpy as np
    import pandas as pd

    v = pd.to_numeric(values, errors="coerce").dropna().to_numpy(dtype=float)
    v = v[np.isfinite(v)]
    if v.size == 0:
        return None, None, None, None
    med = float(np.median(v))
    sig = _robust_sigma(v)
    if not np.isfinite(sig) or sig <= 0:
        return med, med, med, med

    def _floor_1dp(x: float) -> float:
        return math.floor(x * 10.0) / 10.0

    def _ceil_1dp(x: float) -> float:
        return math.ceil(x * 10.0) / 10.0

    l6, u6 = med - 6.0 * sig, med + 6.0 * sig
    l12, u12 = med - 12.0 * sig, med + 12.0 * sig

    # Quantize to 1 decimal: conservative rounding (floor lows, ceil uppers).
    l6 = _floor_1dp(float(l6))
    u6 = _ceil_1dp(float(u6))
    l12 = _floor_1dp(float(l12))
    u12 = _ceil_1dp(float(u12))

    return l6, u6, l12, u12


def _add_overview_sheet(
    workbook,
    *,
    summary_entries: list[dict[str, Any]],
    modules: list[str],
    processed_files: list[str],
    output_folder: Path,
) -> None:
    from collections import Counter
    from datetime import datetime

    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Font, PatternFill

    if "Overview" in workbook.sheetnames:
        workbook.remove(workbook["Overview"])

    ws = workbook.create_sheet("Overview", 0)
    ws.sheet_properties.tabColor = "70AD47"

    title_fill = PatternFill(patternType="solid", fgColor="D9EAD3")
    section_fill = PatternFill(patternType="solid", fgColor="DDEBF7")
    link_font = Font(color="0000EE", underline="single")
    priority_fills = {
        priority: PatternFill(patternType="solid", fgColor=_priority_fill_color(priority))
        for priority in ("HIGH", "MEDIUM", "LOW", "OK")
    }

    ws["A1"] = "Test Data Analysis Overview"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].fill = title_fill

    metric_counts = Counter(
        metric_key
        for entry in summary_entries
        for metric_key in entry.get("metric_keys", [])
    )
    priority_counts = Counter(str(entry.get("priority") or "OK") for entry in summary_entries)
    files_with_issues = sorted({entry["file_name"] for entry in summary_entries})
    total_fail_chips = int(sum(int(entry.get("fail_chips") or 0) for entry in summary_entries))

    summary_rows = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Output folder", str(output_folder)),
        ("Files processed", len(processed_files)),
        ("Files with issues", len(files_with_issues)),
        ("Affected tests", len(summary_entries)),
        ("Total fail chips", total_fail_chips),
        ("High-priority tests", int(priority_counts.get("HIGH", 0))),
        ("Medium-priority tests", int(priority_counts.get("MEDIUM", 0))),
        ("Low-priority tests", int(priority_counts.get("LOW", 0))),
        ("Fails count", int(metric_counts.get(METRIC_YIELD, 0))),
        ("Cpk<1.67 count", int(metric_counts.get(METRIC_CPK_LOW, 0))),
        ("Cpk>20 count", int(metric_counts.get(METRIC_CPK_HIGH, 0))),
        ("Site-to-Site Delta count", int(metric_counts.get(METRIC_SITE_DELTA, 0))),
        ("Unique Values count", int(metric_counts.get(METRIC_UNIQUE_VALUES, 0))),
        ("Skewness count", int(metric_counts.get(METRIC_SKEWNESS, 0))),
        ("Multimodality count", int(metric_counts.get(METRIC_MULTIMODALITY, 0))),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, start=3):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    configured_modules: list[str] = []
    seen_modules: set[str] = set()
    for module_name in modules:
        normalized = str(module_name).strip().upper()
        if not normalized or normalized in seen_modules:
            continue
        configured_modules.append(normalized)
        seen_modules.add(normalized)
    for entry in summary_entries:
        normalized = str(entry.get("module") or "").strip().upper()
        if not normalized or normalized in seen_modules:
            continue
        configured_modules.append(normalized)
        seen_modules.add(normalized)

    module_summary: dict[tuple[str, str], dict[str, int]] = {}
    module_summary_keys: list[tuple[str, str]] = []
    for file_name in sorted(processed_files):
        for module_name in configured_modules:
            key = (file_name, module_name)
            module_summary[key] = {
                **{metric_key: 0 for metric_key in METRIC_DISPLAY_ORDER},
            }
            module_summary_keys.append(key)
    for entry in summary_entries:
        file_name = str(entry.get("file_name") or "").strip()
        module_name = str(entry.get("module") or "").strip().upper()
        if not file_name or not module_name:
            continue
        key = (file_name, module_name)
        if key not in module_summary:
            module_summary[key] = {
                **{metric_key: 0 for metric_key in METRIC_DISPLAY_ORDER},
            }
            module_summary_keys.append(key)
        item = module_summary[key]
        for metric_key in entry.get("metric_keys", []):
            item[metric_key] += 1

    row_cursor = 20
    ws.cell(row=row_cursor, column=1, value="Module level summary").font = Font(bold=True, size=12)
    ws.cell(row=row_cursor, column=1).fill = section_fill
    row_cursor += 1
    traffic_headers = [
        "File",
        "Module",
        "Overall",
        "Fails",
        "Cpk<1.67",
        "Cpk>20",
        "Site-to-Site Delta",
        "Unique Values",
        "Skewness",
        "Multimodality",
    ]
    for col_idx, header in enumerate(traffic_headers, start=1):
        cell = ws.cell(row=row_cursor, column=col_idx, value=header)
        cell.font = Font(bold=True)
    module_metric_header_row = row_cursor
    row_cursor += 1

    for file_name, module_name in module_summary_keys:
        item = module_summary[(file_name, module_name)]
        module_metric_keys = [metric_key for metric_key in METRIC_DISPLAY_ORDER if item.get(metric_key, 0) > 0]
        overall = _status_text_from_metric_keys(
            module_metric_keys,
            yield_threshold=100.0,
            cpk_low=1.67,
            cpk_high=20.0,
        ) or "OK"
        overall_priority = _priority_from_metric_keys(module_metric_keys)
        ws.cell(row=row_cursor, column=1, value=file_name)
        ws.cell(row=row_cursor, column=2, value=module_name)
        status_cell = ws.cell(row=row_cursor, column=3, value=overall)
        status_cell.font = Font(bold=True)
        status_cell.fill = priority_fills[overall_priority]
        for metric_offset, metric_key in enumerate(METRIC_DISPLAY_ORDER, start=4):
            ws.cell(row=row_cursor, column=metric_offset, value=item[metric_key])
        row_cursor += 1

    module_metric_start_row = module_metric_header_row + 1
    module_metric_end_row = row_cursor - 1
    if module_metric_end_row >= module_metric_start_row:
        for col_idx in range(4, 12):
            col_letter = _excel_col_letter(col_idx)
            ws.conditional_formatting.add(
                f"{col_letter}{module_metric_start_row}:{col_letter}{module_metric_end_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="FFFFFF",
                    end_type="max",
                    end_color="F8696B",
                ),
            )

    row_cursor = ws.max_row + 2
    ws.cell(row=row_cursor, column=1, value="File summary").font = Font(bold=True, size=12)
    ws.cell(row=row_cursor, column=1).fill = section_fill
    row_cursor += 1
    file_headers = ["File", "Affected tests", "High priority", "Medium priority", "Low priority", "Total fail chips", "Data sheet", "Plots sheet"]
    for col_idx, header in enumerate(file_headers, start=1):
        cell = ws.cell(row=row_cursor, column=col_idx, value=header)
        cell.font = Font(bold=True)
    row_cursor += 1

    file_summary: dict[str, dict[str, Any]] = {}
    for entry in summary_entries:
        file_name = str(entry["file_name"])
        item = file_summary.setdefault(
            file_name,
            {
                "affected": 0,
                "high": 0,
                "medium": 0,
                "low": 0,
                "fail_chips": 0,
                "sheet_name": entry["sheet_name"],
                "plots_sheet_name": entry["plots_sheet_name"],
            },
        )
        item["affected"] += 1
        priority = str(entry.get("priority") or "OK").lower()
        if priority in {"high", "medium", "low"}:
            item[priority] += 1
        item["fail_chips"] += int(entry.get("fail_chips") or 0)

    for file_name in sorted(file_summary):
        item = file_summary[file_name]
        ws.cell(row=row_cursor, column=1, value=file_name)
        ws.cell(row=row_cursor, column=2, value=item["affected"])
        ws.cell(row=row_cursor, column=3, value=item["high"])
        ws.cell(row=row_cursor, column=4, value=item["medium"])
        ws.cell(row=row_cursor, column=5, value=item["low"])
        ws.cell(row=row_cursor, column=6, value=item["fail_chips"])
        data_cell = ws.cell(row=row_cursor, column=7, value="Open")
        data_cell.hyperlink = f"#{_excel_internal_sheet_ref(str(item['sheet_name']))}!A1"
        data_cell.font = link_font
        plots_cell = ws.cell(row=row_cursor, column=8, value="Open")
        plots_cell.hyperlink = f"#{_excel_internal_sheet_ref(str(item['plots_sheet_name']))}!A1"
        plots_cell.font = link_font
        row_cursor += 1

    _autofit_openpyxl_columns(ws)
    ws.freeze_panes = "A3"


def generate_yield_cpk_report(
    *,
    input_folder: Path,
    output_folder: Path,
    modules: list[str],
    outlier_mad_multiplier: float,
    yield_threshold: float,
    cpk_low: float,
    cpk_high: float,
    max_files: int | None,
    single_file: str | None,
    encoding: str = DEFAULT_ENCODING,
) -> Path:
    import numpy as np
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill

    output_folder.mkdir(parents=True, exist_ok=True)
    plots_root = output_folder / "cdf_plots"
    plots_root.mkdir(parents=True, exist_ok=True)

    csv_paths = _collect_analysis_csv_paths(
        input_folder,
        single_file=single_file,
        max_files=max_files,
    )
    if not csv_paths:
        raise SystemExit(f"No .csv files found in: {input_folder}")

    modules_upper = {m.strip().upper() for m in modules if m.strip()}
    if not modules_upper:
        raise SystemExit("No modules provided. Example: --modules DPLL,TXPA,TXLO")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    plot_image_targets_by_sheet: dict[str, list[str]] = {}
    overview_entries: list[dict[str, Any]] = []

    total_files = len(csv_paths)
    for file_idx, file_path in enumerate(csv_paths, start=1):
        _print_progress("Yield files", file_idx - 1, total_files, f"starting {file_path.name}")
        print(f"Processing: {file_path.name}")
        meta = scan_flat_file_meta(file_path, encoding=encoding)
        wafer_sig = _parse_filename_wafer_signature(file_path.name)
        temp_label = _parse_insertion_temperature_label(file_path.name)

        # Determine tests of interest by module prefix from Test Name row.
        interest_cols: list[str] = []
        interest_names: dict[str, str] = {}
        interest_modules: dict[str, str] = {}
        for test_col in meta.numeric_test_cols:
            tn = _test_name_from_meta(meta, test_col)
            mod = _module_from_test_name(tn)
            if mod in modules_upper:
                interest_cols.append(test_col)
                interest_names[test_col] = tn
                interest_modules[test_col] = mod

        if not interest_cols:
            print(f"  - No tests matched modules {sorted(modules_upper)}; skipping")
            _print_progress("Yield files", file_idx, total_files, f"skipped {file_path.name}")
            continue

        yield_by_col: dict[str, float | None] = {}
        cpk_by_col: dict[str, float | None] = {}
        wanted_meta_cols = [
            c
            for c in ("SITE_NUM", "WAFER", "X", "Y", "LOT", "SUBLOT", "CHIP_ID", "PF", "FIRST_FAIL_TEST")
            if c in meta.header
        ]
        usecols = wanted_meta_cols + interest_cols
        df_units = _read_unit_data(
            file_path,
            data_start_line_index=meta.data_start_line_index,
            usecols=usecols,
            encoding=encoding,
        )

        # Ensure meta col dataframe aligns to series indices for metric/comment generation.
        meta_cols_df = df_units[wanted_meta_cols].copy() if wanted_meta_cols else pd.DataFrame(index=df_units.index)

        affected: list[str] = []
        assessment_by_col: dict[str, TestMetricAssessment] = {}
        numeric_series_by_col: dict[str, Any] = {}
        for test_col in interest_cols:
            numeric_series = pd.to_numeric(df_units[test_col], errors="coerce")
            y, c = _yield_cpk_from_meta(meta, test_col)
            yield_by_col[test_col] = y
            cpk_by_col[test_col] = c
            _, _, unit = _limits_from_meta(meta, test_col)
            assessment = _assess_test_metrics(
                series=numeric_series,
                meta_cols=meta_cols_df,
                unit=unit,
                yield_pct=y,
                cpk=c,
                yield_threshold=yield_threshold,
                cpk_low=cpk_low,
                cpk_high=cpk_high,
                wafer_sig=wafer_sig,
            )
            assessment_by_col[test_col] = assessment
            if assessment.status_text:
                affected.append(test_col)
                numeric_series_by_col[test_col] = numeric_series

        affected.sort(key=lambda c: (interest_modules.get(c, ""), int(c)))

        if not affected:
            print("  - No affected tests (per all configured metrics); skipping sheet")
            _print_progress("Yield files", file_idx, total_files, f"skipped {file_path.name}")
            continue

        # Create sheets.
        sheet_name = _unique_sheet_name(file_path.stem, wb.sheetnames)
        ws = wb.create_sheet(sheet_name)
        plot_sheet_name = _unique_sheet_name((file_path.stem[:25] + "_PLOTS"), wb.sheetnames)
        ws_plots = wb.create_sheet(plot_sheet_name)

        # Tab colors: keep data vs plots tabs distinct.
        ws.sheet_properties.tabColor = "4F81BD"  # blue
        ws_plots.sheet_properties.tabColor = "C0504D"  # red

        headers = [
            "Module",
            "Test Nr",
            "Test Name",
            "Unit",
            "CDF Plot",
            "Yield (%)",
            "Cpk",
            "Failing Chips",
            "Fails",
            "Cpk<1.67",
            "Cpk>20",
            "Site-to-Site Delta",
            "Multimodality",
            "Unique Values",
            "Skewness",
            "Findings",
            "Outliers",
            "N",
            "Original LTL",
            "Original UTL",
            "LTL 6s",
            "UTL 6s",
            "LTL 12s",
            "UTL 12s",
            "TE notes",
        ]
        ws.append(headers)
        metric_header_fill = PatternFill(patternType="solid", fgColor="FFFF00")
        rotated_metric_alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        metric_header_names = {
            "Fails",
            "Cpk<1.67",
            "Cpk>20",
            "Site-to-Site Delta",
            "Multimodality",
            "Unique Values",
            "Skewness",
        }
        for col_idx, header in enumerate(headers, start=1):
            if header in metric_header_names:
                header_cell = ws.cell(row=1, column=col_idx)
                header_cell.fill = metric_header_fill
                header_cell.alignment = rotated_metric_alignment
        ws.row_dimensions[1].height = 72

        ws_plots.append([
            "Test",
            "CDF (fails highlighted)",
            "CDF zoomed",
            "CDF by Site zoomed",
            "Wafer map (static PNG)",
        ])
        ws_plots["A1"].font = Font(bold=True)
        ws_plots["B1"].font = Font(bold=True)
        ws_plots["C1"].font = Font(bold=True)
        ws_plots["D1"].font = Font(bold=True)
        ws_plots["E1"].font = Font(bold=True)
        ws_plots.sheet_view.zoomScale = 85

        plot_anchor_row = 3
        out_rows = 0
        total_affected = len(affected)
        for test_idx, test_col in enumerate(affected, start=1):
            _print_progress(
                "Yield tests",
                test_idx - 1,
                total_affected,
                f"{file_path.name} | test {test_col}",
            )
            test_name = interest_names.get(test_col, "")
            module = interest_modules.get(test_col, "")
            low, high, unit = _limits_from_meta(meta, test_col)
            y = yield_by_col.get(test_col)
            c = cpk_by_col.get(test_col)
            assessment = assessment_by_col[test_col]
            status = assessment.status_text or ""
            metric_key_set = set(assessment.metric_keys)

            numeric = numeric_series_by_col[test_col]
            finite = numeric.dropna().to_numpy(dtype=float)
            finite = finite[np.isfinite(finite)]
            n = int(finite.size)
            if n == 0:
                continue

            low_eval = -np.inf if low is None else float(low)
            high_eval = np.inf if high is None else float(high)
            if low is None and high is None:
                n_fail = 0
            else:
                n_fail = int(((finite < low_eval) | (finite > high_eval)).sum())

            # Outliers
            med = float(np.median(finite))
            mad = _mad(finite)
            n_out = 0
            if mad > 0:
                n_out = int((np.abs(finite - med) > (outlier_mad_multiplier * mad)).sum())

            l6, u6, l12, u12 = _proposed_sigma_limits(numeric)
            comment = _build_comment(
                series=numeric,
                meta_cols=meta_cols_df,
                outlier_mad_multiplier=outlier_mad_multiplier,
                low_limit=low,
                high_limit=high,
                wafer_sig=wafer_sig,
                metric_assessment=assessment,
            )

            # Plot file and embed in plot sheet.
            safe_test = re.sub(r"[^A-Za-z0-9._-]+", "_", test_name)[:80] or test_col
            plot_path = plots_root / file_path.stem / f"{test_col}_{safe_test}.png"
            plot_zoomed_path = plots_root / file_path.stem / f"{test_col}_{safe_test}_zoomed.png"
            include_sigma_limits = (
                METRIC_CPK_LOW in assessment.metric_keys or METRIC_CPK_HIGH in assessment.metric_keys
            )
            title = _build_plot_title(
                test_name=test_name,
                test_col=test_col,
                temp_label=temp_label,
                cpk=c,
                mean_v=float(np.mean(finite)),
                median_v=float(np.median(finite)),
            )
            _cdf_plot_png_pair(
                finite,
                title=title,
                out_path=plot_path,
                zoomed_out_path=plot_zoomed_path,
                low_limit=low,
                high_limit=high,
                proposed_l6=(l6 if include_sigma_limits else None),
                proposed_u6=(u6 if include_sigma_limits else None),
                proposed_l12=(l12 if include_sigma_limits else None),
                proposed_u12=(u12 if include_sigma_limits else None),
            )

            wafer_maps_supported = _supports_wafer_maps(file_path.name)
            wafer_map_path = plots_root / file_path.stem / f"{test_col}_{safe_test}_wafermap.png"
            if wafer_maps_supported:
                _wafer_map_png(
                    numeric,
                    meta_cols=meta_cols_df,
                    title=f"{test_name} ({test_col}) | {temp_label}",
                    out_path=wafer_map_path,
                    low_limit=low,
                    high_limit=high,
                    unit=unit,
                    median_v=float(np.median(finite)),
                )
            site_cdf_path = plots_root / file_path.stem / f"{test_col}_{safe_test}_cdf_by_site_zoomed.png"
            if METRIC_SITE_DELTA in metric_key_set:
                _cdf_plot_by_site_png(
                    numeric,
                    meta_cols=meta_cols_df,
                    title=title,
                    out_path=site_cdf_path,
                    low_limit=low,
                    high_limit=high,
                    zoom_to_limits=True,
                )

            # Write to plots sheet.
            ws_plots[f"A{plot_anchor_row}"] = f"{test_col} {test_name}"
            ws_plots[f"A{plot_anchor_row}"].font = Font(bold=True)

            # Zoomable links (open the PNG in an external viewer).
            cdf_link = ws_plots[f"B{plot_anchor_row}"]
            cdf_link.value = "Open CDF PNG"
            if plot_path.exists():
                plot_uri = plot_path.resolve().as_uri()
                cdf_link.hyperlink = plot_uri
                cdf_link.font = Font(color="0000EE", underline="single")
                plot_image_targets_by_sheet.setdefault(ws_plots.title, []).append(plot_uri)

            cdf_zoomed_link = ws_plots[f"C{plot_anchor_row}"]
            cdf_zoomed_link.value = "Open zoomed CDF PNG"
            plot_zoomed_uri = None
            if plot_zoomed_path.exists():
                plot_zoomed_uri = plot_zoomed_path.resolve().as_uri()
                cdf_zoomed_link.hyperlink = plot_zoomed_uri
                cdf_zoomed_link.font = Font(color="0000EE", underline="single")
                plot_image_targets_by_sheet.setdefault(ws_plots.title, []).append(plot_zoomed_uri)

            site_cdf_link = ws_plots[f"D{plot_anchor_row}"]
            site_cdf_link.value = "Open site CDF zoomed PNG" if site_cdf_path.exists() else "Not generated"
            if site_cdf_path.exists():
                site_cdf_uri = site_cdf_path.resolve().as_uri()
                site_cdf_link.hyperlink = site_cdf_uri
                site_cdf_link.font = Font(color="0000EE", underline="single")
            else:
                site_cdf_uri = None

            wafer_png_link = ws_plots[f"E{plot_anchor_row}"]
            wafer_png_link.value = "Not applicable" if not wafer_maps_supported else "Open wafer PNG"
            wafer_png_uri = None
            if wafer_map_path.exists():
                wafer_png_uri = wafer_map_path.resolve().as_uri()
                wafer_png_link.hyperlink = wafer_png_uri
                wafer_png_link.font = Font(color="0000EE", underline="single")

            image_width = 480
            image_height = 330
            image_row_span = 19
            plot_block_height_rows = 40

            top_image_row = plot_anchor_row + 1
            bottom_image_row = top_image_row + image_row_span
            cdf_anchor_cell = f"B{top_image_row}"
            cdf_zoomed_anchor_cell = f"K{top_image_row}"
            site_cdf_anchor_cell = f"B{bottom_image_row}"
            wafer_anchor_cell = f"K{bottom_image_row}"
            if plot_path.exists():
                img = XLImage(str(plot_path))
                img.width = image_width
                img.height = image_height
                ws_plots.add_image(img, cdf_anchor_cell)
            if plot_zoomed_path.exists():
                zimg = XLImage(str(plot_zoomed_path))
                zimg.width = image_width
                zimg.height = image_height
                ws_plots.add_image(zimg, cdf_zoomed_anchor_cell)

            if wafer_map_path.exists():
                wimg = XLImage(str(wafer_map_path))
                wimg.width = image_width
                wimg.height = image_height
                ws_plots.add_image(wimg, wafer_anchor_cell)
                if wafer_png_uri is not None:
                    plot_image_targets_by_sheet.setdefault(ws_plots.title, []).append(wafer_png_uri)
            if site_cdf_path.exists():
                simg = XLImage(str(site_cdf_path))
                simg.width = image_width
                simg.height = image_height
                ws_plots.add_image(simg, site_cdf_anchor_cell)
                if site_cdf_uri is not None:
                    plot_image_targets_by_sheet.setdefault(ws_plots.title, []).append(site_cdf_uri)
            # Reserve some rows for the image.
            plot_link_target = f"#{_excel_internal_sheet_ref(ws_plots.title)}!{cdf_anchor_cell}"
            plot_anchor_row += plot_block_height_rows

            # Write row to data sheet.
            row = [
                module,
                int(test_col),
                test_name,
                unit,
                "View",
                y,
                c,
                n_fail,
                "YES" if METRIC_YIELD in metric_key_set else "NO",
                "YES" if METRIC_CPK_LOW in metric_key_set else "NO",
                "YES" if METRIC_CPK_HIGH in metric_key_set else "NO",
                "YES" if METRIC_SITE_DELTA in metric_key_set else "NO",
                max(int(assessment.peak_count or 0), 1),
                "NO" if METRIC_UNIQUE_VALUES in metric_key_set else "YES",
                "YES" if METRIC_SKEWNESS in metric_key_set else "NO",
                comment,
                n_out,
                n,
                low,
                high,
                l6,
                u6,
                l12,
                u12,
                "",
            ]
            ws.append(row)
            out_rows += 1

            row_idx = 1 + out_rows
            for col_name in ("Original LTL", "Original UTL"):
                col_idx = headers.index(col_name) + 1
                ws.cell(row=row_idx, column=col_idx).number_format = "0.######"
            for col_name in ("LTL 6s", "UTL 6s", "LTL 12s", "UTL 12s"):
                col_idx = headers.index(col_name) + 1
                ws.cell(row=row_idx, column=col_idx).number_format = "0.0"

            link_cell = ws.cell(row=1 + out_rows, column=headers.index("CDF Plot") + 1)
            link_cell.hyperlink = plot_link_target
            link_cell.font = Font(color="0000EE", underline="single")

            overview_entries.append(
                {
                    "file_name": file_path.name,
                    "sheet_name": ws.title,
                    "plots_sheet_name": ws_plots.title,
                    "module": module,
                    "test_col": int(test_col),
                    "test_name": test_name,
                    "status": status,
                    "metric_keys": assessment.metric_keys,
                    "priority": assessment.priority,
                    "fail_chips": n_fail,
                    "yield_pct": y,
                    "cpk": c,
                    "detail_link": f"#{_excel_internal_sheet_ref(ws.title)}!A{row_idx}",
                }
            )

            _print_progress(
                "Yield tests",
                test_idx,
                total_affected,
                f"{file_path.name} | test {test_col}",
            )

        if out_rows == 0:
            # Avoid leaving empty sheets.
            wb.remove(ws)
            wb.remove(ws_plots)
            _print_progress("Yield files", file_idx, total_files, f"skipped {file_path.name}")
            continue

        fail_chips_col_idx = headers.index("Failing Chips") + 1
        fail_chips_col_letter = _excel_col_letter(fail_chips_col_idx)
        fail_chips_range = f"{fail_chips_col_letter}2:{fail_chips_col_letter}{1 + out_rows}"
        ws.conditional_formatting.add(
            fail_chips_range,
            ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            ),
        )

        yes_fill = PatternFill(patternType="solid", fgColor="FFC7CE")
        no_fill = PatternFill(patternType="solid", fgColor="C6EFCE")
        neutral_fill = PatternFill(patternType="solid", fgColor="FFEB9C")
        yes_font = Font(color="9C0006")
        no_font = Font(color="006100")
        neutral_font = Font(color="9C6500")
        yes_no_metric_headers = [
            "Fails",
            "Cpk<1.67",
            "Cpk>20",
            "Site-to-Site Delta",
            "Skewness",
        ]
        for col_name in yes_no_metric_headers:
            col_idx = headers.index(col_name) + 1
            col_letter = _excel_col_letter(col_idx)
            metric_range = f"{col_letter}2:{col_letter}{1 + out_rows}"
            ws.conditional_formatting.add(
                metric_range,
                FormulaRule(formula=[f'EXACT({col_letter}2,"YES")'], fill=yes_fill, font=yes_font),
            )
            ws.conditional_formatting.add(
                metric_range,
                FormulaRule(formula=[f'EXACT({col_letter}2,"NO")'], fill=no_fill, font=no_font),
            )

        unique_values_col_idx = headers.index("Unique Values") + 1
        unique_values_col_letter = _excel_col_letter(unique_values_col_idx)
        unique_values_range = f"{unique_values_col_letter}2:{unique_values_col_letter}{1 + out_rows}"
        ws.conditional_formatting.add(
            unique_values_range,
            FormulaRule(formula=[f'EXACT({unique_values_col_letter}2,"NO")'], fill=yes_fill, font=yes_font),
        )
        ws.conditional_formatting.add(
            unique_values_range,
            FormulaRule(formula=[f'EXACT({unique_values_col_letter}2,"YES")'], fill=no_fill, font=no_font),
        )

        multimodality_col_idx = headers.index("Multimodality") + 1
        multimodality_col_letter = _excel_col_letter(multimodality_col_idx)
        multimodality_range = f"{multimodality_col_letter}2:{multimodality_col_letter}{1 + out_rows}"
        ws.conditional_formatting.add(
            multimodality_range,
            FormulaRule(formula=[f"{multimodality_col_letter}2=1"], fill=no_fill, font=no_font),
        )
        ws.conditional_formatting.add(
            multimodality_range,
            FormulaRule(formula=[f"{multimodality_col_letter}2<>1"], fill=yes_fill, font=yes_font),
        )

        # Apply module block coloring on the data sheet.
        _apply_module_group_row_colors(ws)
        for row_idx in range(2, 2 + out_rows):
            for col_name in yes_no_metric_headers:
                metric_cell = ws.cell(row=row_idx, column=headers.index(col_name) + 1)
                if metric_cell.value == "YES":
                    metric_cell.fill = yes_fill
                    metric_cell.font = yes_font
                elif metric_cell.value == "NO":
                    metric_cell.fill = no_fill
                    metric_cell.font = no_font

            unique_values_cell = ws.cell(row=row_idx, column=unique_values_col_idx)
            if unique_values_cell.value == "NO":
                unique_values_cell.fill = yes_fill
                unique_values_cell.font = yes_font
            elif unique_values_cell.value == "YES":
                unique_values_cell.fill = no_fill
                unique_values_cell.font = no_font

            multimodality_cell = ws.cell(row=row_idx, column=multimodality_col_idx)
            if multimodality_cell.value == 1:
                multimodality_cell.fill = no_fill
                multimodality_cell.font = no_font
            else:
                multimodality_cell.fill = yes_fill
                multimodality_cell.font = yes_font

            te_notes_cell = ws.cell(row=row_idx, column=headers.index("TE notes") + 1)
            te_notes_cell.fill = PatternFill(fill_type=None)
            te_notes_cell.font = Font(color="000000")

        # Auto-fit columns.
        _autofit_openpyxl_columns(ws)
        _autofit_openpyxl_columns(ws_plots)

        hidden_start_idx = headers.index("Original LTL") + 1
        hidden_end_idx = headers.index("UTL 12s") + 1
        hidden_start = _excel_col_letter(hidden_start_idx)
        hidden_end = _excel_col_letter(hidden_end_idx)
        ws.column_dimensions.group(hidden_start, hidden_end, outline_level=1, hidden=True)
        for col_idx in range(hidden_start_idx, hidden_end_idx + 1):
            col_letter = _excel_col_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True
            ws.column_dimensions[col_letter].outline_level = 1

        # Freeze header row.
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{_excel_col_letter(ws.max_column)}{ws.max_row}"
        _print_progress("Yield files", file_idx, total_files, f"finished {file_path.name}")

    from datetime import datetime

    out_xlsx = output_folder / "Test_Data_Analysis_Report.xlsx"
    if overview_entries:
        _add_overview_sheet(
            wb,
            summary_entries=overview_entries,
            modules=modules,
            processed_files=[p.name for p in csv_paths],
            output_folder=output_folder,
        )
    for sheet in wb.worksheets:
        _autofit_openpyxl_columns(sheet)
    link_issues = _self_check_workbook_internal_hyperlinks(wb)
    if link_issues:
        preview = "\n".join(f"- {msg}" for msg in link_issues[:15])
        more = "" if len(link_issues) <= 15 else f"\n- ... and {len(link_issues) - 15} more"
        raise ValueError("Internal hyperlink self-check failed:\n" + preview + more)

    try:
        wb.save(out_xlsx)
        _enable_clickable_plot_images(out_xlsx, plot_image_targets_by_sheet)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = output_folder / f"Test_Data_Analysis_Report_{ts}.xlsx"
        wb.save(alt)
        _enable_clickable_plot_images(alt, plot_image_targets_by_sheet)
        print(f"Could not overwrite (file open?): {out_xlsx}")
        print(f"Saved instead: {alt}")
        return alt

    print(f"Saved: {out_xlsx}")
    return out_xlsx


def generate_correlation_workbook(
    *,
    input_folder: Path,
    output_folder: Path,
    modules: list[str],
    max_files: int | None,
    single_file: str | None,
    methods: list[Literal["pearson", "spearman"]],
    pearson_abs_min_for_report: float,
    encoding: str = DEFAULT_ENCODING,
) -> Path:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font

    output_folder.mkdir(parents=True, exist_ok=True)
    csv_paths = _collect_analysis_csv_paths(
        input_folder,
        single_file=single_file,
        max_files=max_files,
    )
    if not csv_paths:
        raise SystemExit(f"No .csv files found in: {input_folder}")

    modules_upper = {m.strip().upper() for m in modules if m.strip()}
    if not modules_upper:
        raise SystemExit("No modules provided.")

    wb = Workbook()
    wb.remove(wb.active)
    excel_max_rows = 1_048_576
    excel_max_data_rows = excel_max_rows - 1  # account for header row

    total_files = len(csv_paths)
    for file_idx, file_path in enumerate(csv_paths, start=1):
        _print_progress("Corr files", file_idx - 1, total_files, f"starting {file_path.name}")
        print(f"Correlation: {file_path.name}")
        meta = scan_flat_file_meta(file_path, encoding=encoding)

        test_name_by_col = {c: _test_name_from_meta(meta, c) for c in meta.numeric_test_cols}
        module_by_col = {c: _module_from_test_name(test_name_by_col[c]) for c in meta.numeric_test_cols}
        module_cols = [c for c in meta.numeric_test_cols if module_by_col[c] in modules_upper]
        if not module_cols:
            print("  - No module tests found; skipping")
            _print_progress("Corr files", file_idx, total_files, f"skipped {file_path.name}")
            continue

        # Load ALL numeric tests for correlation (optional heavy step).
        df = _read_unit_data(
            file_path,
            data_start_line_index=meta.data_start_line_index,
            usecols=meta.numeric_test_cols,
            encoding=encoding,
        )
        df = df.apply(pd.to_numeric, errors="coerce")

        sheet = wb.create_sheet(_unique_sheet_name(file_path.stem, wb.sheetnames))
        headers = [
            "Module",
            "Test Nr",
            "Test Name",
            "Other Test Nr",
            "Other Test Name",
        ] + [m.title() for m in methods]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # Precompute ranks for spearman if needed.
        df_rank = None
        if "spearman" in methods:
            df_rank = df.rank(axis=0, method="average", na_option="keep")

        out_rows = 0
        truncated = False
        total_module_cols = len(module_cols)
        for test_idx, test_col in enumerate(module_cols, start=1):
            _print_progress(
                "Corr tests",
                test_idx - 1,
                total_module_cols,
                f"{file_path.name} | test {test_col}",
            )
            if out_rows >= excel_max_data_rows:
                truncated = True
                break
            s = df[test_col]
            if s.dropna().nunique() < 2:
                continue
            pearson = _safe_corr_against_all(df, test_col) if "pearson" in methods else None
            spearman = _safe_corr_against_all(df_rank, test_col) if ("spearman" in methods and df_rank is not None) else None

            for other_col in meta.numeric_test_cols:
                if out_rows >= excel_max_data_rows:
                    truncated = True
                    break
                if other_col == test_col:
                    continue

                pearson_value = None
                if pearson is not None:
                    pearson_value = _to_float(pearson.get(other_col))
                if pearson_value is None or abs(float(pearson_value)) <= float(pearson_abs_min_for_report):
                    continue

                row = [
                    module_by_col[test_col],
                    int(test_col),
                    test_name_by_col[test_col],
                    int(other_col),
                    test_name_by_col[other_col],
                ]
                if pearson is not None:
                    row.append(pearson_value)
                if spearman is not None:
                    row.append(_to_float(spearman.get(other_col)))
                sheet.append(row)
                out_rows += 1

            _print_progress(
                "Corr tests",
                test_idx,
                total_module_cols,
                f"{file_path.name} | test {test_col}",
            )

        if truncated:
            print(f"  - Correlation rows truncated at Excel limit ({excel_max_data_rows} data rows)")

        if out_rows == 0:
            wb.remove(sheet)
            _print_progress("Corr files", file_idx, total_files, f"skipped {file_path.name}")
            continue

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{_excel_col_letter(sheet.max_column)}{sheet.max_row}"
        _autofit_openpyxl_columns(sheet)
        _print_progress("Corr files", file_idx, total_files, f"finished {file_path.name}")

    from datetime import datetime

    out_xlsx = output_folder / "Correlation_Report.xlsx"
    try:
        wb.save(out_xlsx)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = output_folder / f"Correlation_Report_{ts}.xlsx"
        wb.save(alt)
        print(f"Could not overwrite (file open?): {out_xlsx}")
        print(f"Saved instead: {alt}")
        return alt

    print(f"Saved: {out_xlsx}")
    return out_xlsx


def _safe_pair_correlation(a, b) -> float | None:
    """Compute Pearson correlation for two aligned series safely.

    Returns None when there are not enough paired values or one side is constant.
    """
    import numpy as np
    import pandas as pd

    aa = pd.to_numeric(a, errors="coerce")
    bb = pd.to_numeric(b, errors="coerce")
    valid = aa.notna() & bb.notna()
    if int(valid.sum()) < 2:
        return None

    av = aa[valid].to_numpy(dtype=float)
    bv = bb[valid].to_numpy(dtype=float)
    if av.size < 2 or bv.size < 2:
        return None

    std_a = float(np.std(av, ddof=1))
    std_b = float(np.std(bv, ddof=1))
    if not np.isfinite(std_a) or not np.isfinite(std_b) or std_a == 0.0 or std_b == 0.0:
        return None

    corr = float(np.corrcoef(av, bv)[0, 1])
    if not np.isfinite(corr):
        return None
    return corr


def _safe_spearman_correlation(a, b) -> float | None:
    """Compute Spearman correlation without requiring SciPy."""
    import pandas as pd

    aa = pd.to_numeric(a, errors="coerce")
    bb = pd.to_numeric(b, errors="coerce")
    valid = aa.notna() & bb.notna()
    if int(valid.sum()) < 2:
        return None

    aa_rank = aa[valid].rank(method="average")
    bb_rank = bb[valid].rank(method="average")
    return _safe_pair_correlation(aa_rank, bb_rank)


def _safe_corr_against_all(df, target_col: str) -> dict[str, float | None]:
    """Safe correlation of target column against all columns in df."""
    out: dict[str, float | None] = {}
    target = df[target_col]
    for col in df.columns:
        if col == target_col:
            continue
        out[col] = _safe_pair_correlation(target, df[col])
    return out


def run() -> int:
    """Run analysis using runtime configuration supplied by the YAML wrapper."""
    config = _require_runtime_configuration()
    input_folder = config["input_folder"]
    output_folder = config["output_folder"]

    output_folder.mkdir(parents=True, exist_ok=True)

    generate_yield_cpk_report(
        input_folder=input_folder,
        output_folder=output_folder,
        modules=config["modules"],
        outlier_mad_multiplier=config["outlier_mad_multiplier"],
        yield_threshold=config["yield_threshold"],
        cpk_low=config["cpk_low"],
        cpk_high=config["cpk_high"],
        max_files=config["max_files"],
        single_file=config["single_file"],
        encoding=config["encoding"],
    )

    if config["generate_correlation_report"]:
        generate_correlation_workbook(
            input_folder=input_folder,
            output_folder=output_folder,
            modules=config["modules"],
            max_files=config["max_files"],
            single_file=config["single_file"],
            methods=config["correlation_methods"],
            pearson_abs_min_for_report=config["pearson_abs_min_for_report"],
            encoding=config["encoding"],
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(
        "Direct execution of Tests_Data_Analysis.py is disabled. "
        "Use run_tests_data_analysis.py --config <yaml> or run_tests_data_analysis.ps1."
    )
