from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
import textwrap
import unittest
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


TEST_DATA_ANALYSIS_DIR = Path(__file__).resolve().parents[1]
SAMPLE_INPUT_DIR = TEST_DATA_ANALYSIS_DIR / "tests" / "smoke_input"
SAMPLE_FILE_NAME = "smoke_Q2_sample.csv"
OVERVIEW_SHEET_NAME = "Overview"
SAMPLE_SHEET_NAME = "smoke_Q2_sample"
SAMPLE_PLOTS_SHEET_NAME = "smoke_Q2_sample_PLOTS"


class TestsDataAnalysisSmokeTest(unittest.TestCase):
    def test_cli_generates_expected_reports(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            input_dir = tmp_path / "input"
            shutil.copytree(SAMPLE_INPUT_DIR, input_dir)
            (input_dir / "smoke_Q2_sample_dtr_records.csv").write_text(
                "Index;Category;Message\n1;ALARM;Example sidecar\n",
                encoding="utf-8",
            )
            output_dir = tmp_path / "outputs"
            output_dir.mkdir(parents=True, exist_ok=True)
            config_path = tmp_path / "config_smoke.yaml"
            config_path.write_text(
                textwrap.dedent(
                    f"""\
                    input_folder: {input_dir.as_posix()}
                    output_folder: {output_dir.as_posix()}
                    modules:
                      - TXPA
                      - DPLL
                    yield_threshold: 100.0
                    cpk_low: 1.67
                    cpk_high: 20.0
                    outlier_mad_multiplier: 6.0
                    max_files: 1
                    single_file: {SAMPLE_FILE_NAME}
                    encoding: latin1
                    generate_correlation_report: true
                    correlation_methods:
                      - pearson
                    pearson_abs_min_for_report: 0.5
                    wafermap_circle_area_mult: 1.0
                    """
                ),
                encoding="utf-8",
            )

            result = subprocess.run(
                [sys.executable, "run_tests_data_analysis.py", "--config", str(config_path)],
                cwd=TEST_DATA_ANALYSIS_DIR,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(
                result.returncode,
                0,
                msg=(
                    "Smoke test launcher failed.\n"
                    f"STDOUT:\n{result.stdout}\n\n"
                    f"STDERR:\n{result.stderr}"
                ),
            )
            self.assertIn("[Yield files]", result.stdout)
            self.assertIn("[Yield tests]", result.stdout)
            self.assertIn("[Corr files]", result.stdout)
            self.assertIn("[Corr tests]", result.stdout)
            self.assertIn("[Workflow]", result.stdout)
            self.assertIn("100%", result.stdout)

            yield_report = output_dir / "Test_Data_Analysis_Report.xlsx"
            correlation_report = output_dir / "Correlation_Report.xlsx"
            self.assertTrue(yield_report.exists(), "Yield/Cpk report was not created")
            self.assertTrue(correlation_report.exists(), "Correlation report was not created")

            png_files = list((output_dir / "cdf_plots").rglob("*.png"))
            self.assertGreaterEqual(len(png_files), 2, "Expected CDF and wafer map PNG outputs")
            html_files = list((output_dir / "cdf_plots").rglob("*.html"))
            self.assertEqual(len(html_files), 0, "Did not expect interactive wafer map HTML without wafer/XY signature")

            workbook = load_workbook(yield_report, read_only=True, data_only=True)
            try:
                self.assertIn(OVERVIEW_SHEET_NAME, workbook.sheetnames)
                self.assertIn(SAMPLE_SHEET_NAME, workbook.sheetnames)
                self.assertIn(SAMPLE_PLOTS_SHEET_NAME, workbook.sheetnames)
                self.assertEqual(workbook.sheetnames[0], OVERVIEW_SHEET_NAME)
                overview_worksheet = workbook[OVERVIEW_SHEET_NAME]
                worksheet = workbook[SAMPLE_SHEET_NAME]
                plots_worksheet = workbook[SAMPLE_PLOTS_SHEET_NAME]
                data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
                self.assertEqual(overview_worksheet["A1"].value, "Test Data Analysis Overview")
                self.assertEqual(overview_worksheet["A5"].value, "Files processed")
                self.assertEqual(overview_worksheet["B5"].value, 1)
                self.assertEqual(overview_worksheet["A7"].value, "Affected tests")
                self.assertEqual(overview_worksheet["B7"].value, 1)
                self.assertEqual(overview_worksheet["A9"].value, "High-priority tests")
                self.assertEqual(overview_worksheet["A20"].value, "Module level summary")
                self.assertEqual(overview_worksheet["C21"].value, "Fails")
                self.assertEqual(overview_worksheet["G21"].value, "Unique Values")
                self.assertEqual(overview_worksheet["A22"].value, "TXPA")
                self.assertEqual(overview_worksheet["B22"].value, "Fails + Cpk<1.67")
                self.assertEqual(overview_worksheet["A23"].value, "DPLL")
                self.assertEqual(overview_worksheet["B23"].value, "OK")
                self.assertEqual(overview_worksheet["A25"].value, "File summary")
                self.assertEqual(len(data_rows), 1, "Expected exactly one affected test row in the smoke sample")
                self.assertEqual(worksheet["F1"].value, "Fail Chips")
                self.assertEqual(worksheet["H1"].value, "Fails")
                self.assertEqual(worksheet["H1"].fill.fgColor.rgb, "00FFFF00")
                self.assertEqual(worksheet["I1"].value, "Cpk<1.67")
                self.assertEqual(worksheet["L1"].value, "Multimodality")
                self.assertEqual(worksheet["M1"].value, "Unique Values")
                self.assertEqual(worksheet["H1"].alignment.horizontal, "center")
                self.assertEqual(worksheet["H1"].alignment.vertical, "center")
                self.assertEqual(worksheet["H1"].alignment.textRotation, 90)
                self.assertEqual(plots_worksheet["C1"].value, "Wafer map (interactive HTML)")
                self.assertEqual(plots_worksheet["D1"].value, "Wafer map (static PNG)")
                self.assertEqual(plots_worksheet["E1"].value, "CDF by Site")
                self.assertEqual(plots_worksheet["C3"].value, "Not generated")
                self.assertEqual(plots_worksheet["D3"].value, "Open wafer PNG")
                self.assertEqual(plots_worksheet["E3"].value, "Not generated")

                txpa_row = data_rows[0]
                self.assertEqual(txpa_row[0], "TXPA")
                self.assertEqual(txpa_row[1], 520123)
                self.assertEqual(txpa_row[2], "TXPA_OUTPUT_PWR")
                self.assertEqual(txpa_row[5], 2)
                self.assertEqual(txpa_row[7], "YES")
                self.assertEqual(txpa_row[8], "YES")
                self.assertEqual(txpa_row[9], "NO")
                self.assertEqual(txpa_row[10], "NO")
                self.assertEqual(txpa_row[11], 1)
                self.assertEqual(txpa_row[12], "YES")
                self.assertEqual(txpa_row[16], "View")
            finally:
                workbook.close()

            workbook_with_formatting = load_workbook(yield_report, read_only=False, data_only=False)
            try:
                worksheet_with_formatting = workbook_with_formatting[SAMPLE_SHEET_NAME]
                overview_with_formatting = workbook_with_formatting[OVERVIEW_SHEET_NAME]
                cf_rules = list(worksheet_with_formatting.conditional_formatting)
                overview_cf_rules = list(overview_with_formatting.conditional_formatting)
                self.assertTrue(cf_rules, "Expected conditional formatting rules in yield workbook")
                self.assertTrue(
                    any(str(rule.sqref) in {"F2", "F2:F2"} for rule in cf_rules),
                    "Expected Fail Chips conditional formatting on column F",
                )
                self.assertTrue(
                    any(str(rule.sqref) in {"H2", "H2:H2"} for rule in cf_rules),
                    "Expected Fails YES/NO conditional formatting on column H",
                )
                self.assertTrue(
                    any(str(rule.sqref) in {"I2", "I2:I2"} for rule in cf_rules),
                    "Expected Cpk<1.67 YES/NO conditional formatting on column I",
                )
                self.assertTrue(
                    any(str(rule.sqref) in {"L2", "L2:L2"} for rule in cf_rules),
                    "Expected Multimodality conditional formatting on column L",
                )
                self.assertEqual(worksheet_with_formatting["H2"].fill.fgColor.rgb, "00FFC7CE")
                self.assertEqual(worksheet_with_formatting["I2"].fill.fgColor.rgb, "00FFC7CE")
                self.assertEqual(worksheet_with_formatting["J2"].fill.fgColor.rgb, "00C6EFCE")
                self.assertEqual(worksheet_with_formatting["M2"].fill.fgColor.rgb, "00C6EFCE")
                self.assertLessEqual(float(worksheet_with_formatting.column_dimensions["H"].width), 12.0)
                self.assertLessEqual(float(worksheet_with_formatting.column_dimensions["I"].width), 12.0)
                self.assertLessEqual(float(worksheet_with_formatting.column_dimensions["J"].width), 12.0)
                self.assertGreater(float(worksheet_with_formatting.column_dimensions["C"].width), 12.0)
                self.assertTrue(
                    any(str(rule.sqref) in {"C22:C23", "C22:C23"} for rule in overview_cf_rules),
                    "Expected module-level Fails color scale formatting on column C",
                )
                self.assertTrue(
                    any(str(rule.sqref) in {"G22:G23", "G22:G23"} for rule in overview_cf_rules),
                    "Expected module-level Unique Values color scale formatting on column G",
                )
            finally:
                workbook_with_formatting.close()

            with zipfile.ZipFile(yield_report, "r") as zf:
                drawing_xml_names = [name for name in zf.namelist() if name.startswith("xl/drawings/drawing") and name.endswith(".xml")]
                drawing_rels_names = [
                    name for name in zf.namelist() if name.startswith("xl/drawings/_rels/drawing") and name.endswith(".xml.rels")
                ]
                self.assertTrue(drawing_xml_names, "Expected drawing XML parts in yield workbook")
                self.assertTrue(drawing_rels_names, "Expected drawing relationship parts in yield workbook")

                drawing_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
                package_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

                click_targets: list[str] = []
                for rel_name in drawing_rels_names:
                    rel_root = ET.fromstring(zf.read(rel_name))
                    for rel in rel_root.findall(f"{{{package_ns}}}Relationship"):
                        if rel.attrib.get("Type", "").endswith("/hyperlink"):
                            click_targets.append(rel.attrib.get("Target", ""))

                self.assertGreaterEqual(len(click_targets), 2, "Expected clickable hyperlinks for embedded plots")
                self.assertTrue(any(target.endswith(".png") for target in click_targets))
                self.assertFalse(any(target.endswith(".html") for target in click_targets))

                click_count = 0
                for drawing_name in drawing_xml_names:
                    drawing_root = ET.fromstring(zf.read(drawing_name))
                    click_count += len(drawing_root.findall(f".//{{{drawing_ns}}}hlinkClick"))

                self.assertGreaterEqual(click_count, 2, "Expected embedded plot images to have click hyperlinks")

            corr_workbook = load_workbook(correlation_report, read_only=True, data_only=True)
            try:
                self.assertIn(SAMPLE_SHEET_NAME, corr_workbook.sheetnames)
                corr_sheet = corr_workbook[SAMPLE_SHEET_NAME]
                corr_rows = list(corr_sheet.iter_rows(min_row=2, values_only=True))
                self.assertGreaterEqual(len(corr_rows), 2, "Expected correlation rows in smoke report")
            finally:
                corr_workbook.close()


if __name__ == "__main__":
    unittest.main()
