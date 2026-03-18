from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


IMPORTANT_TESTS = 1000
ENGINEER_DAY_HOURS = 8
DETECTION_CONSERVATIVE = 0.8501
DETECTION_AUTOMATED = 0.9735


def compute_row(problematic_pct: float) -> dict[str, float]:
    issue_tests = IMPORTANT_TESTS * (problematic_pct / 100.0)
    full_manual_hours = (IMPORTANT_TESTS * 14 + issue_tests * 80 + 25 * 60) / 3600
    conservative_hours = (IMPORTANT_TESTS * 2.2 + issue_tests * 28 + 30 * 60) / 3600
    automated_hours = (12 * 60 + (issue_tests * 1.15) * 22 + 15 * 60) / 3600
    full_manual_reviews_per_day = ENGINEER_DAY_HOURS / full_manual_hours
    conservative_reviews_per_day = ENGINEER_DAY_HOURS / conservative_hours
    automated_reviews_per_day = ENGINEER_DAY_HOURS / automated_hours
    throughput_gain_vs_conservative = conservative_hours / automated_hours
    throughput_gain_vs_full_manual = full_manual_hours / automated_hours

    return {
        "Problematic Tests (%)": problematic_pct,
        "Problematic Tests Count": issue_tests,
        "Full Manual Review Time (h)": full_manual_hours,
        "Conservative Review Time (h)": conservative_hours,
        "Automated Review Time (h)": automated_hours,
        "Full Manual Productivity (reviews/8h day)": full_manual_reviews_per_day,
        "Conservative Productivity (reviews/8h day)": conservative_reviews_per_day,
        "Automated Productivity (reviews/8h day)": automated_reviews_per_day,
        "Automation Time Saved vs Full Manual (h)": full_manual_hours - automated_hours,
        "Automation Time Saved vs Conservative (h)": conservative_hours - automated_hours,
        "Automation Throughput Gain vs Full Manual (x)": throughput_gain_vs_full_manual,
        "Automation Throughput Gain vs Conservative (x)": throughput_gain_vs_conservative,
        "Automation Productivity Improvement vs Full Manual (%)": (throughput_gain_vs_full_manual - 1.0) * 100.0,
        "Automation Productivity Improvement vs Conservative (%)": (throughput_gain_vs_conservative - 1.0) * 100.0,
        "Conservative Detection (%)": DETECTION_CONSERVATIVE * 100.0,
        "Automated Detection (%)": DETECTION_AUTOMATED * 100.0,
        "Conservative Issues Found": issue_tests * DETECTION_CONSERVATIVE,
        "Automated Issues Found": issue_tests * DETECTION_AUTOMATED,
        "Conservative Issues Missed": issue_tests * (1.0 - DETECTION_CONSERVATIVE),
        "Automated Issues Missed": issue_tests * (1.0 - DETECTION_AUTOMATED),
    }


def autosize_worksheet(worksheet) -> None:
    for column in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column]
        width = max(len(value) for value in values) + 2
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(width, 32)


def build_plot(curve_df: pd.DataFrame, plot_path: Path) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5), constrained_layout=True)
    productivity_pct_df = curve_df.copy()
    productivity_pct_df["Conservative Productivity (%)"] = 100.0
    productivity_pct_df["Automated Productivity (%)"] = (
        productivity_pct_df["Automation Throughput Gain vs Conservative (x)"] * 100.0
    )

    axes[0].plot(
        curve_df["Problematic Tests (%)"],
        curve_df["Conservative Review Time (h)"],
        marker="o",
        linewidth=2,
        label="Conservative reviewer",
        color="#b14623",
    )
    axes[0].plot(
        curve_df["Problematic Tests (%)"],
        curve_df["Automated Review Time (h)"],
        marker="o",
        linewidth=2,
        label="Automated review",
        color="#0f6cbd",
    )
    axes[0].set_title("Review Time vs Problematic Tests")
    axes[0].set_xlabel("Problematic tests (%)")
    axes[0].set_ylabel("Review time (h)")
    axes[0].grid(alpha=0.3)
    axes[0].legend()

    axes[1].plot(
        productivity_pct_df["Problematic Tests (%)"],
        productivity_pct_df["Conservative Productivity (%)"],
        marker="o",
        linewidth=2,
        label="Conservative reviewer",
        color="#b14623",
    )
    axes[1].plot(
        productivity_pct_df["Problematic Tests (%)"],
        productivity_pct_df["Automated Productivity (%)"],
        marker="o",
        linewidth=2,
        label="Automated review",
        color="#0f6cbd",
    )
    axes[1].set_title("Productivity vs Problematic Tests")
    axes[1].set_xlabel("Problematic tests (%)")
    axes[1].set_ylabel("Relative productivity (%)")
    axes[1].grid(alpha=0.3)
    axes[1].legend()

    fig.suptitle("1000-Test Review Comparison: Automated vs Conservative Reviewer", fontsize=14, fontweight="bold")
    fig.savefig(plot_path, dpi=200)
    plt.close(fig)


def main() -> None:
    output_dir = Path(__file__).resolve().parent / "analysis_outputs"
    output_dir.mkdir(exist_ok=True)

    summary_df = pd.DataFrame(
        [compute_row(problematic_pct) for problematic_pct in (5.0, 2.0, 1.0)]
    )
    curve_df = pd.DataFrame(
        [compute_row(problematic_pct) for problematic_pct in range(0, 11)]
    )

    workbook_path = output_dir / "review_productivity_summary_1000_tests.xlsx"
    plot_path = output_dir / "review_productivity_comparison_1000_tests.png"

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary_1000_Tests", index=False)
        curve_df.to_excel(writer, sheet_name="Curve_0_to_10pct", index=False)

    workbook = load_workbook(workbook_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        autosize_worksheet(worksheet)
    workbook.save(workbook_path)

    build_plot(curve_df, plot_path)

    print(f"Workbook: {workbook_path}")
    print(f"Plot: {plot_path}")


if __name__ == "__main__":
    main()