from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
import textwrap
import unittest
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


TEST_DATA_ANALYSIS_DIR = Path(__file__).resolve().parents[1]
SAMPLE_INPUT_DIR = TEST_DATA_ANALYSIS_DIR / "tests" / "smoke_input"
SAMPLE_FILE_NAME = "smoke_Q2_sample.csv"
OVERVIEW_SHEET_NAME = "Overview"
SAMPLE_SHEET_NAME = "smoke_Q2_sample"
SAMPLE_PLOTS_SHEET_NAME = "smoke_Q2_sample_PLOTS"


class TestDataReviewerSmokeTest(unittest.TestCase):
    def test_cli_generates_expected_reports(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            input_dir = tmp_path / "input"
            shutil.copytree(SAMPLE_INPUT_DIR, input_dir)
            (input_dir / "smoke_Q2_sample_dtr_records.csv").write_text(
                "Index;Category;Message\n1;ALARM;Example sidecar\n",
                encoding="utf-8",
            )
            output_dir = tmp_path / "outputs"
            output_dir.mkdir(parents=True, exist_ok=True)
            config_path = tmp_path / "config_smoke.yaml"
            config_path.write_text(
                textwrap.dedent(
                    f"""\
                    input_folder: {input_dir.as_posix()}
                    output_folder: {output_dir.as_posix()}
                    modules:
                      - TXPA
                      - DPLL
                    yield_threshold: 100.0
                    cpk_low: 1.67
                    cpk_high: 20.0
                    generate_correlation_report: true
                    """
                ),
                encoding="utf-8",
            )

            result = subprocess.run(
                [sys.executable, "run_test_data_reviewer.py", "--config", str(config_path)],
                cwd=TEST_DATA_ANALYSIS_DIR,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(
                result.returncode,
                0,
                msg=(
                    "Smoke test launcher failed.\n"
                    f"STDOUT:\n{result.stdout}\n\n"
                    f"STDERR:\n{result.stderr}"
                ),
            )
            self.assertIn("[Yield files]", result.stdout)
            self.assertIn("[Yield tests]", result.stdout)
            self.assertIn("[Corr files]", result.stdout)
            self.assertIn("[Corr tests]", result.stdout)
            self.assertIn("[Workflow]", result.stdout)
            self.assertIn("100%", result.stdout)

            yield_report = output_dir / "Test_Data_Reviewer_Report.xlsx"
            correlation_report = output_dir / "Correlation_Report.xlsx"
            self.assertTrue(yield_report.exists(), "Yield/Cpk report was not created")
            self.assertTrue(correlation_report.exists(), "Correlation report was not created")

            png_files = list((output_dir / "cdf_plots").rglob("*.png"))
            self.assertGreaterEqual(len(png_files), 2, "Expected CDF and wafer map PNG outputs")
            html_files = list((output_dir / "cdf_plots").rglob("*.html"))
            self.assertEqual(len(html_files), 0, "Did not expect any interactive wafer map HTML outputs")

            workbook = load_workbook(yield_report, read_only=True, data_only=True)
            try:
                self.assertIn(OVERVIEW_SHEET_NAME, workbook.sheetnames)
                self.assertIn(SAMPLE_SHEET_NAME, workbook.sheetnames)
                self.assertIn(SAMPLE_PLOTS_SHEET_NAME, workbook.sheetnames)
                self.assertEqual(workbook.sheetnames[0], OVERVIEW_SHEET_NAME)
                overview_worksheet = workbook[OVERVIEW_SHEET_NAME]
                worksheet = workbook[SAMPLE_SHEET_NAME]
                plots_worksheet = workbook[SAMPLE_PLOTS_SHEET_NAME]
                data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
                self.assertEqual(overview_worksheet["A1"].value, "Test Data Reviewer Overview")
                self.assertEqual(overview_worksheet["A5"].value, "Files processed")
                self.assertEqual(overview_worksheet["B5"].value, 1)
                self.assertEqual(overview_worksheet["A7"].value, "Affected tests")
                self.assertEqual(overview_worksheet["B7"].value, 1)
                self.assertEqual(overview_worksheet["A9"].value, "High-priority tests")
                self.assertEqual(overview_worksheet["A20"].value, "Module level summary")
                self.assertEqual(overview_worksheet["D21"].value, "Fails")
                self.assertEqual(overview_worksheet["H21"].value, "Unique Values")
                self.assertEqual(overview_worksheet["I21"].value, "Multimodality")
                self.assertEqual(overview_worksheet["A22"].value, SAMPLE_FILE_NAME)
                self.assertEqual(overview_worksheet["B22"].value, "TXPA")
                self.assertEqual(overview_worksheet["C22"].value, "Fails + Cpk<1.67")
                self.assertEqual(overview_worksheet["A23"].value, SAMPLE_FILE_NAME)
                self.assertEqual(overview_worksheet["B23"].value, "DPLL")
                self.assertEqual(overview_worksheet["C23"].value, "OK")
                self.assertEqual(overview_worksheet["A25"].value, "File summary")
                self.assertEqual(len(data_rows), 1, "Expected exactly one affected test row in the smoke sample")
                self.assertEqual(worksheet["E1"].value, "CDF Plot")
                self.assertEqual(worksheet["F1"].value, "Yield (%)")
                self.assertEqual(worksheet["G1"].value, "Cpk")
                self.assertEqual(worksheet["H1"].value, "Failing Chips")
                self.assertEqual(worksheet["I1"].value, "Fails")
                self.assertEqual(worksheet["I1"].fill.fgColor.rgb, "00FFFF00")
                self.assertEqual(worksheet["J1"].value, "Cpk<1.67")
                self.assertEqual(worksheet["M1"].value, "Multimodality")
                self.assertEqual(worksheet["N1"].value, "Unique Values")
                self.assertEqual(worksheet["O1"].value, "Findings")
                self.assertEqual(worksheet["P1"].value, "Outliers")
                self.assertEqual(worksheet["Q1"].value, "N")
                self.assertEqual(worksheet["X1"].value, "TE notes")
                self.assertEqual(worksheet["I1"].alignment.horizontal, "center")
                self.assertEqual(worksheet["I1"].alignment.vertical, "center")
                self.assertEqual(worksheet["I1"].alignment.textRotation, 90)
                self.assertEqual(plots_worksheet["C1"].value, "CDF zoomed")
                self.assertEqual(plots_worksheet["D1"].value, "CDF by Site zoomed")
                self.assertEqual(plots_worksheet["E1"].value, "Wafer map (static PNG)")
                self.assertEqual(plots_worksheet.max_column, 5)
                self.assertEqual(plots_worksheet["C3"].value, "Open zoomed CDF PNG")
                self.assertEqual(plots_worksheet["D3"].value, "Not generated")
                self.assertEqual(plots_worksheet["E3"].value, "Open wafer PNG")
                self.assertIsNone(plots_worksheet["F1"].value)
                self.assertIsNone(plots_worksheet["F3"].value)

                txpa_row = data_rows[0]
                self.assertEqual(txpa_row[0], "TXPA")
                self.assertEqual(txpa_row[1], 520123)
                self.assertEqual(txpa_row[2], "TXPA_OUTPUT_PWR")
                self.assertEqual(txpa_row[4], "View")
                self.assertEqual(txpa_row[6], 1.2)
                self.assertEqual(txpa_row[7], 2)
                self.assertEqual(txpa_row[8], "YES")
                self.assertEqual(txpa_row[9], "YES")
                self.assertEqual(txpa_row[10], "NO")
                self.assertEqual(txpa_row[11], "NO")
                self.assertEqual(txpa_row[12], 1)
                self.assertEqual(txpa_row[13], "YES")
                self.assertIn("large spread", str(txpa_row[14]).lower())
                self.assertEqual(txpa_row[15], 0)
                self.assertEqual(txpa_row[16], 4)
                self.assertEqual(txpa_row[23], None)
            finally:
                workbook.close()

            workbook_with_formatting = load_workbook(yield_report, read_only=False, data_only=False)
            try:
                worksheet_with_formatting = workbook_with_formatting[SAMPLE_SHEET_NAME]
                overview_with_formatting = workbook_with_formatting[OVERVIEW_SHEET_NAME]
                cf_rules = list(worksheet_with_formatting.conditional_formatting)
                overview_cf_rules = list(overview_with_formatting.conditional_formatting)
                self.assertTrue(cf_rules, "Expected conditional formatting rules in yield workbook")
                self.assertTrue(
                    any(str(rule.sqref) in {"H2", "H2:H2"} for rule in cf_rules),
                    "Expected Failing Chips conditional formatting on column H",
                )
                self.assertTrue(
                    any(str(rule.sqref) in {"I2", "I2:I2"} for rule in cf_rules),
                    "Expected Fails YES/NO conditional formatting on column I",
                )
                self.assertTrue(
                    any(str(rule.sqref) in {"J2", "J2:J2"} for rule in cf_rules),
                    "Expected Cpk<1.67 YES/NO conditional formatting on column J",
                )
                self.assertTrue(
                    any(str(rule.sqref) in {"M2", "M2:M2"} for rule in cf_rules),
                    "Expected Multimodality conditional formatting on column M",
                )
                self.assertEqual(worksheet_with_formatting["I2"].fill.fgColor.rgb, "00FFC7CE")
                self.assertEqual(worksheet_with_formatting["J2"].fill.fgColor.rgb, "00FFC7CE")
                self.assertEqual(worksheet_with_formatting["K2"].fill.fgColor.rgb, "00C6EFCE")
                self.assertEqual(worksheet_with_formatting["M2"].fill.fgColor.rgb, "00C6EFCE")
                self.assertEqual(worksheet_with_formatting["N2"].fill.fgColor.rgb, "00C6EFCE")
                self.assertEqual(worksheet_with_formatting["R2"].number_format, "0.######")
                self.assertEqual(worksheet_with_formatting["S2"].number_format, "0.######")
                self.assertTrue(bool(worksheet_with_formatting.column_dimensions["R"].hidden))
                self.assertTrue(bool(worksheet_with_formatting.column_dimensions["S"].hidden))
                self.assertTrue(bool(worksheet_with_formatting.column_dimensions["T"].hidden))
                self.assertTrue(bool(worksheet_with_formatting.column_dimensions["U"].hidden))
                self.assertTrue(bool(worksheet_with_formatting.column_dimensions["V"].hidden))
                self.assertTrue(bool(worksheet_with_formatting.column_dimensions["W"].hidden))
                self.assertFalse(bool(worksheet_with_formatting.column_dimensions["E"].hidden))
                self.assertFalse(bool(worksheet_with_formatting.column_dimensions["X"].hidden))
                self.assertIsNone(worksheet_with_formatting["X2"].fill.patternType)
                self.assertLessEqual(float(worksheet_with_formatting.column_dimensions["I"].width), 12.0)
                self.assertLessEqual(float(worksheet_with_formatting.column_dimensions["J"].width), 12.0)
                self.assertLessEqual(float(worksheet_with_formatting.column_dimensions["K"].width), 12.0)
                self.assertGreater(float(worksheet_with_formatting.column_dimensions["C"].width), 12.0)
                self.assertTrue(
                    any(str(rule.sqref) in {"D22:D23", "D22:D23"} for rule in overview_cf_rules),
                    "Expected module-level Fails color scale formatting on column D",
                )
                self.assertTrue(
                    any(str(rule.sqref) in {"H22:H23", "H22:H23"} for rule in overview_cf_rules),
                    "Expected module-level Unique Values color scale formatting on column H",
                )
                self.assertTrue(
                    any(str(rule.sqref) in {"I22:I23", "I22:I23"} for rule in overview_cf_rules),
                    "Expected module-level Multimodality color scale formatting on column I",
                )
            finally:
                workbook_with_formatting.close()

            with zipfile.ZipFile(yield_report, "r") as zf:
                drawing_xml_names = [name for name in zf.namelist() if name.startswith("xl/drawings/drawing") and name.endswith(".xml")]
                drawing_rels_names = [
                    name for name in zf.namelist() if name.startswith("xl/drawings/_rels/drawing") and name.endswith(".xml.rels")
                ]
                self.assertTrue(drawing_xml_names, "Expected drawing XML parts in yield workbook")
                self.assertTrue(drawing_rels_names, "Expected drawing relationship parts in yield workbook")

                drawing_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
                package_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

                click_targets: list[str] = []
                for rel_name in drawing_rels_names:
                    rel_root = ET.fromstring(zf.read(rel_name))
                    for rel in rel_root.findall(f"{{{package_ns}}}Relationship"):
                        if rel.attrib.get("Type", "").endswith("/hyperlink"):
                            click_targets.append(rel.attrib.get("Target", ""))

                self.assertGreaterEqual(len(click_targets), 2, "Expected clickable hyperlinks for embedded plots")
                self.assertTrue(any(target.endswith(".png") for target in click_targets))
                self.assertFalse(any(target.endswith(".html") for target in click_targets))

                click_count = 0
                for drawing_name in drawing_xml_names:
                    drawing_root = ET.fromstring(zf.read(drawing_name))
                    click_count += len(drawing_root.findall(f".//{{{drawing_ns}}}hlinkClick"))

                self.assertGreaterEqual(click_count, 2, "Expected embedded plot images to have click hyperlinks")

            corr_workbook = load_workbook(correlation_report, read_only=True, data_only=True)
            try:
                self.assertIn(SAMPLE_SHEET_NAME, corr_workbook.sheetnames)
                corr_sheet = corr_workbook[SAMPLE_SHEET_NAME]
                corr_rows = list(corr_sheet.iter_rows(min_row=2, values_only=True))
                self.assertGreaterEqual(len(corr_rows), 2, "Expected correlation rows in smoke report")
            finally:
                corr_workbook.close()


if __name__ == "__main__":
    unittest.main()