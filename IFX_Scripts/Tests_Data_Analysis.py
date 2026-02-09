"""Flat CSV test data analysis + reporting.

This script is designed for large, semicolon-separated "flat" CSV exports where:
- The header row contains metadata columns (e.g. LOT/WAFER/X/Y/SITE_NUM) and many
  numeric test columns (e.g. 520123, 530045, ...).
- A small block of meta rows follows the header ("Test Name", "Low", "High",
  "Unit", "Cpk", "Yield", ...).
- Unit/device rows come after that meta block, with measurements per test column.

The report is module-agnostic: a test's "module" is extracted from the first 4
characters of its test name (e.g. DPLL, TXPA, TXLO). Modules to analyze are
provided via a user configuration section at the top of this file.

Outputs:
- One Excel workbook containing one sheet per input file with only the tests of
  interest that have yield < 100% or Cpk outside thresholds.
- A per-input-file plots sheet embedding CDF plots; hyperlinks in the data sheet
  jump to the embedded plot.
- Optionally, a separate correlation workbook (one sheet per input file)
  computing Pearson and Spearman correlations for each module test vs all tests.

Note on "hover previews": Excel doesn't support showing an image on hyperlink
hover via openpyxl. This implementation embeds images in a dedicated plots sheet
and links to those locations.
"""

from __future__ import annotations
import csv
import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Literal


DEFAULT_ENCODING = "latin1"
DELIMITER = ";"


# ================================
# USER PARAMETERS (edit here)
# ================================
REPO_ROOT = Path(__file__).resolve().parents[2]

# Input/Output
INPUT_FOLDER: Path = Path(r"C:/UserData/Infineon/TE_CTRX/CTRX_8188_8144/Data_Reviews/FE_Test_D")
OUTPUT_FOLDER: Path = Path(r"C:/UserData/Infineon/TE_CTRX/CTRX_8188_8144/Data_Reviews/FE_Test_D")

# Which modules to analyze (first 4 chars of test name)
MODULES: list[str] = ["TXGE","TXVC","DPLL","TXPA","TXPB","TXPC","TXPD","TXLO","TXPS"]

# Thresholds
YIELD_THRESHOLD: float = 100.0
CPK_LOW: float = 1.67
CPK_HIGH: float = 20.0

# Outlier detection (|x-median| > OUTLIER_MAD_MULTIPLIER * MAD)
OUTLIER_MAD_MULTIPLIER: float = 6.0

# Optional controls
MAX_FILES: int | None = None  # e.g. 1 for quick test
SINGLE_FILE: str | None = None  # e.g. "...csv" to process one file
ENCODING: str = DEFAULT_ENCODING

# Optional correlation workbook
GENERATE_CORRELATION_REPORT: bool = False
CORRELATION_METHODS: list[Literal["pearson", "spearman"]] = ["pearson", "spearman"]

# Wafer map display controls
# Scales the *area* of the wafer outline circle; 2.0 => 2× area, 3.0 => 3× area.
WAFERMAP_CIRCLE_AREA_MULT: float = 1.0


def _as_path(value: str | os.PathLike[str] | Path) -> Path:
    return value if isinstance(value, Path) else Path(value)


def _excel_col_letter(col_idx_1_based: int) -> str:
    """Convert a 1-based column index to an Excel column letter."""
    if col_idx_1_based < 1:
        raise ValueError("Column index must be >= 1")
    n = col_idx_1_based
    letters: list[str] = []
    while n:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def _safe_sheet_name(name: str) -> str:
    # Excel sheet limit is 31 chars and cannot contain some characters.
    safe = re.sub(r"[\\/*?:\[\]]", "_", name)
    safe = safe.strip() or "Sheet"
    return safe[:31]


def _autofit_openpyxl_columns(ws, *, min_width: int = 8, max_width: int = 70, padding: int = 2) -> None:
    if ws.max_column < 1 or ws.max_row < 1:
        return

    for col_idx, col_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
        start=1,
    ):
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            text = str(cell.value)
            if "\n" in text:
                text = max(text.splitlines(), key=len)
            max_len = max(max_len, len(text))

        width = max(min_width, min(max_width, max_len + padding))
        ws.column_dimensions[_excel_col_letter(col_idx)].width = width


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    text = str(value).strip().strip('"')
    if text == "" or text.lower() in {"nan", "na", "none"}:
        return None
    try:
        return float(text)
    except ValueError:
        # Handle comma decimals.
        try:
            return float(text.replace(",", "."))
        except ValueError:
            return None


def _module_from_test_name(test_name: str) -> str:
    if not test_name:
        return ""
    s = str(test_name).strip()
    if len(s) < 4:
        return s.upper()
    return s[:4].upper()


@dataclass(frozen=True)
class FlatFileMeta:
    header: list[str]
    numeric_test_cols: list[str]
    data_start_line_index: int  # 0-based line index where unit data begins
    meta_rows: dict[str, dict[str, str]]  # row_name -> {col_name -> raw string}


def scan_flat_file_meta(
    file_path: Path,
    *,
    encoding: str = DEFAULT_ENCODING,
    delimiter: str = DELIMITER,
    needed_meta_rows: Iterable[str] = ("Test Name", "Low", "High", "Unit", "Cpk", "Yield", "Mean", "Stddev"),
    max_scan_lines: int = 200,
) -> FlatFileMeta:
    needed = {r.strip() for r in needed_meta_rows}
    meta_rows: dict[str, dict[str, str]] = {}

    with file_path.open("r", encoding=encoding, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError(f"Empty file: {file_path}")

        header = [h.strip() for h in header]
        numeric_test_cols = [h for h in header if h.strip().isdigit()]
        if not numeric_test_cols:
            raise ValueError("Could not find numeric test columns in header")

        data_start_line_index = 1
        for line_idx, row in enumerate(reader, start=1):
            if line_idx >= max_scan_lines:
                break
            if not row:
                continue
            key = (row[0] or "").strip().strip('"')
            if key.isdigit():
                data_start_line_index = line_idx
                break

            if key in needed:
                row_map: dict[str, str] = {}
                # Zip may truncate; that's ok.
                for col_name, cell in zip(header, row, strict=False):
                    if col_name in numeric_test_cols:
                        row_map[col_name] = (cell or "").strip()
                meta_rows[key] = row_map

        else:
            # If we didn't break on a unit row, we still set start line to after scanned block.
            data_start_line_index = max_scan_lines

    return FlatFileMeta(
        header=header,
        numeric_test_cols=numeric_test_cols,
        data_start_line_index=data_start_line_index,
        meta_rows=meta_rows,
    )


def _mad(values):
    import numpy as np

    med = np.nanmedian(values)
    return float(np.nanmedian(np.abs(values - med)))


def _robust_sigma(values) -> float:
    import numpy as np

    m = _mad(values)
    if m > 0:
        return 1.4826 * m
    # Fallback to std if MAD collapses.
    return float(np.nanstd(values, ddof=1))


def _count_hist_peaks(values) -> int:
    """Very lightweight multimodality heuristic based on histogram peaks."""
    import numpy as np

    v = values[np.isfinite(values)]
    if v.size < 80:
        return 1
    q25, q75 = np.percentile(v, [25, 75])
    iqr = q75 - q25
    if iqr <= 0:
        return 1
    bin_width = 2 * iqr / (v.size ** (1 / 3))
    if not np.isfinite(bin_width) or bin_width <= 0:
        return 1
    bins = max(10, min(80, int((v.max() - v.min()) / bin_width) + 1))
    hist, _ = np.histogram(v, bins=bins)
    if hist.size < 5:
        return 1

    # Smooth with a small moving average.
    kernel = np.array([1, 2, 3, 2, 1], dtype=float)
    kernel /= kernel.sum()
    smooth = np.convolve(hist.astype(float), kernel, mode="same")
    # Local maxima with a basic prominence threshold.
    threshold = max(5.0, 0.10 * float(smooth.max()))
    peaks = 0
    for i in range(1, len(smooth) - 1):
        if smooth[i] > smooth[i - 1] and smooth[i] > smooth[i + 1] and smooth[i] >= threshold:
            peaks += 1
    return max(1, peaks)


def _cdf_plot_png(
    values,
    *,
    title: str,
    out_path: Path,
    low_limit: float | None = None,
    high_limit: float | None = None,
    proposed_l6: float | None = None,
    proposed_u6: float | None = None,
    proposed_l12: float | None = None,
    proposed_u12: float | None = None,
) -> None:
    import pandas as pd
    import numpy as np

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        # Plotting is optional; if matplotlib isn't available just skip.
        return

    v = pd.to_numeric(values, errors="coerce").dropna().to_numpy(dtype=float)
    v = v[np.isfinite(v)]
    if v.size == 0:
        return
    v.sort()
    y = np.arange(1, v.size + 1) / v.size

    mean_v = float(np.mean(v))
    median_v = float(np.median(v))

    fig, ax = plt.subplots(figsize=(7.0, 4.0), dpi=140)

    # CDF points (not a continuous line)
    ax.plot(v, y, linestyle="None", marker=".", markersize=3.0, alpha=0.85, label="Data")

    # Spec limits (red)
    if low_limit is not None and np.isfinite(low_limit):
        ax.axvline(float(low_limit), color="#D62728", linestyle="-", linewidth=1.6, label=f"LTL={_fmt_num(float(low_limit))}")
    if high_limit is not None and np.isfinite(high_limit):
        ax.axvline(float(high_limit), color="#D62728", linestyle="-", linewidth=1.6, label=f"UTL={_fmt_num(float(high_limit))}")

    # Proposed 6σ/12σ limits for Cpk-issue tests (orange/purple)
    if proposed_l6 is not None and np.isfinite(proposed_l6):
        ax.axvline(float(proposed_l6), color="#FF7F0E", linestyle="--", linewidth=1.2, label=f"LTL 6s={_fmt_1dp(float(proposed_l6))}")
    if proposed_u6 is not None and np.isfinite(proposed_u6):
        ax.axvline(float(proposed_u6), color="#FF7F0E", linestyle="--", linewidth=1.2, label=f"UTL 6s={_fmt_1dp(float(proposed_u6))}")
    if proposed_l12 is not None and np.isfinite(proposed_l12):
        ax.axvline(float(proposed_l12), color="#9467BD", linestyle=":", linewidth=1.2, label=f"LTL 12s={_fmt_1dp(float(proposed_l12))}")
    if proposed_u12 is not None and np.isfinite(proposed_u12):
        ax.axvline(float(proposed_u12), color="#9467BD", linestyle=":", linewidth=1.2, label=f"UTL 12s={_fmt_1dp(float(proposed_u12))}")

    ax.grid(True, alpha=0.3)
    ax.set_title(title)
    ax.set_xlabel("Value")
    ax.set_ylabel("CDF")
    ax.legend(loc="best", fontsize=8, framealpha=0.9)
    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, format="png")
    plt.close(fig)


def _wafer_map_png(
    values,
    *,
    meta_cols,
    title: str,
    out_path: Path,
    low_limit: float | None = None,
    high_limit: float | None = None,
) -> None:
    """Create a wafer map scatter plot using X/Y and (optionally) WAFER.

    Colors show the test value distribution; spec-fail points are highlighted.
    """
    import pandas as pd
    import numpy as np

    if meta_cols is None:
        return
    if not all(c in getattr(meta_cols, "columns", []) for c in ("X", "Y")):
        return

    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.patches import Circle
        import matplotlib.colors as mcolors
    except Exception:
        return

    import warnings

    area_mult = float(WAFERMAP_CIRCLE_AREA_MULT) if "WAFERMAP_CIRCLE_AREA_MULT" in globals() else 3.0
    if not np.isfinite(area_mult) or area_mult <= 0:
        area_mult = 3.0

    v = pd.to_numeric(values, errors="coerce")
    x = pd.to_numeric(meta_cols["X"], errors="coerce")
    y = pd.to_numeric(meta_cols["Y"], errors="coerce")
    wafer = meta_cols["WAFER"].astype("string").str.strip() if "WAFER" in meta_cols.columns else None

    df = pd.DataFrame({"v": v, "X": x, "Y": y})
    if wafer is not None:
        wafer = wafer.mask(wafer.eq("") | wafer.str.lower().eq("nan"))
        df["WAFER"] = wafer

    df = df.dropna(subset=["v", "X", "Y"]).copy()
    if df.empty:
        return

    # Choose wafers to plot.
    if "WAFER" in df.columns and df["WAFER"].notna().any():
        counts = df.dropna(subset=["WAFER"]).groupby("WAFER")["v"].size().sort_values(ascending=False)
        wafers = [str(w) for w in counts.index.tolist()]
    else:
        wafers = ["ALL"]
        df["WAFER"] = "ALL"

    max_wafers = 6
    wafers = wafers[:max_wafers]

    # Shared color scaling across subplots.
    all_v = df["v"].to_numpy(dtype=float)
    vmin = float(np.nanpercentile(all_v, 1))
    vmax = float(np.nanpercentile(all_v, 99))
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmin >= vmax:
        vmin = float(np.nanmin(df["v"]))
        vmax = float(np.nanmax(df["v"]))

    # Vivid multi-color gradient that makes extremes pop.
    try:
        cmap = plt.get_cmap("turbo")
    except Exception:
        cmap = plt.get_cmap("viridis")

    norm = mcolors.Normalize(vmin=vmin, vmax=vmax, clip=True)

    n = len(wafers)
    ncols = 3 if n >= 3 else n
    nrows = int(math.ceil(n / max(1, ncols)))
    fig_w = 7.0
    fig_h = 4.2 if n <= 2 else 4.8
    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(fig_w, fig_h), dpi=140)
    if not isinstance(axes, np.ndarray):
        axes = np.array([axes])
    axes = axes.ravel()

    # Reserve a right margin for the colorbar so it never covers the wafer maps.
    fig.subplots_adjust(left=0.07, right=0.84, bottom=0.10, top=0.86, wspace=0.25, hspace=0.35)

    low = -np.inf if low_limit is None else float(low_limit)
    high = np.inf if high_limit is None else float(high_limit)

    mappable = None
    for ax, w in zip(axes, wafers, strict=False):
        d = df[df["WAFER"].astype(str) == str(w)]
        if d.empty:
            ax.axis("off")
            continue

        xv = d["X"].to_numpy(dtype=float)
        yv = d["Y"].to_numpy(dtype=float)
        vv = d["v"].to_numpy(dtype=float)

        fails = (vv < low) | (vv > high)
        n_fail = int(np.count_nonzero(fails))

        sc = ax.scatter(
            xv,
            yv,
            c=vv,
            cmap=cmap,
            norm=norm,
            s=18,
            marker="s",
            linewidths=0,
            alpha=0.95,
        )
        mappable = sc

        if np.any(fails):
            ax.scatter(
                xv[fails],
                yv[fails],
                facecolors="none",
                edgecolors="#D62728",
                linewidths=1.2,
                s=65,
                marker="s",
                label=f"fails={n_fail}",
            )

        # Simple wafer outline (circle around the die cloud).
        cx = float(np.mean(xv))
        cy = float(np.mean(yv))
        base_r = float(np.max(np.sqrt((xv - cx) ** 2 + (yv - cy) ** 2))) + 0.8
        r = base_r * float(math.sqrt(area_mult))
        ax.add_patch(Circle((cx, cy), r, fill=False, color="#666666", linewidth=0.8, alpha=0.8))

        ax.set_title(f"WAFER={w}  N={len(vv)}  fails={n_fail}", fontsize=9)
        ax.set_aspect("equal", adjustable="box")
        ax.grid(True, alpha=0.15)
        ax.tick_params(labelsize=8)

    # Hide any unused axes.
    for ax in axes[len(wafers) :]:
        ax.axis("off")

    if mappable is not None:
        # Dedicated axis for colorbar (no overlap with subplot area).
        cax = fig.add_axes([0.87, 0.18, 0.03, 0.62])
        cbar = fig.colorbar(mappable, cax=cax)
        cbar.ax.tick_params(labelsize=8)
        cbar.set_label("Value", fontsize=9)

    fig.suptitle(title, fontsize=10)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, format="png")
    plt.close(fig)


def _parse_filename_wafer_signature(file_name: str) -> str | None:
    m = re.search(r"\b(S11P|S21P|S31P|B11P|B21P)\b", file_name, flags=re.IGNORECASE)
    if not m:
        return None
    return m.group(1).upper()


def _parse_insertion_temperature_label(file_name: str) -> str:
    """Best-effort extraction of insertion temperature from file name."""
    upper = file_name.upper()
    if "S11P" in upper or "B11P" in upper or "HT" in upper or "Q3" in upper:
        return "Hot (135°C)"
    if "S21P" in upper or "Q2" in upper:
        return "Cold (-40°C)"
    if "S31P" in upper or "B21P" in upper or "RT" in upper or "Q1" in upper:
        return "Ambient (25°C)"
    return "Unknown"


def _fmt_num(x: float) -> str:
    return f"{float(x):.6g}"


def _fmt_1dp(x: float) -> str:
    return f"{float(x):.1f}"


def _build_plot_title(
    *,
    test_name: str,
    test_col: str,
    temp_label: str,
    mean_v: float,
    median_v: float,
) -> str:
    first = f"{test_name} ({test_col}) | {temp_label}"
    return first + "\n" + f"mean={_fmt_num(mean_v)}; median={_fmt_num(median_v)}"


def _apply_module_group_row_colors(ws, *, module_col_header: str = "Module", first_data_row: int = 2) -> None:
    """Apply alternating light fills per contiguous module block (for filtering/readability)."""
    try:
        from openpyxl.styles import PatternFill
    except Exception:
        return

    if ws.max_row < first_data_row or ws.max_column < 1:
        return

    module_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        if str(ws.cell(row=1, column=col_idx).value).strip() == module_col_header:
            module_col_idx = col_idx
            break
    if module_col_idx is None:
        return

    palette = [
        "D9E1F2",  # light blue
        "E2EFDA",  # light green
        "FFF2CC",  # light yellow
        "FCE4D6",  # light orange
        "EDEDED",  # light gray
    ]

    current_module = None
    color_idx = -1
    fill = None
    for row_idx in range(first_data_row, ws.max_row + 1):
        module_val = ws.cell(row=row_idx, column=module_col_idx).value
        module_str = "" if module_val is None else str(module_val)
        if module_str != current_module:
            current_module = module_str
            color_idx = (color_idx + 1) % len(palette)
            fill = PatternFill(patternType="solid", fgColor=palette[color_idx])

        if fill is None:
            continue
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _build_comment(
    *,
    series,
    meta_cols,
    outlier_mad_multiplier: float,
    low_limit: float | None,
    high_limit: float | None,
    wafer_sig: str | None,
) -> str:
    import numpy as np
    import pandas as pd

    vals = pd.to_numeric(series, errors="coerce")
    finite = vals.dropna().to_numpy(dtype=float)
    finite = finite[np.isfinite(finite)]
    if finite.size == 0:
        return "No numeric data"

    med = float(np.median(finite))
    mad = _mad(finite)
    sig = _robust_sigma(finite)

    parts: list[str] = []

    if mad > 0:
        outliers = np.abs(finite - med) > (outlier_mad_multiplier * mad)
        n_out = int(outliers.sum())
        if n_out > 0:
            parts.append(f"Outliers: {n_out}/{finite.size} (>{outlier_mad_multiplier:g}×MAD)")

    # Instability / large spread heuristic.
    if sig > 0 and abs(med) > 0:
        robust_cv = sig / abs(med)
        if robust_cv >= 0.05:
            parts.append(f"Large spread (robust CV={robust_cv:.2%})")

    # Multi-modality heuristic.
    peaks = _count_hist_peaks(finite)
    if peaks >= 2:
        parts.append(f"Possible multi-modal distribution (peaks≈{peaks})")

    # Site effect heuristic.
    if "SITE_NUM" in meta_cols.columns:
        df_site = pd.DataFrame({"v": vals, "SITE_NUM": pd.to_numeric(meta_cols["SITE_NUM"], errors="coerce")})
        df_site = df_site.dropna(subset=["v", "SITE_NUM"])
        if not df_site.empty and df_site["SITE_NUM"].nunique() >= 2:
            med_by_site = df_site.groupby("SITE_NUM")["v"].median().sort_index()
            rng = float(med_by_site.max() - med_by_site.min())
            if sig > 0 and rng / sig >= 3.0:
                worst = med_by_site.idxmax() if abs(med_by_site.max() - med) >= abs(med_by_site.min() - med) else med_by_site.idxmin()
                parts.append(f"Site effect (site medians range={rng:.4g}, worst site={int(worst)})")

    # Wafer signature heuristics.
    if wafer_sig:
        parts.append(f"File wafer signature: {wafer_sig}")

    if "WAFER" in meta_cols.columns:
        df_w = pd.DataFrame({"v": vals, "WAFER": meta_cols["WAFER"].astype(str)})
        df_w = df_w.dropna(subset=["v"])  # keep WAFER even if blank
        wafer_series = df_w["WAFER"].astype("string").str.strip()
        wafer_series = wafer_series.mask(wafer_series.eq("") | wafer_series.str.lower().eq("nan"))
        wafers = wafer_series.dropna().unique()
        if wafers.size >= 2:
            med_by_wafer = df_w.dropna(subset=["WAFER"]).groupby("WAFER")["v"].median()
            rng = float(med_by_wafer.max() - med_by_wafer.min())
            if sig > 0 and rng / sig >= 3.0:
                parts.append("Wafer signature suspected (median shifts across wafers)")

    # Coordinate signature (very lightweight).
    for axis in ("X", "Y"):
        if axis in meta_cols.columns:
            coord = pd.to_numeric(meta_cols[axis], errors="coerce")
            df_xy = pd.DataFrame({"v": vals, axis: coord}).dropna()
            if df_xy.shape[0] >= 50:
                rho = df_xy["v"].corr(df_xy[axis], method="spearman")
                if rho is not None and np.isfinite(rho) and abs(float(rho)) >= 0.30:
                    parts.append(f"Coordinate signature: spearman(v,{axis})={float(rho):+.2f}")

    # Fail clustering note.
    if low_limit is not None or high_limit is not None:
        low = -np.inf if low_limit is None else low_limit
        high = np.inf if high_limit is None else high_limit
        fails = (finite < low) | (finite > high)
        n_fail = int(fails.sum())
        if n_fail > 0:
            parts.append(f"Spec fails: {n_fail}/{finite.size}")

    return "; ".join(parts) if parts else "OK"


def _read_unit_data(
    file_path: Path,
    *,
    data_start_line_index: int,
    usecols: list[str],
    encoding: str = DEFAULT_ENCODING,
) -> Any:
    import pandas as pd

    # Keep header (line 0), skip meta rows 1..data_start_line_index-1.
    def _skip(i: int) -> bool:
        return 0 < i < data_start_line_index

    df = pd.read_csv(
        file_path,
        sep=DELIMITER,
        encoding=encoding,
        low_memory=False,
        usecols=usecols,
        skiprows=_skip,
    )
    df.columns = [c.strip() for c in df.columns]
    return df


def _limits_from_meta(meta: FlatFileMeta, test_col: str) -> tuple[float | None, float | None, str | None]:
    low_raw = meta.meta_rows.get("Low", {}).get(test_col)
    high_raw = meta.meta_rows.get("High", {}).get(test_col)
    unit_raw = meta.meta_rows.get("Unit", {}).get(test_col)
    return _to_float(low_raw), _to_float(high_raw), (unit_raw.strip() if unit_raw else None)


def _yield_cpk_from_meta(meta: FlatFileMeta, test_col: str) -> tuple[float | None, float | None]:
    y_raw = meta.meta_rows.get("Yield", {}).get(test_col)
    c_raw = meta.meta_rows.get("Cpk", {}).get(test_col)
    return _to_float(y_raw), _to_float(c_raw)


def _test_name_from_meta(meta: FlatFileMeta, test_col: str) -> str:
    return (meta.meta_rows.get("Test Name", {}).get(test_col) or "").strip().strip('"')


def _status_for_test(
    *,
    yield_pct: float | None,
    cpk: float | None,
    yield_threshold: float,
    cpk_low: float,
    cpk_high: float,
) -> str | None:
    if yield_pct is not None and yield_pct < yield_threshold:
        return "FAILS"
    if cpk is not None and cpk < cpk_low:
        return f"Cpk<{cpk_low:g}"
    if cpk is not None and cpk > cpk_high:
        return f"Cpk>{cpk_high:g}"
    return None


def _proposed_sigma_limits(values) -> tuple[float | None, float | None, float | None, float | None]:
    import numpy as np
    import pandas as pd

    v = pd.to_numeric(values, errors="coerce").dropna().to_numpy(dtype=float)
    v = v[np.isfinite(v)]
    if v.size == 0:
        return None, None, None, None
    med = float(np.median(v))
    sig = _robust_sigma(v)
    if not np.isfinite(sig) or sig <= 0:
        return med, med, med, med

    def _floor_1dp(x: float) -> float:
        return math.floor(x * 10.0) / 10.0

    def _ceil_1dp(x: float) -> float:
        return math.ceil(x * 10.0) / 10.0

    l6, u6 = med - 6.0 * sig, med + 6.0 * sig
    l12, u12 = med - 12.0 * sig, med + 12.0 * sig

    # Quantize to 1 decimal: conservative rounding (floor lows, ceil uppers).
    l6 = _floor_1dp(float(l6))
    u6 = _ceil_1dp(float(u6))
    l12 = _floor_1dp(float(l12))
    u12 = _ceil_1dp(float(u12))

    return l6, u6, l12, u12


def generate_yield_cpk_report(
    *,
    input_folder: Path,
    output_folder: Path,
    modules: list[str],
    outlier_mad_multiplier: float,
    yield_threshold: float,
    cpk_low: float,
    cpk_high: float,
    max_files: int | None,
    single_file: str | None,
    encoding: str = DEFAULT_ENCODING,
) -> Path:
    import numpy as np
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font

    output_folder.mkdir(parents=True, exist_ok=True)
    plots_root = output_folder / "cdf_plots"
    plots_root.mkdir(parents=True, exist_ok=True)

    if single_file:
        csv_paths = [input_folder / single_file]
    else:
        csv_paths = sorted([p for p in input_folder.glob("*.csv") if p.is_file()])
    if max_files is not None:
        csv_paths = csv_paths[:max_files]
    if not csv_paths:
        raise SystemExit(f"No .csv files found in: {input_folder}")

    modules_upper = {m.strip().upper() for m in modules if m.strip()}
    if not modules_upper:
        raise SystemExit("No modules provided. Example: --modules DPLL,TXPA,TXLO")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for file_path in csv_paths:
        print(f"Processing: {file_path.name}")
        meta = scan_flat_file_meta(file_path, encoding=encoding)
        wafer_sig = _parse_filename_wafer_signature(file_path.name)
        temp_label = _parse_insertion_temperature_label(file_path.name)

        # Determine tests of interest by module prefix from Test Name row.
        interest_cols: list[str] = []
        interest_names: dict[str, str] = {}
        interest_modules: dict[str, str] = {}
        for test_col in meta.numeric_test_cols:
            tn = _test_name_from_meta(meta, test_col)
            mod = _module_from_test_name(tn)
            if mod in modules_upper:
                interest_cols.append(test_col)
                interest_names[test_col] = tn
                interest_modules[test_col] = mod

        if not interest_cols:
            print(f"  - No tests matched modules {sorted(modules_upper)}; skipping")
            continue

        # Identify affected tests based on Yield/Cpk thresholds.
        affected: list[str] = []
        status_by_col: dict[str, str] = {}
        yield_by_col: dict[str, float | None] = {}
        cpk_by_col: dict[str, float | None] = {}
        for test_col in interest_cols:
            y, c = _yield_cpk_from_meta(meta, test_col)
            yield_by_col[test_col] = y
            cpk_by_col[test_col] = c
            status = _status_for_test(
                yield_pct=y,
                cpk=c,
                yield_threshold=yield_threshold,
                cpk_low=cpk_low,
                cpk_high=cpk_high,
            )
            if status:
                affected.append(test_col)
                status_by_col[test_col] = status

        # Group affected tests by module for readability/formatting.
        affected.sort(key=lambda c: (interest_modules.get(c, ""), int(c)))

        if not affected:
            print("  - No affected tests (per thresholds); skipping sheet")
            continue

        # Load only needed columns from unit data.
        wanted_meta_cols = [
            c
            for c in ("SITE_NUM", "WAFER", "X", "Y", "LOT", "SUBLOT", "CHIP_ID", "PF", "FIRST_FAIL_TEST")
            if c in meta.header
        ]
        usecols = wanted_meta_cols + affected
        df_units = _read_unit_data(
            file_path,
            data_start_line_index=meta.data_start_line_index,
            usecols=usecols,
            encoding=encoding,
        )

        # Ensure meta col dataframe aligns to series indices for comment generation.
        meta_cols_df = df_units[wanted_meta_cols].copy() if wanted_meta_cols else pd.DataFrame(index=df_units.index)

        # Create sheets.
        sheet_name = _safe_sheet_name(file_path.stem)
        ws = wb.create_sheet(sheet_name)
        plot_sheet_name = _safe_sheet_name((file_path.stem[:25] + "_PLOTS"))
        ws_plots = wb.create_sheet(plot_sheet_name)

        # Tab colors: keep data vs plots tabs distinct.
        ws.sheet_properties.tabColor = "4F81BD"  # blue
        ws_plots.sheet_properties.tabColor = "C0504D"  # red

        headers = [
            "File",
            "Module",
            "Test Nr",
            "Test Name",
            "Unit",
            "Yield (%)",
            "Cpk",
            "Status",
            "N",
            "Outliers",
            "LTL 6s",
            "UTL 6s",
            "LTL 12s",
            "UTL 12s",
            "Comment",
            "CDF Plot",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws_plots.append(["Test", "CDF (embedded)", "Wafer map (fails highlighted)"])
        ws_plots["A1"].font = Font(bold=True)
        ws_plots["B1"].font = Font(bold=True)
        ws_plots["C1"].font = Font(bold=True)

        plot_anchor_row = 3
        out_rows = 0
        for test_col in affected:
            test_name = interest_names.get(test_col, "")
            module = interest_modules.get(test_col, "")
            low, high, unit = _limits_from_meta(meta, test_col)
            y = yield_by_col.get(test_col)
            c = cpk_by_col.get(test_col)
            status = status_by_col[test_col]

            series = df_units[test_col]
            numeric = pd.to_numeric(series, errors="coerce")
            finite = numeric.dropna().to_numpy(dtype=float)
            finite = finite[np.isfinite(finite)]
            n = int(finite.size)
            if n == 0:
                continue

            # Outliers
            med = float(np.median(finite))
            mad = _mad(finite)
            n_out = 0
            if mad > 0:
                n_out = int((np.abs(finite - med) > (outlier_mad_multiplier * mad)).sum())

            l6, u6, l12, u12 = _proposed_sigma_limits(numeric)
            comment = _build_comment(
                series=numeric,
                meta_cols=meta_cols_df,
                outlier_mad_multiplier=outlier_mad_multiplier,
                low_limit=low,
                high_limit=high,
                wafer_sig=wafer_sig,
            )

            # Plot file and embed in plot sheet.
            safe_test = re.sub(r"[^A-Za-z0-9._-]+", "_", test_name)[:80] or test_col
            plot_path = plots_root / file_path.stem / f"{test_col}_{safe_test}.png"
            include_sigma_limits = status.startswith("Cpk<") or status.startswith("Cpk>")
            title = _build_plot_title(
                test_name=test_name,
                test_col=test_col,
                temp_label=temp_label,
                mean_v=float(np.mean(finite)),
                median_v=float(np.median(finite)),
            )
            _cdf_plot_png(
                numeric,
                title=title,
                out_path=plot_path,
                low_limit=low,
                high_limit=high,
                proposed_l6=(l6 if include_sigma_limits else None),
                proposed_u6=(u6 if include_sigma_limits else None),
                proposed_l12=(l12 if include_sigma_limits else None),
                proposed_u12=(u12 if include_sigma_limits else None),
            )

            wafer_map_path = plots_root / file_path.stem / f"{test_col}_{safe_test}_wafermap.png"
            comment_text = str(comment)
            spatial_signature = (
                "Wafer signature suspected" in comment_text
                or "Coordinate signature" in comment_text
            )
            if spatial_signature:
                _wafer_map_png(
                    numeric,
                    meta_cols=meta_cols_df,
                    title=f"{test_name} ({test_col}) | {temp_label}",
                    out_path=wafer_map_path,
                    low_limit=low,
                    high_limit=high,
                )

            # Write to plots sheet.
            ws_plots[f"A{plot_anchor_row}"] = f"{test_col} {test_name}"
            ws_plots[f"A{plot_anchor_row}"].font = Font(bold=True)

            # Zoomable links (open the PNG in an external viewer).
            cdf_link = ws_plots[f"B{plot_anchor_row}"]
            cdf_link.value = "Open CDF PNG"
            if plot_path.exists():
                cdf_link.hyperlink = plot_path.resolve().as_uri()
                cdf_link.font = Font(color="0000EE", underline="single")

            wafer_link = ws_plots[f"C{plot_anchor_row}"]
            wafer_link.value = "Open wafer PNG"
            if wafer_map_path.exists():
                wafer_link.hyperlink = wafer_map_path.resolve().as_uri()
                wafer_link.font = Font(color="0000EE", underline="single")
            cdf_anchor_cell = f"B{plot_anchor_row + 1}"
            wafer_anchor_cell = f"K{plot_anchor_row + 1}"
            if plot_path.exists():
                img = XLImage(str(plot_path))
                img.width = 520
                img.height = 360
                ws_plots.add_image(img, cdf_anchor_cell)

            if wafer_map_path.exists():
                wimg = XLImage(str(wafer_map_path))
                wimg.width = 520
                wimg.height = 360
                ws_plots.add_image(wimg, wafer_anchor_cell)
            # Reserve some rows for the image.
            plot_link_target = f"#{plot_sheet_name}!{cdf_anchor_cell}"
            plot_anchor_row += 30

            # Write row to data sheet.
            row = [
                file_path.name,
                module,
                int(test_col),
                test_name,
                unit,
                y,
                c,
                status,
                n,
                n_out,
                l6,
                u6,
                l12,
                u12,
                comment,
                "View",
            ]
            ws.append(row)
            out_rows += 1

            # Force 1-decimal display for sigma limits in Excel.
            row_idx = 1 + out_rows
            for col_name in ("LTL 6s", "UTL 6s", "LTL 12s", "UTL 12s"):
                col_idx = headers.index(col_name) + 1
                ws.cell(row=row_idx, column=col_idx).number_format = "0.0"

            link_cell = ws.cell(row=1 + out_rows, column=headers.index("CDF Plot") + 1)
            link_cell.hyperlink = plot_link_target
            link_cell.font = Font(color="0000EE", underline="single")

        if out_rows == 0:
            # Avoid leaving empty sheets.
            wb.remove(ws)
            wb.remove(ws_plots)
            continue

        # Apply module block coloring on the data sheet.
        _apply_module_group_row_colors(ws)

        # Auto-fit columns.
        _autofit_openpyxl_columns(ws)
        _autofit_openpyxl_columns(ws_plots)

        # Freeze header row.
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{_excel_col_letter(ws.max_column)}{ws.max_row}"

    from datetime import datetime

    out_xlsx = output_folder / "Yield_Cpk_Report.xlsx"
    try:
        wb.save(out_xlsx)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = output_folder / f"Yield_Cpk_Report_{ts}.xlsx"
        wb.save(alt)
        print(f"Could not overwrite (file open?): {out_xlsx}")
        print(f"Saved instead: {alt}")
        return alt

    print(f"Saved: {out_xlsx}")
    return out_xlsx


def generate_correlation_workbook(
    *,
    input_folder: Path,
    output_folder: Path,
    modules: list[str],
    max_files: int | None,
    single_file: str | None,
    methods: list[Literal["pearson", "spearman"]],
    encoding: str = DEFAULT_ENCODING,
) -> Path:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font

    output_folder.mkdir(parents=True, exist_ok=True)
    if single_file:
        csv_paths = [input_folder / single_file]
    else:
        csv_paths = sorted([p for p in input_folder.glob("*.csv") if p.is_file()])
    if max_files is not None:
        csv_paths = csv_paths[:max_files]
    if not csv_paths:
        raise SystemExit(f"No .csv files found in: {input_folder}")

    modules_upper = {m.strip().upper() for m in modules if m.strip()}
    if not modules_upper:
        raise SystemExit("No modules provided.")

    wb = Workbook()
    wb.remove(wb.active)

    for file_path in csv_paths:
        print(f"Correlation: {file_path.name}")
        meta = scan_flat_file_meta(file_path, encoding=encoding)

        test_name_by_col = {c: _test_name_from_meta(meta, c) for c in meta.numeric_test_cols}
        module_by_col = {c: _module_from_test_name(test_name_by_col[c]) for c in meta.numeric_test_cols}
        module_cols = [c for c in meta.numeric_test_cols if module_by_col[c] in modules_upper]
        if not module_cols:
            print("  - No module tests found; skipping")
            continue

        # Load ALL numeric tests for correlation (optional heavy step).
        df = _read_unit_data(
            file_path,
            data_start_line_index=meta.data_start_line_index,
            usecols=meta.numeric_test_cols,
            encoding=encoding,
        )
        df = df.apply(pd.to_numeric, errors="coerce")

        sheet = wb.create_sheet(_safe_sheet_name(file_path.stem))
        headers = [
            "Module",
            "Test Nr",
            "Test Name",
            "Other Test Nr",
            "Other Test Name",
        ] + [m.title() for m in methods]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # Precompute ranks for spearman if needed.
        df_rank = None
        if "spearman" in methods:
            df_rank = df.rank(axis=0, method="average", na_option="keep")

        out_rows = 0
        for test_col in module_cols:
            s = df[test_col]
            if s.dropna().nunique() < 2:
                continue
            pearson = df.corrwith(s, method="pearson") if "pearson" in methods else None
            spearman = None
            if "spearman" in methods and df_rank is not None:
                spearman = df_rank.corrwith(df_rank[test_col], method="pearson")

            for other_col in meta.numeric_test_cols:
                if other_col == test_col:
                    continue
                row = [
                    module_by_col[test_col],
                    int(test_col),
                    test_name_by_col[test_col],
                    int(other_col),
                    test_name_by_col[other_col],
                ]
                if pearson is not None:
                    row.append(_to_float(pearson.get(other_col)))
                if spearman is not None:
                    row.append(_to_float(spearman.get(other_col)))
                sheet.append(row)
                out_rows += 1

        if out_rows == 0:
            wb.remove(sheet)
            continue

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{_excel_col_letter(sheet.max_column)}{sheet.max_row}"
        _autofit_openpyxl_columns(sheet)

    from datetime import datetime

    out_xlsx = output_folder / "Correlation_Report.xlsx"
    try:
        wb.save(out_xlsx)
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = output_folder / f"Correlation_Report_{ts}.xlsx"
        wb.save(alt)
        print(f"Could not overwrite (file open?): {out_xlsx}")
        print(f"Saved instead: {alt}")
        return alt

    print(f"Saved: {out_xlsx}")
    return out_xlsx


def run() -> int:
    """Run analysis using the USER PARAMETERS at the top of this file."""
    input_folder = _as_path(INPUT_FOLDER)
    output_folder = _as_path(OUTPUT_FOLDER)

    output_folder.mkdir(parents=True, exist_ok=True)

    generate_yield_cpk_report(
        input_folder=input_folder,
        output_folder=output_folder,
        modules=MODULES,
        outlier_mad_multiplier=OUTLIER_MAD_MULTIPLIER,
        yield_threshold=YIELD_THRESHOLD,
        cpk_low=CPK_LOW,
        cpk_high=CPK_HIGH,
        max_files=MAX_FILES,
        single_file=SINGLE_FILE,
        encoding=ENCODING,
    )

    if GENERATE_CORRELATION_REPORT:
        generate_correlation_workbook(
            input_folder=input_folder,
            output_folder=output_folder,
            modules=MODULES,
            max_files=MAX_FILES,
            single_file=SINGLE_FILE,
            methods=CORRELATION_METHODS,
            encoding=ENCODING,
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(run())
