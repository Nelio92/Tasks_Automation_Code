from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook


def main() -> None:
    xlsx = Path(__file__).resolve().parents[2] / "CTRX8188A_TE_TX.xlsx"
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb["52_DPLL"]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [h if isinstance(h, str) else "" for h in header]

    tn_idx = headers.index("Test Number")
    odi_idxs = [i for i, h in enumerate(headers) if isinstance(h, str) and h.strip().startswith("ODI")]

    print("ODI columns:")
    for i in odi_idxs:
        print(" ", i + 1, headers[i])

    count_nonzero = 0
    examples: list[tuple[int, str, object]] = []
    blank_streak = 0
    rows_scanned = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        rows_scanned += 1
        tn = row[tn_idx] if tn_idx < len(row) else None
        if tn in (None, ""):
            blank_streak += 1
            if blank_streak > 300:
                break
            continue
        blank_streak = 0

        for i in odi_idxs:
            v = row[i] if i < len(row) else None
            if v in (None, 0, "0", ""):
                continue
            try:
                fv = float(v)
                if abs(fv) <= 1e-12:
                    continue
            except Exception:
                pass

            count_nonzero += 1
            if len(examples) < 20:
                examples.append((int(float(tn)), headers[i], v))

    print("rows_scanned:", rows_scanned)
    print("nonzero cached ODI cells:", count_nonzero)
    print("examples:")
    for ex in examples:
        print(" ", ex)


if __name__ == "__main__":
    main()
