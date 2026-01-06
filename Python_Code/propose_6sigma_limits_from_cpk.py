from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


WORKBOOK_NAME = "FE_Test_CS_DoE_W2_W9_W15_All_Temps.xlsx"
TEST_NUMBERS_SHEET = "Test_Numbers"

CPK_GOOD_MIN = 1.67
CPK_GOOD_MAX = 4.0
SIGMA_MULTIPLIER = 6.0

# Rounding rules for proposed limits.
# - Always integers: LTL floored, UTL ceiled.
# - Prefer snapping to multiples of 5, else multiples of 2, but only if the
#   extra widening vs the exact 6σ limit is acceptable.
PREFERRED_STEPS: tuple[int, ...] = (5, 2)
ROUNDING_MAX_EXTRA_SIGMA = 0.2


@dataclass(frozen=True)
class Cols:
    test_nr: int
    cpk: int
    mean: int
    stddev: int
    low: int | None
    high: int | None
    min_: int | None
    max_: int | None


@dataclass(frozen=True)
class ProposalRow:
    sheet: str
    row: int
    group_key: tuple[str, float | None, float | None]
    mean: float
    stddev: float
    exact_ltl: float
    exact_utl: float
    low: float | None
    high: float | None
    min_: float | None
    max_: float | None


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value))


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if _is_number(value):
        return float(value)
    try:
        s = str(value).strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None


def _format_test_number(value: Any) -> str | None:
    """Normalize a test number to a stable string key."""
    if value is None:
        return None
    if _is_number(value):
        f = float(value)
        if f.is_integer():
            return str(int(f))
        return str(f)
    s = str(value).strip()
    return s or None


def _find_col(headers: list[str], want: Iterable[str]) -> int | None:
    want_set = {w.lower() for w in want}
    for idx, h in enumerate(headers, start=1):
        if h in want_set:
            return idx
    return None


def _find_cols(ws) -> Cols | None:
    headers = [_norm_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    test_nr = _find_col(headers, ["test nr", "test number", "test no", "test #"])
    cpk = _find_col(headers, ["cpk", "cpk value"])
    mean = _find_col(headers, ["mean", "average"])
    stddev = _find_col(headers, ["stddev", "std dev", "std", "stdev", "sigma"])

    if not (test_nr and cpk and mean and stddev):
        return None

    low = _find_col(headers, ["low", "ltl", "lsl"])
    high = _find_col(headers, ["high", "utl", "usl"])
    min_ = _find_col(headers, ["min"])
    max_ = _find_col(headers, ["max"])

    return Cols(
        test_nr=test_nr,
        cpk=cpk,
        mean=mean,
        stddev=stddev,
        low=low,
        high=high,
        min_=min_,
        max_=max_,
    )


def _floor_to_step(x: float, step: int) -> int:
    return int(math.floor(x / step) * step)


def _ceil_to_step(x: float, step: int) -> int:
    return int(math.ceil(x / step) * step)


def _choose_integer_limits(exact_ltl: float, exact_utl: float, stddev: float) -> tuple[int, int, str]:
    """Return (ltl, utl, rounding_mode).

    Always returns integers, using floor/ceil. Optionally snaps to a preferred step size
    if the extra widening relative to the exact value is acceptable.
    """

    base_ltl = int(math.floor(exact_ltl))
    base_utl = int(math.ceil(exact_utl))

    # Strict widening tolerance relative to the exact 6σ limit.
    # Note: this governs optional snapping to step sizes (5/2).
    allowance = ROUNDING_MAX_EXTRA_SIGMA * stddev

    for step in PREFERRED_STEPS:
        cand_ltl = _floor_to_step(exact_ltl, step)
        cand_utl = _ceil_to_step(exact_utl, step)
        delta_ltl = abs(cand_ltl - exact_ltl)
        delta_utl = abs(cand_utl - exact_utl)
        if delta_ltl <= allowance and delta_utl <= allowance:
            return cand_ltl, cand_utl, f"step={step}"

    return base_ltl, base_utl, "int"


def _ensure_headers(ws, headers: list[str]) -> list[int]:
    """Ensure headers exist; reuse existing cols if present, else append at end."""

    existing = {
        _norm_header(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)
    }

    col_indices: list[int] = []
    missing: list[str] = []
    for header in headers:
        col = existing.get(_norm_header(header))
        if col is None:
            missing.append(header)
        else:
            col_indices.append(col)

    if missing:
        start_col = ws.max_column + 1
        for offset, header in enumerate(missing):
            col = start_col + offset
            ws.cell(1, col).value = header
            col_indices.append(col)

    # Return in the same order as 'headers'
    lookup = {_norm_header(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    return [lookup[_norm_header(h)] for h in headers]


def _build_outlier_and_notes(
    *,
    proposed_ltl: float,
    proposed_utl: float,
    low: float | None,
    high: float | None,
    min_: float | None,
    max_: float | None,
) -> tuple[str, str]:
    risks: list[str] = []
    notes: list[str] = []

    if min_ is not None and min_ < proposed_ltl:
        risks.append("MIN < 6σ LTL")
    if max_ is not None and max_ > proposed_utl:
        risks.append("MAX > 6σ UTL")

    if min_ is not None and min_ < proposed_ltl:
        notes.append("Some data below 6σ LTL (possible outliers / heavy tail)")
    if max_ is not None and max_ > proposed_utl:
        notes.append("Some data above 6σ UTL (possible outliers / heavy tail)")

    # Context vs current spec limits (not necessarily a problem, but useful for review)
    if low is not None and proposed_ltl < low:
        notes.append("Proposed LTL is lower (wider) than current Low")
    if high is not None and proposed_utl > high:
        notes.append("Proposed UTL is higher (wider) than current High")

    # If the computed 6σ window is actually tighter than observed Min/Max, call it out.
    if min_ is not None and proposed_ltl > min_:
        risks.append("6σ LTL > MIN")
        notes.append("Computed 6σ LTL would exclude some observed data")
    if max_ is not None and proposed_utl < max_:
        risks.append("6σ UTL < MAX")
        notes.append("Computed 6σ UTL would exclude some observed data")

    risk_str = "; ".join(dict.fromkeys(risks)) if risks else ""
    note_str = "; ".join(dict.fromkeys(notes)) if notes else ""
    return risk_str, note_str


def main() -> int:
    repo_root = Path(__file__).resolve().parents[2]
    input_path = repo_root / WORKBOOK_NAME

    if not input_path.exists():
        raise FileNotFoundError(f"Workbook not found: {input_path}")

    # Load twice: one for numeric values (cached results), one for editing/preserving formulas.
    wb_values = load_workbook(input_path, data_only=True)
    wb_edit = load_workbook(input_path, data_only=False)

    if TEST_NUMBERS_SHEET not in wb_values.sheetnames:
        raise ValueError(f"Sheet '{TEST_NUMBERS_SHEET}' not found")

    ws_tests_val = wb_values[TEST_NUMBERS_SHEET]

    test_numbers: set[str] = set()
    for r in range(1, ws_tests_val.max_row + 1):
        key = _format_test_number(ws_tests_val.cell(r, 1).value)
        if key:
            test_numbers.add(key)

    if not test_numbers:
        raise ValueError("No test numbers found in Test_Numbers column A")

    highlight_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow
    highlight_font = Font(bold=True)

    # Two-pass approach:
    # 1) Collect all rows that need new limits.
    # 2) Uniformize within groups that share the same existing limits.
    proposals = 0
    touched_sheets: list[str] = []
    proposal_rows: list[ProposalRow] = []

    for sheet_name in wb_edit.sheetnames:
        if sheet_name == TEST_NUMBERS_SHEET:
            continue

        ws_edit = wb_edit[sheet_name]
        ws_val = wb_values[sheet_name]

        cols = _find_cols(ws_val)
        if cols is None:
            continue

        sheet_proposals = 0

        for r in range(2, ws_val.max_row + 1):
            test_key = _format_test_number(ws_val.cell(r, cols.test_nr).value)
            if not test_key or test_key not in test_numbers:
                continue

            cpk = _to_float(ws_val.cell(r, cols.cpk).value)
            mean = _to_float(ws_val.cell(r, cols.mean).value)
            stddev = _to_float(ws_val.cell(r, cols.stddev).value)

            low = _to_float(ws_val.cell(r, cols.low).value) if cols.low else None
            high = _to_float(ws_val.cell(r, cols.high).value) if cols.high else None
            min_ = _to_float(ws_val.cell(r, cols.min_).value) if cols.min_ else None
            max_ = _to_float(ws_val.cell(r, cols.max_).value) if cols.max_ else None

            is_good = cpk is not None and (CPK_GOOD_MIN <= cpk <= CPK_GOOD_MAX)

            if is_good:
                continue

            # If missing mean/stddev, we can't compute proposals.
            if mean is None or stddev is None or stddev <= 0:
                continue

            proposed_ltl = mean - SIGMA_MULTIPLIER * stddev
            proposed_utl = mean + SIGMA_MULTIPLIER * stddev

            group_key = (test_key, low, high)
            proposal_rows.append(
                ProposalRow(
                    sheet=sheet_name,
                    row=r,
                    group_key=group_key,
                    mean=mean,
                    stddev=stddev,
                    exact_ltl=proposed_ltl,
                    exact_utl=proposed_utl,
                    low=low,
                    high=high,
                    min_=min_,
                    max_=max_,
                )
            )

            sheet_proposals += 1
            proposals += 1

        if sheet_proposals:
            touched_sheets.append(sheet_name)

    # Build unified proposals per group_key.
    by_group: dict[tuple[str, float | None, float | None], list[ProposalRow]] = {}
    for pr in proposal_rows:
        by_group.setdefault(pr.group_key, []).append(pr)

    unified: dict[tuple[str, float | None, float | None], tuple[int, int, str]] = {}
    for key, rows in by_group.items():
        exact_ltl = min(r.exact_ltl for r in rows)
        exact_utl = max(r.exact_utl for r in rows)
        strict_stddev = min(r.stddev for r in rows)
        ltl_i, utl_i, mode = _choose_integer_limits(exact_ltl, exact_utl, strict_stddev)
        unified[key] = (ltl_i, utl_i, mode)

    # Write results + formatting.
    for sheet_name in wb_edit.sheetnames:
        if sheet_name == TEST_NUMBERS_SHEET:
            continue

        ws_edit = wb_edit[sheet_name]
        ws_val = wb_values[sheet_name]
        cols = _find_cols(ws_val)
        if cols is None:
            continue

        col_prop_ltl, col_prop_utl, col_flag, col_risk, col_notes, col_mode = _ensure_headers(
            ws_edit,
            [
                "Proposed LTL (6-sigma)",
                "Proposed UTL (6-sigma)",
                "New limits proposed",
                "Outlier risk",
                "Robustness notes",
                "Rounding mode",
            ],
        )

        for r in range(2, ws_val.max_row + 1):
            test_key = _format_test_number(ws_val.cell(r, cols.test_nr).value)
            if not test_key or test_key not in test_numbers:
                continue

            cpk = _to_float(ws_val.cell(r, cols.cpk).value)
            mean = _to_float(ws_val.cell(r, cols.mean).value)
            stddev = _to_float(ws_val.cell(r, cols.stddev).value)
            low = _to_float(ws_val.cell(r, cols.low).value) if cols.low else None
            high = _to_float(ws_val.cell(r, cols.high).value) if cols.high else None
            min_ = _to_float(ws_val.cell(r, cols.min_).value) if cols.min_ else None
            max_ = _to_float(ws_val.cell(r, cols.max_).value) if cols.max_ else None

            is_good = cpk is not None and (CPK_GOOD_MIN <= cpk <= CPK_GOOD_MAX)
            if is_good:
                ws_edit.cell(r, col_flag).value = "NO"
                continue

            if mean is None or stddev is None or stddev <= 0:
                ws_edit.cell(r, col_flag).value = "NO"
                ws_edit.cell(r, col_notes).value = "Missing/invalid Mean or Stddev; cannot compute 6σ limits"
                continue

            group_key = (test_key, low, high)
            ltl_i, utl_i, mode = unified.get(group_key, _choose_integer_limits(mean - 6 * stddev, mean + 6 * stddev, stddev))

            risk_str, note_str = _build_outlier_and_notes(
                proposed_ltl=float(ltl_i),
                proposed_utl=float(utl_i),
                low=low,
                high=high,
                min_=min_,
                max_=max_,
            )

            ws_edit.cell(r, col_prop_ltl).value = int(ltl_i)
            ws_edit.cell(r, col_prop_utl).value = int(utl_i)
            ws_edit.cell(r, col_flag).value = "YES"
            ws_edit.cell(r, col_risk).value = risk_str
            ws_edit.cell(r, col_notes).value = note_str
            ws_edit.cell(r, col_mode).value = mode

            # Highlight: Test Nr cell + the appended proposal columns.
            ws_edit.cell(r, cols.test_nr).fill = highlight_fill
            ws_edit.cell(r, cols.test_nr).font = highlight_font
            for c in [col_prop_ltl, col_prop_utl, col_flag, col_risk, col_notes, col_mode]:
                cell = ws_edit.cell(r, c)
                cell.fill = highlight_fill
                cell.font = highlight_font

    out_path = input_path.with_name(input_path.stem + "_with_proposed_limits.xlsx")
    wb_edit.save(out_path)

    print(f"Test numbers loaded: {len(test_numbers)}")
    print(f"Sheets updated: {len(touched_sheets)}")
    if touched_sheets:
        print(" - " + "\n - ".join(touched_sheets))
    print(f"Rows flagged with new limits: {proposals}")
    print(f"Saved: {out_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
