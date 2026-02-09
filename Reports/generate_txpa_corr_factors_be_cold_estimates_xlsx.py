"""Generate TXPA correlation factor cold (-40C) estimates as formatted XLSX.

Inputs:
  - TXPA_Corr_Factors_BE.xlsx (repo root)
Outputs (under Tasks_Automation_Code/Reports/):
  - txpa_corr_factors_be_cold_estimates.xlsx
  - txpa_corr_factors_be_cold_high_risk.xlsx
  - txpa_corr_factors_be_group_summary.xlsx

This script intentionally produces XLSX (not CSV) and auto-fits column widths.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
REPORTS_DIR = REPO_ROOT / "Tasks_Automation_Code" / "Reports"
INPUT_XLSX = REPO_ROOT / "TXPA_Corr_Factors_BE.xlsx"


def _autofit_worksheet_columns(ws, max_width: int = 70) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, float) and np.isnan(v):
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        width = min(max_len + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _format_header_row(ws) -> None:
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_df_to_xlsx(df: pd.DataFrame, out_path: Path, sheet_name: str) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        _format_header_row(ws)
        _autofit_worksheet_columns(ws)


def _compute_tables() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = pd.read_excel(INPUT_XLSX, sheet_name=0)

    col_135 = [c for c in df.columns if "135" in str(c)][0]
    col_25 = [c for c in df.columns if "25" in str(c)][0]

    f25_s = pd.to_numeric(df[col_25], errors="coerce")
    f135_s = pd.to_numeric(df[col_135], errors="coerce")
    valid = f25_s.notna() & f135_s.notna()

    df = df.loc[valid].copy()
    f25 = f25_s.loc[valid].to_numpy(dtype=float)
    f135 = f135_s.loc[valid].to_numpy(dtype=float)

    # Temperatures
    T1, T2, T3 = 25.0, 135.0, -40.0

    def K(Tc: float) -> float:
        return Tc + 273.15

    # Model A: linear in T (Celsius)
    slope_T = (f135 - f25) / (T2 - T1)
    f3_T = f25 + slope_T * (T3 - T1)

    # Model B: linear in invK
    x1, x2, x3 = 1.0 / K(T1), 1.0 / K(T2), 1.0 / K(T3)
    a = (f135 - f25) / (x2 - x1)
    b = f25 - a * x1
    f3_invK = a * x3 + b

    # Model C: linear in logK
    y1, y2, y3 = np.log(K(T1)), np.log(K(T2)), np.log(K(T3))
    a2 = (f135 - f25) / (y2 - y1)
    b2 = f25 - a2 * y1
    f3_logK = a2 * y3 + b2

    cold_models = np.vstack([f3_T, f3_invK, f3_logK])
    f3_min = cold_models.min(axis=0)
    f3_max = cold_models.max(axis=0)
    f3_med = np.median(cold_models, axis=0)
    f3_mean = cold_models.mean(axis=0)
    env_width = f3_max - f3_min

    abs_delta = np.abs(f135 - f25)
    rel_delta = abs_delta / np.where(f25 == 0.0, np.nan, np.abs(f25))

    sens_class = np.where(abs_delta <= 0.05, "low", np.where(abs_delta <= 0.2, "medium", "high"))
    unc_class = np.where(env_width <= 0.1, "low", np.where(env_width <= 0.2, "medium", "high"))

    # Recommendation policy
    recommended = np.where(abs_delta <= 0.05, f25, f3_med)
    rec_method = np.where(abs_delta <= 0.05, "use_f25_low_sensitivity", "ensemble_median_3models")

    base_cols = [
        "DataSheet",
        "LUT value",
        "Test Name",
        "Voltage corner",
        "Frequency_GHz",
        "PA Channel",
        "N",
    ]
    out = df[base_cols].copy()

    out["f_25"] = f25
    out["f_135"] = f135
    out["delta_135_minus_25"] = f135 - f25
    out["abs_delta_135_25"] = abs_delta
    out["rel_change_abs_135_25"] = rel_delta

    out["f_-40_linearT"] = f3_T
    out["f_-40_linearInvK"] = f3_invK
    out["f_-40_linearLogK"] = f3_logK

    out["f_-40_models_min"] = f3_min
    out["f_-40_models_max"] = f3_max
    out["f_-40_models_mean"] = f3_mean
    out["f_-40_models_median"] = f3_med
    out["f_-40_models_envelope_width"] = env_width

    out["temp_sensitivity_class"] = sens_class
    out["cold_model_uncertainty_class"] = unc_class
    out["f_-40_recommended"] = recommended
    out["f_-40_recommended_method"] = rec_method

    out["flag_high_sensitivity_abs_delta_gt_0p2"] = abs_delta > 0.2
    out["flag_high_model_uncertainty_env_gt_0p2"] = env_width > 0.2
    out["flag_sign_change_25_to_135"] = (np.sign(f25) != np.sign(f135)) & (f25 != 0.0) & (f135 != 0.0)

    risk = out.loc[out["flag_high_sensitivity_abs_delta_gt_0p2"] | out["flag_high_model_uncertainty_env_gt_0p2"]].copy()
    risk = risk.sort_values(
        [
            "flag_high_model_uncertainty_env_gt_0p2",
            "flag_high_sensitivity_abs_delta_gt_0p2",
            "f_-40_models_envelope_width",
            "abs_delta_135_25",
        ],
        ascending=[False, False, False, False],
    )

    # Group summary (by Voltage corner, Frequency_GHz, PA Channel)
    keys = ["Voltage corner", "Frequency_GHz", "PA Channel"]
    summary = (
        out.groupby(keys, dropna=False)
        .agg(
            rows=("f_25", "size"),
            f25_median=("f_25", "median"),
            f135_median=("f_135", "median"),
            delta_median=("delta_135_minus_25", "median"),
            abs_delta_median=("abs_delta_135_25", "median"),
            env_width_median=("f_-40_models_envelope_width", "median"),
            env_width_p90=("f_-40_models_envelope_width", lambda s: s.quantile(0.90)),
        )
        .reset_index()
    )

    return out, risk, summary


def main() -> int:
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"Missing input workbook: {INPUT_XLSX}")

    out, risk, summary = _compute_tables()

    _write_df_to_xlsx(out, REPORTS_DIR / "txpa_corr_factors_be_cold_estimates.xlsx", "cold_estimates")
    _write_df_to_xlsx(risk, REPORTS_DIR / "txpa_corr_factors_be_cold_high_risk.xlsx", "high_risk")
    _write_df_to_xlsx(summary, REPORTS_DIR / "txpa_corr_factors_be_group_summary.xlsx", "group_summary")

    print("Wrote:")
    print(" -", REPORTS_DIR / "txpa_corr_factors_be_cold_estimates.xlsx")
    print(" -", REPORTS_DIR / "txpa_corr_factors_be_cold_high_risk.xlsx")
    print(" -", REPORTS_DIR / "txpa_corr_factors_be_group_summary.xlsx")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
